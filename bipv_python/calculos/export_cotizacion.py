"""
export_cotizacion.py
====================
Genera una COTIZACIÓN presentable para el cliente final del sistema BIPV,
en formato Excel (openpyxl) y PDF (fpdf2, la misma librería del Reporte técnico).

Funciones puras: reciben un dict `datos` armado por la página del Presupuesto
(no leen st.session_state) y devuelven los bytes del archivo.

Estructura esperada de `datos`
------------------------------
{
    "empresa":         str,      # nombre de la empresa que cotiza
    "proyecto":        str,      # nombre del proyecto
    "cliente":         str,      # nombre del cliente/destinatario
    "fecha":           str,      # fecha de emisión "dd/mm/aaaa"
    "validez_dias":    int,      # validez de la oferta (default 15)
    "trm":             float,    # COP/USD (0 o None si no aplica)
    "items": [                   # SOLO ítems activos
        {"categoria": str, "descripcion": str, "cantidad": float,
         "unidad": str, "unitario_cop": float, "total_cop": float},
        ...
    ],
    "subtotal_cop":       float, # suma de ítems (CAPEX directo + catálogo)
    "costos_blandos_cop": float, # 0 si no aplica
    "indirectos_cop":     float, # 0 si no aplica (AUI / administración / utilidad)
    "contingencia_cop":   float, # 0 si no aplica
    "total_cop":          float, # TOTAL final en COP
    "total_usd":          float, # TOTAL en USD (0 si no hay TRM)
    "notas":              str,   # condiciones/notas al pie (texto libre)
}

Dependencias: openpyxl, fpdf2 >= 2.7
"""
from __future__ import annotations

import io
from typing import Any

# ── Notas / condiciones por defecto ───────────────────────────────────────────
def _txt_xlsx(s) -> str:
    """Neutraliza inyección de fórmulas en Excel (OWASP CSV/Formula injection).

    Cualquier texto controlado por el usuario que empiece por '=', '+', '-',
    '@', TAB o CR se prefija con apóstrofo para que Excel lo trate como texto
    literal y nunca lo ejecute como fórmula.
    """
    s = str(s or "")
    return "'" + s if s[:1] in ("=", "+", "-", "@", "\t", "\r") else s


NOTAS_DEFAULT = (
    "• Los valores están expresados en pesos colombianos (COP) e incluyen los "
    "conceptos detallados en esta cotización.\n"
    "• Precios sujetos a variación de la TRM (tasa de cambio) y a la disponibilidad "
    "de los equipos importados al momento de la orden de compra.\n"
    "• No incluye obras civiles adicionales, adecuaciones eléctricas del predio ni "
    "trámites no mencionados salvo indicación expresa.\n"
    "• Tiempo de ejecución y forma de pago se acuerdan en el contrato final.\n"
    "• Esta cotización no reemplaza una ingeniería de detalle certificada."
)

VALIDEZ_DEFAULT_DIAS = 15


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS DE FORMATO
# ══════════════════════════════════════════════════════════════════════════════

def formato_cop(valor: float) -> str:
    """Formato colombiano: separador de miles con punto → '$ 12.345.678'."""
    try:
        v = float(valor)
    except (TypeError, ValueError):
        return "$ 0"
    # Formato estándar en-US (coma miles) y luego swap a punto.
    entero = f"{round(v):,.0f}".replace(",", ".")
    return f"$ {entero}"


def formato_usd(valor: float) -> str:
    """Formato USD con separador de miles con coma → 'USD 12,345'."""
    try:
        v = float(valor)
    except (TypeError, ValueError):
        return "USD 0"
    return f"USD {v:,.0f}"


def _num(valor: Any, default: float = 0.0) -> float:
    try:
        return float(valor)
    except (TypeError, ValueError):
        return default


def _latin1(texto: str) -> str:
    """Sanitiza texto para las fuentes core de fpdf2 (Helvetica = latin-1)."""
    s = str(texto)
    reemplazos = {
        "•": "-", "–": "-", "—": "-", "’": "'", "‘": "'",
        "“": '"', "”": '"', "…": "...", "→": "->", "≈": "~",
        "²": "2", "³": "3", "·": "-", "✔": "OK", "€": "EUR",
        "™": "", "®": "", "©": "",
    }
    for k, v in reemplazos.items():
        s = s.replace(k, v)
    return s.encode("latin-1", "replace").decode("latin-1")


def _validar_datos(datos: dict) -> list[dict]:
    """Valida y normaliza. Lanza ValueError si no hay ítems activos."""
    if not isinstance(datos, dict):
        raise ValueError("Los datos de la cotización deben ser un diccionario.")
    items = datos.get("items") or []
    items_norm = []
    for it in items:
        cant = _num(it.get("cantidad"))
        total = _num(it.get("total_cop"))
        # Un ítem cuenta si tiene descripción y algún valor > 0.
        if str(it.get("descripcion", "")).strip() and (total > 0 or cant > 0):
            items_norm.append({
                "categoria":    str(it.get("categoria", "General")).strip() or "General",
                "descripcion":  str(it.get("descripcion", "")).strip(),
                "ref":          str(it.get("ref", "") or "").strip(),
                "cantidad":     cant,
                "unidad":       str(it.get("unidad", "")).strip(),
                "unitario_usd": _num(it.get("unitario_usd")),
                "total_usd_item": _num(it.get("total_usd")),
                "unitario_cop": _num(it.get("unitario_cop")),
                "total_cop":    total,
            })
    if not items_norm:
        raise ValueError(
            "No hay ítems activos en el presupuesto. Marca al menos un ítem como "
            "activo (✔) y con valor mayor que cero antes de generar la cotización."
        )
    return items_norm


def _agrupar_por_categoria(items: list[dict]) -> dict[str, list[dict]]:
    grupos: dict[str, list[dict]] = {}
    for it in items:
        grupos.setdefault(it["categoria"], []).append(it)
    return grupos


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def generar_cotizacion_excel(datos: dict) -> bytes:
    """Genera la cotización en Excel (.xlsx) y devuelve los bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    items = _validar_datos(datos)
    grupos = _agrupar_por_categoria(items)

    empresa   = str(datos.get("empresa", "") or "")
    proyecto  = str(datos.get("proyecto", "Proyecto BIPV") or "Proyecto BIPV")
    cliente   = str(datos.get("cliente", "") or "")
    fecha     = str(datos.get("fecha", "") or "")
    validez   = int(_num(datos.get("validez_dias", VALIDEZ_DEFAULT_DIAS), VALIDEZ_DEFAULT_DIAS))
    trm       = _num(datos.get("trm"))
    subtotal  = _num(datos.get("subtotal_cop"))
    blandos   = _num(datos.get("costos_blandos_cop"))
    indirect  = _num(datos.get("indirectos_cop"))
    conting   = _num(datos.get("contingencia_cop"))
    total_cop = _num(datos.get("total_cop"))
    total_usd = _num(datos.get("total_usd"))
    notas     = str(datos.get("notas") or NOTAS_DEFAULT)

    # Seguridad: texto libre del usuario nunca debe interpretarse como fórmula
    empresa, proyecto, cliente = _txt_xlsx(empresa), _txt_xlsx(proyecto), _txt_xlsx(cliente)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cotización"

    # Estilos
    AZUL      = "1A5276"
    AZUL_CLARO= "D6EAF8"
    GRIS      = "F4F6F7"
    f_titulo  = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    f_sub     = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    f_hdr     = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    f_cat     = Font(name="Calibri", size=10, bold=True, color=AZUL)
    f_bold    = Font(name="Calibri", size=10, bold=True)
    f_norm    = Font(name="Calibri", size=10)
    f_nota    = Font(name="Calibri", size=8, italic=True, color="666666")
    fill_azul = PatternFill("solid", fgColor=AZUL)
    fill_hdr  = PatternFill("solid", fgColor=AZUL)
    fill_cat  = PatternFill("solid", fgColor=AZUL_CLARO)
    fill_tot  = PatternFill("solid", fgColor=AZUL_CLARO)
    right     = Alignment(horizontal="right")
    center    = Alignment(horizontal="center")
    left_wrap = Alignment(horizontal="left", wrap_text=True, vertical="top")
    borde     = Border(bottom=Side(style="thin", color="D5D8DC"))

    # Determinar si algún ítem tiene precios en USD (para mostrar columnas USD)
    _tiene_usd = any(_num(it.get("unitario_usd")) > 0 for it in items)

    if _tiene_usd and trm > 0:
        # 8 columnas: Descripción, Ref, Cantidad, Unidad, USD/un, Total USD, $ unitario (COP), Total COP
        ancho = 8
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 9
        ws.column_dimensions["D"].width = 7
        ws.column_dimensions["E"].width = 13
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 16
        ws.column_dimensions["H"].width = 16
    else:
        # 6 columnas: Descripción, Ref, Cantidad, Unidad, $ unitario (COP), Total COP
        ancho = 6
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 9
        ws.column_dimensions["D"].width = 7
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18

    r = 1
    # ── Encabezado ────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ancho)
    c = ws.cell(r, 1, "COTIZACIÓN — SISTEMA SOLAR BIPV")
    c.font = f_titulo; c.fill = fill_azul; c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 26
    r += 2

    def _kv(label, valor):
        nonlocal r
        a = ws.cell(r, 1, label); a.font = f_bold
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=ancho)
        b = ws.cell(r, 2, valor); b.font = f_norm
        r += 1

    if empresa:
        _kv("Empresa:", empresa)
    _kv("Proyecto:", proyecto)
    if cliente:
        _kv("Cliente:", cliente)
    if fecha:
        _kv("Fecha de emisión:", fecha)
    _kv("Validez de la oferta:", f"{validez} días")
    if trm > 0:
        _kv("TRM de referencia:", f"$ {round(trm):,.0f}".replace(",", ".") + " COP/USD")
    r += 1

    # ── Cabecera de tabla ───────────────────────────────────────────────────────
    if _tiene_usd and trm > 0:
        headers = ["Descripción", "Ref.", "Cantidad", "Unidad",
                   "USD/un", "Total USD", "$ unitario (COP)", "Total COP"]
    else:
        headers = ["Descripción", "Ref.", "Cantidad", "Unidad",
                   "$ unitario (COP)", "Total COP"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(r, j, h)
        c.font = f_hdr; c.fill = fill_hdr
        c.alignment = center if j > 1 else Alignment(horizontal="left")
    r += 1

    # ── Ítems por categoría ─────────────────────────────────────────────────────
    for categoria, filas in grupos.items():
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ancho)
        c = ws.cell(r, 1, _txt_xlsx(categoria))
        c.font = f_cat; c.fill = fill_cat
        r += 1
        for it in filas:
            ws.cell(r, 1, _txt_xlsx(it["descripcion"])).font = f_norm
            ws.cell(r, 1).alignment = left_wrap
            cr = ws.cell(r, 2, _txt_xlsx(it.get("ref", ""))); cr.font = f_norm; cr.alignment = center
            cn = ws.cell(r, 3, round(it["cantidad"], 2)); cn.font = f_norm; cn.alignment = right
            cu = ws.cell(r, 4, _txt_xlsx(it["unidad"])); cu.font = f_norm; cu.alignment = center
            if _tiene_usd and trm > 0:
                c_uusd = ws.cell(r, 5, round(it.get("unitario_usd", 0), 2))
                c_uusd.font = f_norm; c_uusd.number_format = '"USD" #,##0.00'; c_uusd.alignment = right
                c_tusd = ws.cell(r, 6, round(it.get("total_usd_item", 0), 2))
                c_tusd.font = f_norm; c_tusd.number_format = '"USD" #,##0.00'; c_tusd.alignment = right
                cvu = ws.cell(r, 7, round(it["unitario_cop"])); cvu.font = f_norm
                cvu.number_format = '"$" #,##0'; cvu.alignment = right
                cvt = ws.cell(r, 8, round(it["total_cop"])); cvt.font = f_norm
                cvt.number_format = '"$" #,##0'; cvt.alignment = right
            else:
                cvu = ws.cell(r, 5, round(it["unitario_cop"])); cvu.font = f_norm
                cvu.number_format = '"$" #,##0'; cvu.alignment = right
                cvt = ws.cell(r, 6, round(it["total_cop"])); cvt.font = f_norm
                cvt.number_format = '"$" #,##0'; cvt.alignment = right
            for col in range(1, ancho + 1):
                ws.cell(r, col).border = borde
            r += 1

    r += 1

    # ── Totales ──────────────────────────────────────────────────────────────
    _col_total_cop = ancho  # última columna = Total COP
    _col_total_usd = ancho - 2 if (_tiene_usd and trm > 0) else None

    def _fila_total(label, valor_cop, valor_usd=None, negrita=False, resaltar=False):
        nonlocal r
        # El label solo puede fusionar hasta ANTES de la columna de total USD
        # (si existe) -- fusionar hasta _col_total_cop - 1 la incluiría (es
        # menor que _col_total_cop), y escribir el valor USD en una celda
        # fusionada no-ancla lanza 'MergedCell' object attribute 'value' is
        # read-only. Encontrado en producción: cualquier cotización con al
        # menos un ítem con precio en USD (p.ej. equipo importado) y TRM > 0
        # disparaba esto al generar el Excel.
        _col_label_fin = (_col_total_usd - 1) if _col_total_usd else (_col_total_cop - 1)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=_col_label_fin)
        a = ws.cell(r, 1, label)
        a.font = f_bold if (negrita or resaltar) else f_norm
        a.alignment = right
        v = ws.cell(r, _col_total_cop, round(valor_cop))
        v.number_format = '"$" #,##0'
        v.font = f_bold if (negrita or resaltar) else f_norm
        v.alignment = right
        if _col_total_usd and valor_usd is not None:
            vu = ws.cell(r, _col_total_usd, round(valor_usd, 2))
            vu.number_format = '"USD" #,##0.00'
            vu.font = f_bold if (negrita or resaltar) else f_norm
            vu.alignment = right
        if resaltar:
            for col in range(1, ancho + 1):
                ws.cell(r, col).fill = fill_tot
        r += 1

    _sub_usd = total_usd * (subtotal / total_cop) if total_cop > 0 else 0
    _fila_total("Subtotal", subtotal, _sub_usd if trm > 0 else None)
    if blandos > 0:
        _bl_usd = blandos / trm if trm > 0 else 0
        _fila_total("Costos blandos (ingeniería, trámites, gestión)", blandos,
                    _bl_usd if trm > 0 else None)
    if indirect > 0:
        _ind_usd = indirect / trm if trm > 0 else 0
        _fila_total("Costos indirectos (administración y utilidad)", indirect,
                    _ind_usd if trm > 0 else None)
    if conting > 0:
        _cont_usd = conting / trm if trm > 0 else 0
        _fila_total("Contingencia", conting, _cont_usd if trm > 0 else None)
    _fila_total("TOTAL", total_cop, total_usd if trm > 0 else None, resaltar=True)

    if trm > 0 and total_usd > 0 and not (_tiene_usd and trm > 0):
        # Solo si las columnas USD no están visibles, mostrar total USD como fila aparte
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=_col_total_cop - 1)
        a = ws.cell(r, 1, "Equivalente aproximado (USD)"); a.font = f_norm; a.alignment = right
        v = ws.cell(r, _col_total_cop, round(total_usd)); v.number_format = '"USD" #,##0'
        v.font = f_norm; v.alignment = right
        r += 1

    r += 2

    # ── Notas / condiciones ────────────────────────────────────────────────────
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ancho)
    c = ws.cell(r, 1, "Notas y condiciones"); c.font = f_sub; c.fill = fill_azul
    r += 1
    for linea in notas.split("\n"):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ancho)
        c = ws.cell(r, 1, _txt_xlsx(linea)); c.font = f_nota; c.alignment = left_wrap
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# PDF (fpdf2 — misma librería del Reporte técnico de la página 10 / utils)
# ══════════════════════════════════════════════════════════════════════════════

def generar_cotizacion_pdf(datos: dict) -> bytes:
    """Genera la cotización en PDF y devuelve los bytes."""
    from fpdf import FPDF, XPos, YPos

    items = _validar_datos(datos)
    grupos = _agrupar_por_categoria(items)

    empresa   = str(datos.get("empresa", "") or "")
    proyecto  = str(datos.get("proyecto", "Proyecto BIPV") or "Proyecto BIPV")
    cliente   = str(datos.get("cliente", "") or "")
    fecha     = str(datos.get("fecha", "") or "")
    validez   = int(_num(datos.get("validez_dias", VALIDEZ_DEFAULT_DIAS), VALIDEZ_DEFAULT_DIAS))
    trm       = _num(datos.get("trm"))
    subtotal  = _num(datos.get("subtotal_cop"))
    blandos   = _num(datos.get("costos_blandos_cop"))
    indirect  = _num(datos.get("indirectos_cop"))
    conting   = _num(datos.get("contingencia_cop"))
    total_cop = _num(datos.get("total_cop"))
    total_usd = _num(datos.get("total_usd"))
    notas     = str(datos.get("notas") or NOTAS_DEFAULT)

    AZUL   = (26, 82, 118)
    AZUL_C = (214, 234, 248)
    GRIS   = (100, 100, 100)
    NEGRO  = (0, 0, 0)
    BLANCO = (255, 255, 255)

    class _Cotizacion(FPDF):
        def __init__(self):
            super().__init__(orientation="P", unit="mm", format="A4")
            self.set_auto_page_break(auto=True, margin=18)
            self.set_margins(left=15, top=16, right=15)

        def header(self):
            self.set_fill_color(*AZUL)
            self.rect(0, 0, 210, 12, style="F")
            self.set_font("Helvetica", "B", 10)
            self.set_text_color(*BLANCO)
            self.set_xy(15, 3)
            self.cell(0, 6, _latin1(empresa or "Cotización BIPV"),
                      new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            self.set_text_color(*NEGRO)
            self.set_y(16)

        def footer(self):
            self.set_y(-14)
            self.set_font("Helvetica", "I", 7)
            self.set_text_color(*GRIS)
            self.cell(0, 5, _latin1(
                "Cotización generada por la Calculadora BIPV. Valores estimativos; "
                "no reemplazan una ingeniería de detalle certificada."),
                align="C", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            self.cell(0, 4, f"Pagina {self.page_no()}", align="C",
                      new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            self.set_text_color(*NEGRO)

    pdf = _Cotizacion()
    pdf.add_page()

    # ── Título ──────────────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 15)
    pdf.set_text_color(*AZUL)
    pdf.cell(0, 9, _latin1("COTIZACION - SISTEMA SOLAR BIPV"),
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(*NEGRO)
    pdf.ln(2)

    # ── Datos del proyecto/cliente ────────────────────────────────────────────
    def _kv(label, valor):
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(45, 6, _latin1(label), new_x=XPos.RIGHT, new_y=YPos.TOP)
        pdf.set_font("Helvetica", "", 9)
        pdf.multi_cell(0, 6, _latin1(valor), new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    _kv("Proyecto:", proyecto)
    if cliente:
        _kv("Cliente:", cliente)
    if fecha:
        _kv("Fecha de emision:", fecha)
    _kv("Validez de la oferta:", f"{validez} dias")
    if trm > 0:
        _kv("TRM de referencia:", f"$ {round(trm):,.0f}".replace(",", ".") + " COP/USD")
    pdf.ln(3)

    # Anchos de columnas de la tabla (suman 180 mm útiles)
    W_DESC, W_CANT, W_UNI, W_VU, W_VT = 82, 20, 16, 31, 31

    def _cabecera_tabla():
        pdf.set_fill_color(*AZUL)
        pdf.set_text_color(*BLANCO)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(W_DESC, 7, " Descripcion", fill=True, new_x=XPos.RIGHT, new_y=YPos.TOP)
        pdf.cell(W_CANT, 7, "Cantidad", fill=True, align="C", new_x=XPos.RIGHT, new_y=YPos.TOP)
        pdf.cell(W_UNI, 7, "Unidad", fill=True, align="C", new_x=XPos.RIGHT, new_y=YPos.TOP)
        pdf.cell(W_VU, 7, "Unitario (COP)", fill=True, align="C", new_x=XPos.RIGHT, new_y=YPos.TOP)
        pdf.cell(W_VT, 7, "Total (COP)", fill=True, align="C", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_text_color(*NEGRO)

    _cabecera_tabla()

    for categoria, filas in grupos.items():
        # Fila de categoría
        pdf.set_fill_color(*AZUL_C)
        pdf.set_font("Helvetica", "B", 8.5)
        pdf.set_text_color(*AZUL)
        pdf.cell(W_DESC + W_CANT + W_UNI + W_VU + W_VT, 6, _latin1(" " + categoria),
                 fill=True, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_text_color(*NEGRO)
        for it in filas:
            pdf.set_font("Helvetica", "", 8)
            # Descripción puede envolver: calcular alto con multi_cell simulada.
            x0, y0 = pdf.get_x(), pdf.get_y()
            desc = _latin1(it["descripcion"])
            pdf.multi_cell(W_DESC, 5, " " + desc, border=0,
                           new_x=XPos.RIGHT, new_y=YPos.TOP, max_line_height=5)
            alto = max(5.0, pdf.get_y() - y0)
            # Reposicionar para las columnas numéricas a la misma altura de la fila.
            pdf.set_xy(x0 + W_DESC, y0)
            pdf.cell(W_CANT, alto, f"{it['cantidad']:,.2f}", align="C",
                     new_x=XPos.RIGHT, new_y=YPos.TOP)
            pdf.cell(W_UNI, alto, _latin1(it["unidad"]), align="C",
                     new_x=XPos.RIGHT, new_y=YPos.TOP)
            pdf.cell(W_VU, alto, formato_cop(it["unitario_cop"]), align="R",
                     new_x=XPos.RIGHT, new_y=YPos.TOP)
            pdf.cell(W_VT, alto, formato_cop(it["total_cop"]), align="R",
                     new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            # Línea separadora
            pdf.set_draw_color(213, 216, 220)
            y_lin = pdf.get_y()
            pdf.line(15, y_lin, 195, y_lin)

    pdf.ln(3)

    # ── Totales ────────────────────────────────────────────────────────────────
    W_LBL = W_DESC + W_CANT + W_UNI
    W_VAL = W_VU + W_VT

    def _fila_total(label, valor, resaltar=False):
        if resaltar:
            pdf.set_fill_color(*AZUL_C)
            pdf.set_font("Helvetica", "B", 10)
        else:
            pdf.set_font("Helvetica", "", 9)
        pdf.cell(W_LBL, 7, _latin1(label + "  "), align="R", fill=resaltar,
                 new_x=XPos.RIGHT, new_y=YPos.TOP)
        pdf.cell(W_VAL, 7, formato_cop(valor), align="R", fill=resaltar,
                 new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    _fila_total("Subtotal", subtotal)
    if blandos > 0:
        _fila_total("Costos blandos (ingenieria, tramites, gestion)", blandos)
    if indirect > 0:
        _fila_total("Costos indirectos (administracion y utilidad)", indirect)
    if conting > 0:
        _fila_total("Contingencia", conting)
    _fila_total("TOTAL (COP)", total_cop, resaltar=True)

    if trm > 0 and total_usd > 0:
        pdf.set_font("Helvetica", "", 8.5)
        pdf.set_text_color(*GRIS)
        pdf.cell(W_LBL, 6, _latin1("Equivalente aproximado (USD)  "), align="R",
                 new_x=XPos.RIGHT, new_y=YPos.TOP)
        pdf.cell(W_VAL, 6, f"USD {total_usd:,.0f}", align="R",
                 new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_text_color(*NEGRO)

    pdf.ln(5)

    # ── Notas / condiciones ────────────────────────────────────────────────────
    pdf.set_fill_color(*AZUL)
    pdf.set_text_color(*BLANCO)
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(0, 7, _latin1(" Notas y condiciones"), fill=True,
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(*NEGRO)
    pdf.ln(1)
    pdf.set_font("Helvetica", "", 8)
    for linea in notas.split("\n"):
        pdf.multi_cell(0, 5, _latin1(linea), new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf_bytes = pdf.output()
    return bytes(pdf_bytes)


def nombre_archivo_cotizacion(proyecto: str, fecha_iso: str, ext: str) -> str:
    """Nombre sugerido: Cotizacion_<proyecto>_<fecha>.<ext>."""
    nombre = _latin1(str(proyecto or "BIPV")).replace(" ", "_")
    nombre = "".join(ch for ch in nombre if ch.isalnum() or ch in ("_", "-")) or "BIPV"
    return f"Cotizacion_{nombre}_{fecha_iso}.{ext}"
