# -*- coding: utf-8 -*-
"""Regresión: generar_cotizacion_excel() lanzaba
`AttributeError: 'MergedCell' object attribute 'value' is read-only`
para cualquier cotización con al menos un ítem con precio en USD (p.ej.
un equipo importado) y TRM > 0 -- reportado por un usuario real intentando
descargar la cotización de un proyecto con estructura importada.

Causa raíz: _fila_total() fusionaba las columnas 1..(_col_total_cop-1)
para el label, pero esa fusión INCLUÍA la columna _col_total_usd (donde
la misma función escribe el total en USD) cuando había columnas USD
visibles -- escribir .value en una celda fusionada no-ancla revienta con
ese AttributeError, no con un ValueError controlado.
"""
import io

import pytest
from openpyxl import load_workbook

from calculos.export_cotizacion import generar_cotizacion_excel


def _datos_con_item_usd(**overrides):
    base = {
        "empresa": "Estrategias Químicas", "proyecto": "Granja FV Test",
        "cliente": "Cliente", "fecha": "20/08/2026", "validez_dias": 15,
        "trm": 3100.0,
        "items": [
            {"categoria": "Estructura", "descripcion": "Estructura Mibet",
             "ref": "EST-MIBET", "cantidad": 1, "unidad": "glb",
             "unitario_usd": 57796.0, "total_usd": 57796.0,
             "unitario_cop": 57796.0 * 3100, "total_cop": 57796.0 * 3100},
        ],
        "subtotal_cop": 57796.0 * 3100, "costos_blandos_cop": 0.0,
        "indirectos_cop": 0.0, "contingencia_cop": 0.0,
        "total_cop": 57796.0 * 3100, "total_usd": 57796.0,
        "notas": "test",
    }
    base.update(overrides)
    return base


def test_cotizacion_con_item_usd_y_trm_no_revienta_con_mergedcell():
    # Antes de la regresión: AttributeError: 'MergedCell' object attribute
    # 'value' is read-only. Ahora debe generar bytes válidos sin excepción.
    b = generar_cotizacion_excel(_datos_con_item_usd())
    assert isinstance(b, bytes)
    assert len(b) > 0


def test_totales_usd_y_cop_caen_en_las_columnas_correctas():
    datos = _datos_con_item_usd()
    b = generar_cotizacion_excel(datos)
    wb = load_workbook(io.BytesIO(b))
    ws = wb.active

    filas = {
        str(row[0].value): row
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row)
        if row[0].value in ("Subtotal", "TOTAL")
    }
    assert "Subtotal" in filas and "TOTAL" in filas
    for etiqueta, row in filas.items():
        # Columnas: 1 Descripcion .. 5 USD/un, 6 Total USD, 7 $/un COP, 8 Total COP
        assert row[5].value == pytest.approx(57796.0), etiqueta       # Total USD
        assert row[7].value == pytest.approx(57796.0 * 3100), etiqueta  # Total COP


def test_cotizacion_solo_cop_sin_trm_sigue_funcionando():
    # Camino sin columnas USD (ancho=6) -- no debe regresionar por el fix.
    datos = _datos_con_item_usd(trm=0.0)
    datos["items"][0]["unitario_usd"] = 0.0
    datos["items"][0]["total_usd"] = 0.0
    datos["total_usd"] = 0.0
    b = generar_cotizacion_excel(datos)
    wb = load_workbook(io.BytesIO(b))
    ws = wb.active
    fila_total = next(
        row for row in ws.iter_rows(min_row=1, max_row=ws.max_row)
        if row[0].value == "TOTAL"
    )
    assert fila_total[5].value == pytest.approx(57796.0 * 3100)  # col 6 = Total COP (ancho=6)
