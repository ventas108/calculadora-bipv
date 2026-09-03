# -*- coding: utf-8 -*-
"""
Script de uso unico: agrega paneles reales de First Solar (CdTe) al
catalogo real (datos/paneles_catalogo.xlsx::Catalogo_Paneles_FV) -- ver
DIAGNOSTICO_CATALOGO_FIRSTSOLAR_NREL.md para el detalle completo.

CRITERIO DISTINTO al de los 5 fabricantes cristalinos ya importados (JA
Solar, Trina, Jinko, Canadian Solar, LONGi):

1. NO se excluye por tolerancia SDM (6%, validar_sdm_vs_ficha). Ver
   DIAGNOSTICO_JRC_HULD_PRIMARIO_CDTE.md: calculos/produccion.py YA usa
   JRC/Huld como motor de ENERGIA para CdTe, no SDM (migrado el 2-sep-2026
   por un defecto estructural real del SDM para esta tecnologia -- la
   "joroba" de eficiencia >100%). Excluir por un chequeo de ajuste SDM
   descartaria paneles CdTe reales buenos por una metrica irrelevante para
   su uso principal (produccion de energia). Se calcula igual para
   transparencia en Notas (Motor IV/Mismatch/MPPT combinado SI siguen en
   SDM, con la limitacion estructural ya documentada, sin exclusion).
2. Se reclasifica la columna Tecnologia a "CdTe" para TODOS los
   candidatos, incluidos los que la fuente etiqueta genericamente "Thin
   Film" -- verificado (3-sep-2026): First Solar solo fabrica CdTe, y esos
   modulos "Thin Film" son First Solar mas antiguos/pequenos (formato
   clasico 1.2x0.6m de series 2-4), no un producto distinto. Sin esto,
   clasificar_tecnologia_jrc() (calculos/modelo_jrc_huld.py) no los
   reconoceria como CdTe y produccion.py los dejaria caer en SDM por
   defecto -- justo el problema que este ajuste evita.

Verificado con 2 fichas oficiales reales (Series 6 Plus, Series 7 TR1,
adjuntadas por el usuario): los datos de CEC para "CdTe" coinciden EXACTOS
(Voc/Vmp/Isc/Imp, 268 celdas). NOCT de CEC corre ~2.3C por encima del real
(mediana 47.3C vs 45C oficial confirmado en ambas fichas) -- documentado
en Notas de cada fila, no oculto.

Fuentes (mismas 2 que los imports anteriores):
1. PVS_params_translated.csv (Deville et al. 2025 IEEE JPV, Zenodo
   10.5281/zenodo.14173605) -- params PVsyst-v6.
2. CEC Modules.csv (NREL/SAM) -- NOCT/dimensiones/coef. temp. REPORTADOS.

DECISION DE DISENO (igual que los imports anteriores): no se inyectan
I_L_ref/I_o_ref/R_s/R_sh_ref -- el catalogo Excel nunca los guarda. El
unico dato del paper que si pasa al catalogo es gamma_ref, como "n
(Factor Idealidad)" -- solo usado por Motor IV/Mismatch/MPPT (SDM), NO
por JRC/Huld (que solo necesita Pmax_stc).

Uso:
    cd bipv_python
    python datos/agregar_paneles_firstsolar_nrel.py
"""
import csv
import re
import sys
from pathlib import Path
import openpyxl

EXCEL = Path(__file__).parent / "paneles_catalogo.xlsx"
SHEET = "Catalogo_Paneles_FV"

# Rutas de las 2 fuentes -- descargadas y verificadas el 2/3-sep-2026.
# No se distribuyen con el repo por tamano.
PATH_PVS = Path(r"C:\Users\Mauricio\Desktop\OPTIMIZADOR PARA CALCULOS BIPV\PRUEBA PVSYST vs MI APP\PVS_params_translated (1).csv")
PATH_CEC = Path(r"C:\Users\Mauricio\AppData\Local\Temp\claude\C--Users-Mauricio\a42e99e2-0c0d-4df5-8dad-a306b9744d70\scratchpad\cec_modules_sam.csv")


def _norm(s: str) -> str:
    return re.sub(r"[.,]", "", s).strip().lower()


def _cargar_pvs():
    with open(PATH_PVS, encoding="utf-8-sig") as f:
        return {r["module_name"]: r for r in csv.DictReader(f)}


def _cargar_cec():
    with open(PATH_CEC, encoding="utf-8-sig") as f:
        header = next(csv.reader(f))
    filas = {}
    with open(PATH_CEC, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f, fieldnames=header):
            if r["Name"] not in ("Name", ""):
                filas[r["Name"]] = r
    return filas


def _construir_pares_firstsolar(pvs, cec):
    """Cruce normalizado, solo First Solar, primer match por clave
    normalizada."""
    pvs_norm = {}
    for n in pvs:
        pvs_norm.setdefault(_norm(n), n)
    cec_norm = {}
    for n in cec:
        cec_norm.setdefault(_norm(n), n)
    comunes = set(pvs_norm) & set(cec_norm)
    return [
        (cec_norm[k], pvs_norm[k])
        for k in comunes
        if cec[cec_norm[k]]["Manufacturer"].strip().lower() == "first solar inc"
    ]


def _construir_fila(nombre_cec, pv, c):
    modelo = nombre_cec.replace("First Solar Inc ", "").strip()
    tech_original = c["Technology"].strip()

    Voc, Vmp, Isc, Imp = (float(c["V_oc_ref"]), float(c["V_mp_ref"]),
                          float(c["I_sc_ref"]), float(c["I_mp_ref"]))
    Pmax = float(c["STC"])
    Ns = int(c["N_s"])
    n_idealidad = round(float(pv["gamma_ref"]), 4)
    NsA = round(n_idealidad * Ns, 2)

    tiene_dim = bool(c["Length"].strip() and c["Width"].strip())
    dims_mm = f"{round(float(c['Length'])*1000)}x{round(float(c['Width'])*1000)}" if tiene_dim else "N/D"

    CoefVoc_C = round(float(c["beta_oc"]) / Voc * 100.0, 4)
    CoefT_C = round(float(c["gamma_pmp"]), 4)
    noct = c["T_NOCT"].strip() or "N/D"

    sys.path.insert(0, str(Path(__file__).parent.parent))
    from calculos.modelo_iv import validar_sdm_vs_ficha  # import perezoso
    Tk_alfa = float(c["alpha_sc"]) / Isc * 100.0
    panel_audit = {
        "nombre": modelo, "tecnologia": "CdTe", "N_s": Ns,
        "gamma_ref": n_idealidad, "mu_gamma": float(pv["mu_gamma"]),
        "I_L_ref": float(pv["I_L_ref"]), "I_o_ref": float(pv["I_o_ref"]),
        "R_s": float(pv["R_s"]), "R_sh_ref": float(pv["R_sh_ref"]),
        "R_sh_0": float(pv["R_sh_0"]), "Tk_alfa": Tk_alfa,
        "Voc_stc": Voc, "Isc_stc": Isc, "Vmp_stc": Vmp, "Imp_stc": Imp, "Pmax_stc": Pmax,
    }
    try:
        val = validar_sdm_vs_ficha(panel_audit, tolerancia_pct=6.0)
        sdm_ok = val["validacion_ok"]
        peor = max(("Voc", "Isc", "Vmp", "Imp", "Pmax"), key=lambda p: val[p]["error_pct"])
        peor_val = val[peor]["error_pct"]
    except Exception:
        sdm_ok, peor, peor_val = None, None, None

    pmp_nmbe = float(pv["pmp_nmbe"])
    partes = [
        "Fuente: NREL/SAM CEC Modules.csv (verificado exacto contra 2 fichas "
        "oficiales reales First Solar, Series 6 Plus y Series 7 TR1 -- Voc/Vmp/"
        "Isc/Imp coinciden) + Deville et al. 2025 IEEE JPV (params PVsyst-v6 "
        f"traducidos, pmp_nmbe={pmp_nmbe:.3f}%).",
        f"Tecnologia reclasificada a CdTe (fuente original: '{tech_original}') "
        "-- First Solar solo fabrica CdTe; necesario para que "
        "clasificar_tecnologia_jrc() la enrute a JRC/Huld en Produccion.",
        "⚡ Motor de ENERGIA real usado en Producción: JRC/Huld (no SDM) -- "
        "ver DIAGNOSTICO_JRC_HULD_PRIMARIO_CDTE.md. Motor IV/Mismatch/MPPT "
        "combinado SÍ siguen usando SDM para todas las tecnologías, con la "
        "limitación estructural ya documentada para CdTe (no exclusión, solo "
        "aviso).",
    ]
    if sdm_ok is False:
        partes.append(f"ℹ️ SDM (solo relevante para Motor IV/MPPT, no energía): {peor} se desvía {peor_val:.1f}% del ajuste PVsyst-v6 -- esperado para CdTe, no bloquea el import.")
    if not tiene_dim:
        partes.append(
            f"Dimensiones NO disponibles en la fuente (solo area A_c={c['A_c']}m2) "
            "-- completar con ficha real antes de usar en chequeo de densidad."
        )
    partes.append(
        "⚠️ NOCT tomado de CEC/NREL -- verificado contra 2 fichas reales que "
        "el NOCT oficial real de First Solar es 45°C; CEC reporta hasta "
        f"{noct}°C para este modelo (mediana del lote: 47.3°C, ~2.3°C por "
        "encima del real) -- confirmar con ficha específica antes de un "
        "proyecto real."
    )

    return {
        "ID": None, "Marca": "First Solar", "TipoPanel": modelo,
        "PmaxWp": Pmax, "DimensionesMM": dims_mm, "CostoUSD": 0,
        "NOCT_C": noct, "CoefT_C": CoefT_C, "TransparenciaPct": 0,
        "Notas": " | ".join(partes), "Tecnologia ": "CdTe",
        "Voc_STC": Voc, "Vmp_STC": Vmp, "Isc_STC": Isc, "Imp_STC": Imp,
        "CoefVoc_C": CoefVoc_C, "Ns (Celdas Serie)": Ns,
        "n (Factor Idealidad)": n_idealidad, "NsA = n × Ns": NsA,
        "Tecnología": "CdTe",
        "Fuente NsA": "NREL/SAM CEC Modules.csv + Sandia JPV 2025 (gamma_ref)",
        "Confianza": "alta (CEC verificado exacto contra 2 fichas oficiales reales)",
        "BifacialidadPct": 0,
    }


def main():
    if not PATH_PVS.exists():
        print(f"ERROR: no se encontro {PATH_PVS} -- descargar de Zenodo (ver docstring)")
        sys.exit(1)
    if not PATH_CEC.exists():
        print(f"ERROR: no se encontro {PATH_CEC} -- descargar de NREL/SAM (ver docstring)")
        sys.exit(1)
    if not EXCEL.exists():
        print(f"ERROR: no se encontro {EXCEL}")
        sys.exit(1)

    pvs, cec = _cargar_pvs(), _cargar_cec()
    pares = _construir_pares_firstsolar(pvs, cec)
    print(f"Candidatos First Solar (match normalizado): {len(pares)}")

    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]
    # SIN .strip() a proposito -- ver DIAGNOSTICO_CATALOGO_TRINA_NREL.md,
    # bug real del 3-sep-2026: el encabezado real tiene "Tecnologia " con
    # espacio al final; con .strip() aqui esa columna queda sin mapear y se
    # escribe en blanco sin ningun aviso.
    headers = [str(ws.cell(1, c).value) if ws.cell(1, c).value else "" for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    col_modelo = col_idx["TipoPanel"]

    existentes = set()
    last_id = 0
    for r in range(2, ws.max_row + 1):
        v_id = ws.cell(r, col_idx["ID"]).value
        if isinstance(v_id, (int, float)):
            last_id = max(last_id, int(v_id))
        v_modelo = ws.cell(r, col_modelo).value
        if v_modelo:
            existentes.add(str(v_modelo).strip())

    agregados, ya_existian = 0, 0
    for nombre_cec, nombre_pv in pares:
        pv, c = pvs[nombre_pv], cec[nombre_cec]
        fila = _construir_fila(nombre_cec, pv, c)
        if fila["TipoPanel"] in existentes:
            ya_existian += 1
            continue

        last_id += 1
        fila["ID"] = last_id
        target_row = ws.max_row + 1
        for col_name, value in fila.items():
            if col_name not in col_idx:
                print(f"  AVISO: columna '{col_name}' no encontrada en el catalogo real (se omite)")
                continue
            ws.cell(row=target_row, column=col_idx[col_name], value=value)
        agregados += 1

    wb.save(EXCEL)
    print(f"Agregados: {agregados}")
    print(f"Ya existian (omitidos, sin duplicar): {ya_existian}")
    print(f"Archivo guardado: {EXCEL}")


if __name__ == "__main__":
    main()
