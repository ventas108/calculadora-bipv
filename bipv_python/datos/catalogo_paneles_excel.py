"""Loader Catalogo_Paneles_FV — estructura unificada con costos y parámetros IV."""
import math
import re
import pandas as pd
import streamlit as st

import os as _os
# Ruta relativa al propio módulo (funciona en el servidor y en desarrollo);
# fallback a la ruta histórica del servidor por si el archivo se movió.
_EXCEL = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "paneles_catalogo.xlsx")
if not _os.path.exists(_EXCEL):
    _EXCEL = "/var/www/bipv/calculadora-bipv/bipv_python/datos/paneles_catalogo.xlsx"
_SHEET = "Catalogo_Paneles_FV"

def _f(val, default=None):
    # float(val) NO lanza excepción para NaN -- una celda vacía en una
    # columna numérica pandas la lee como NaN (no None), así que sin el
    # isfinite() de abajo esta función devolvía NaN en vez de `default`.
    # Bug real encontrado el 28-ago-2026 insertando paneles nuevos vía
    # guardar_panel_excel() con campos en None (Tk_gamma entre ellos):
    # nan or -0.45 (idiom usado en calculos.produccion) da NaN, no -0.45,
    # porque NaN es truthy en Python -- causaba
    # "P_dc_kW contiene valores no finitos" en cualquier simulación con esos
    # paneles. Mismo patrón que ya usa correctamente
    # datos/catalogo_inversores_excel.py::_f(), aplicado aquí también.
    try:
        v = float(val)
        return v if math.isfinite(v) else default
    except (TypeError, ValueError):
        return default

def _parse_area(dims):
    if not dims or str(dims).strip() in ('', 'N/D', 'Variable'):
        return None
    m = re.search(r'(\d+)\s*[xX×]\s*(\d+)', str(dims))
    return round(float(m.group(1)) * float(m.group(2)) / 1e6, 4) if m else None


def excel_mtime() -> float:
    """mtime del Excel de paneles -- mismo patrón #26 que ya usa
    datos/catalogo_inversores_excel.py::excel_mtime_inv(). Agregado
    4-sep-2026 para que optimization.variables._catalogo_paneles_real()
    (y su equivalente de inversores) puedan cachear el catálogo UNIDO sin
    reconstruirlo en cada llamada -- ver auditoría de rendimiento de CI en
    DIAGNOSTICO_CACHE_CATALOGO_UNIDO_SCENARIO_GENERATOR.md: sin esto,
    generar_candidatos() rehacía la unión completa (hasta ~3.100 paneles)
    en cada intento del muestreo (hasta 1.800 veces en un solo test)."""
    try:
        return _os.path.getmtime(_EXCEL)
    except OSError:
        return 0.0


@st.cache_data(ttl=3600)
def cargar_catalogo_paneles() -> dict:
    df = pd.read_excel(_EXCEL, sheet_name=_SHEET, header=0)
    df.columns = [str(c).strip() for c in df.columns]
    paneles = {}
    for _, r in df.iterrows():
        pmax = _f(r.get("PmaxWp"))
        if not pmax or pmax <= 0:
            continue
        Voc = _f(r.get("Voc_STC"))
        Isc = _f(r.get("Isc_STC"))
        if Voc is not None and Voc < 10:   continue
        if Isc is not None and Isc > 100:  continue
        nombre = str(r.get("TipoPanel", "")).strip()
        if not nombre or len(nombre) < 5: continue  # filtra entradas basura/prueba (ej. "yaya")
        Vmp = _f(r.get("Vmp_STC"))
        Imp = _f(r.get("Imp_STC"))
        costo = _f(r.get("CostoUSD"))
        paneles[nombre] = {
            "nombre":            nombre,
            "marca":             str(r.get("Marca", "")).strip(),
            "tecnologia":        str(r.get("Tecnologia", r.get("Tecnología", ""))).strip(),
            "Pmax_stc":          pmax,
            "dimensiones_mm":    str(r.get("DimensionesMM", "")).strip(),
            "area_m2":           _parse_area(r.get("DimensionesMM")),
            "costo_usd":         costo if (costo and costo > 0) else None,
            "NOCT":              _f(r.get("NOCT_C")),
            "beta_mp":           _f(r.get("CoefT_C")),
            "CoefVoc_C":         _f(r.get("CoefVoc_C")),
            "transparencia_pct": _f(r.get("TransparenciaPct"), 0),
            "bifacialidad_pct":  _f(r.get("BifacialidadPct"), 0),
            "Voc": Voc, "Vmp": Vmp,
            "Isc": Isc, "Imp": Imp,
            "N_s":         _f(r.get("Ns (Celdas Serie)")),
            "n_idealidad": _f(r.get("n (Factor Idealidad)")),
            "NsA":         _f(r.get("NsA = n × Ns")),
            "fuente_NsA":  str(r.get("Fuente NsA", "")).strip(),
            "confianza":   str(r.get("Confianza", "")).strip(),
            "notas":       str(r.get("Notas", "")).strip(),
            "I_sc_ref": Isc, "V_oc_ref": Voc,
            "I_mp_ref": Imp, "V_mp_ref": Vmp,
            "alpha_sc": None, "beta_oc": _f(r.get("CoefVoc_C")),
            "gamma_mp": _f(r.get("CoefT_C")),
            # ── Aliases para dimensionamiento.py ─────────────────────────
            "Voc_stc":  Voc,                          # = Voc_STC  del Excel
            "Vmp_stc":  Vmp,                          # = Vmp_STC  del Excel
            "Isc_stc":  Isc,                          # = Isc_STC  del Excel
            "Imp_stc":  Imp,                          # = Imp_STC  del Excel -- bug real
            # (30-ago-2026): faltaba mientras los otros 3 sí tenían alias
            # "_stc". calculos.modelo_iv.validar_sdm_vs_ficha() accede a
            # panel["Imp_stc"] con subíndice directo (sin .get()) -- sin este
            # alias, CUALQUIER panel del catálogo Excel sin SDM precalibrado
            # lanzaba KeyError dentro de preparar_panel_iv() (capturado y
            # silenciado como "datos insuficientes"), así que el ajuste
            # on-demand fit_desoto() nunca llegaba a activarse para ningún
            # panel real de este catálogo -- Motor IV quedaba mudo pese al
            # aviso "🟢 se activará automáticamente" en Dimensionamiento.
            "Tk_beta":  _f(r.get("CoefVoc_C")),       # coef. temp. Voc  (%/°C)
            "Tk_gamma": _f(r.get("CoefT_C")),         # coef. temp. Pmax (%/°C)
            "I_L_ref": None, "I_o_ref": None,
            "R_s": None, "R_sh_ref": None,
        }
    return paneles

def guardar_panel_excel(datos: dict, merge_conservador: bool = False) -> str:
    """
    Agrega o actualiza un panel en el Excel del catálogo.

    Parámetros
    ----------
    datos : dict
        Campos a escribir (claves = nombres de columna del Excel).
        Los campos con valor None no se escriben en modo merge_conservador.
    merge_conservador : bool (default False)
        Si True y el panel ya existe: solo sobreescribe los campos que traigan
        un valor no-None en `datos`. Los campos con None preservan el valor
        que ya había en el Excel (Ns, costo, notas, coeficientes curados a mano).
        Si False (inserción o actualización completa): escribe todos los campos,
        incluyendo None (borra el valor anterior).

    Retorna el nombre del panel guardado.
    Invalida el cache de st.cache_data para que el próximo cargar_catalogo_paneles() lea el nuevo panel.
    """
    import openpyxl, datetime

    nombre = str(datos.get("TipoPanel", "")).strip()
    if not nombre:
        raise ValueError("El campo TipoPanel (nombre del modelo) es obligatorio.")

    try:
        wb = openpyxl.load_workbook(_EXCEL)
    except FileNotFoundError:
        raise FileNotFoundError(f"No se encontró el archivo: {_EXCEL}")

    if _SHEET not in wb.sheetnames:
        raise ValueError(f"La hoja '{_SHEET}' no existe en {_EXCEL}.")

    ws = wb[_SHEET]

    # Leer encabezados de la primera fila
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

    # Comprobar si ya existe una fila con ese nombre (actualizar) o agregar
    fila_existente = None
    es_actualizacion = False
    col_tipo = None
    try:
        col_tipo = headers.index("TipoPanel") + 1   # 1-based
    except ValueError:
        raise ValueError("La hoja no tiene columna 'TipoPanel'.")

    for row in ws.iter_rows(min_row=2):
        if str(row[col_tipo - 1].value or "").strip() == nombre:
            fila_existente = row[0].row
            es_actualizacion = True
            break

    if fila_existente is None:
        fila_existente = ws.max_row + 1

    # Escribir datos — solo columnas que existen en el encabezado
    for col_nombre, valor in datos.items():
        if col_nombre not in headers:
            continue
        # Merge conservador: si el panel ya existía y el nuevo valor es None,
        # no sobreescribir — preservar lo que había en el Excel.
        if merge_conservador and es_actualizacion and valor is None:
            continue
        col_idx = headers.index(col_nombre) + 1
        ws.cell(row=fila_existente, column=col_idx, value=valor)

    # Anotar fecha de ingreso si hay columna
    for meta_col in ("FechaIngreso", "Fecha_Ingreso", "Fecha"):
        if meta_col in headers:
            ws.cell(
                row=fila_existente,
                column=headers.index(meta_col) + 1,
                value=datetime.date.today().isoformat()
            )
            break

    wb.save(_EXCEL)

    # Invalidar cache para que la próxima carga refleje el nuevo panel
    try:
        cargar_catalogo_paneles.clear()
    except Exception:
        pass

    return nombre


def eliminar_panel_excel(nombre: str) -> bool:
    """
    Elimina la fila del panel con TipoPanel == nombre del Excel.
    Retorna True si se eliminó, False si no se encontró.
    Invalida el cache.
    """
    import openpyxl

    nombre = nombre.strip()
    if not nombre:
        raise ValueError("Nombre vacío.")

    try:
        wb = openpyxl.load_workbook(_EXCEL)
    except FileNotFoundError:
        raise FileNotFoundError(f"No se encontró el archivo: {_EXCEL}")

    if _SHEET not in wb.sheetnames:
        raise ValueError(f"La hoja '{_SHEET}' no existe.")

    ws = wb[_SHEET]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    try:
        col_tipo = headers.index("TipoPanel") + 1
    except ValueError:
        raise ValueError("La hoja no tiene columna 'TipoPanel'.")

    fila_borrar = None
    for row in ws.iter_rows(min_row=2):
        if str(row[col_tipo - 1].value or "").strip() == nombre:
            fila_borrar = row[0].row
            break

    if fila_borrar is None:
        return False

    ws.delete_rows(fila_borrar)
    wb.save(_EXCEL)

    try:
        cargar_catalogo_paneles.clear()
    except Exception:
        pass

    return True


def actualizar_panel_excel(nombre_original: str, datos: dict) -> str:
    """
    Actualiza los campos de un panel existente (por nombre_original).
    Si datos contiene 'TipoPanel' distinto, también renombra el panel.
    Retorna el nombre final guardado.
    """
    datos_completos = {"TipoPanel": nombre_original}
    datos_completos.update(datos)
    return guardar_panel_excel(datos_completos)


def obtener_panel_excel(nombre: str) -> dict:
    return cargar_catalogo_paneles().get(nombre, {})

def lista_paneles_excel() -> list:
    return sorted(cargar_catalogo_paneles().keys())

# Alias de compatibilidad con Dimensionamiento.py
cargar_catalogo_excel = cargar_catalogo_paneles
