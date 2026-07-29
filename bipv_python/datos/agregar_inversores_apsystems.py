"""
Script de uso único: agrega APsystems AHS-6.3-SP, AHS-6.3 y AHS-5-LV
al catálogo de inversores si todavía no existen en la hoja Catalogo_Inversores.

Uso en el servidor:
    cd /var/www/bipv/calculadora-bipv
    source bipv_python/venv/bin/activate
    python bipv_python/datos/agregar_inversores_apsystems.py
"""
import sys
from pathlib import Path
import openpyxl

EXCEL = Path("/var/www/bipv/calculadora-bipv/bipv_python/datos/inversores_catalogo.xlsx")
SHEET = "Catalogo_Inversores"

# Datos técnicos extraídos de las fichas oficiales APsystems (ES) Rev 2026
INVERSORES = [
    {
        "Modelo":                                    "AHS-6.3-SP",
        "Datos completos (Si/No)":                   "Si",
        "Costo Inversor":                            "",
        # PV Input
        "Tension DC Maxima (V)":                     500,
        "Tension Arranque (V)":                      60,        # límite inferior MPPT
        "Rango MPPT Min (V)":                        60,
        "Rango MPPT Max (V)":                        450,
        "Tension Minima MPPT Activo (V)":            60,
        "N Trackers":                                1,
        "N Strings/Tracker":                         1,
        "Corriente Maxima Tracker (A)":              22,
        "Corriente Cortocircuito Max Tracker (A)":   "",        # no especificada en ficha
        "Potencia FV Max Recomendada (W)":           6500,
    },
    {
        "Modelo":                                    "AHS-6.3",
        "Datos completos (Si/No)":                   "Si",
        "Costo Inversor":                            "",
        "Tension DC Maxima (V)":                     500,
        "Tension Arranque (V)":                      60,
        "Rango MPPT Min (V)":                        60,
        "Rango MPPT Max (V)":                        450,
        "Tension Minima MPPT Activo (V)":            60,
        "N Trackers":                                1,
        "N Strings/Tracker":                         1,
        "Corriente Maxima Tracker (A)":              22,
        "Corriente Cortocircuito Max Tracker (A)":   "",
        "Potencia FV Max Recomendada (W)":           6500,
    },
    {
        "Modelo":                                    "AHS-5-LV",
        "Datos completos (Si/No)":                   "Si",
        "Costo Inversor":                            "",
        "Tension DC Maxima (V)":                     350,
        "Tension Arranque (V)":                      60,
        "Rango MPPT Min (V)":                        60,
        "Rango MPPT Max (V)":                        300,
        "Tension Minima MPPT Activo (V)":            60,
        "N Trackers":                                1,
        "N Strings/Tracker":                         1,
        "Corriente Maxima Tracker (A)":              22,
        "Corriente Cortocircuito Max Tracker (A)":   "",
        "Potencia FV Max Recomendada (W)":           6000,
    },
]


def main():
    if not EXCEL.exists():
        print(f"ERROR: No se encontró el archivo: {EXCEL}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL)

    if SHEET not in wb.sheetnames:
        print(f"ERROR: No existe la hoja '{SHEET}' en {EXCEL.name}")
        print(f"Hojas disponibles: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[SHEET]

    # Header en fila 3 (header=2 en pandas)
    header_row = 3
    headers = [str(ws.cell(header_row, c).value).strip()
               for c in range(1, ws.max_column + 1)]

    col_modelo = headers.index("Modelo") + 1  # 1-based

    # Modelos ya existentes en el catálogo
    existentes = set()
    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        val = str(row[col_modelo - 1].value).strip() if row[col_modelo - 1].value else ""
        if val:
            existentes.add(val.upper())

    agregados = 0
    for inv in INVERSORES:
        modelo = inv["Modelo"]
        if modelo.upper() in existentes:
            print(f"✓ '{modelo}' ya existe — omitido")
            continue

        next_row = ws.max_row + 1
        for col_name, value in inv.items():
            if col_name not in headers:
                print(f"  AVISO: columna '{col_name}' no encontrada (se omite)")
                continue
            col_idx = headers.index(col_name) + 1
            ws.cell(row=next_row, column=col_idx,
                    value=value if value != "" else None)

        print(f"✓ Agregado: {modelo} en fila {next_row}")
        existentes.add(modelo.upper())
        agregados += 1

    if agregados:
        wb.save(EXCEL)
        print(f"\nArchivo guardado: {EXCEL}")
        print(f"Total agregados: {agregados} inversor(es)")
    else:
        print("\nNo se realizaron cambios.")


if __name__ == "__main__":
    main()
