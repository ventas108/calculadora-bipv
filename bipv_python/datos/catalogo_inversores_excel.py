"""Loader Catalogo_Inversores con costos y semáforo de completitud."""
import math as _math
import os as _os

import pandas as pd
import streamlit as st

# Ruta relativa al propio módulo (funciona en el servidor y en desarrollo);
# fallback a la ruta histórica del servidor por si el archivo se movió. Antes
# hardcodeada solo a /var/www/... (sin el mismo fallback que ya tenía
# catalogo_paneles_excel.py) -- en cualquier entorno de desarrollo local esto
# hacía que cargar_catalogo_inversores()/guardar_inversor_excel() reventaran
# con FileNotFoundError, y optimization.variables._catalogo_inversores_real()
# caía en silencio al catálogo Python chico de 7 inversores en vez del Excel
# real de 105 -- encontrado 28-ago-2026 al intentar guardar un inversor real
# desde este entorno.
_EXCEL = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "inversores_catalogo.xlsx")
if not _os.path.exists(_EXCEL):
    _EXCEL = "/var/www/bipv/calculadora-bipv/bipv_python/datos/inversores_catalogo.xlsx"
_SHEET = "Catalogo_Inversores"

def _f(val, default=None):
    try:
        _value = float(val)
        return _value if _math.isfinite(_value) else default
    except (TypeError, ValueError):
        return default


def excel_mtime_inv() -> float:
    """mtime del Excel — entra como parte de la clave de caché para que
    cualquier edición directa al archivo invalide la caché sola (patrón #26)."""
    try:
        return _os.path.getmtime(_EXCEL)
    except OSError:
        return 0.0


def cargar_catalogo_inversores() -> dict:
    """API pública sin argumentos (7 llamadores). #205: inyecta el mtime del
    Excel en la clave de caché — editar el archivo en el servidor se refleja
    al siguiente rerun, sin esperar el TTL de 1 hora ni reiniciar PM2."""
    return _cargar_catalogo_inversores_cached(excel_mtime_inv())


@st.cache_data(ttl=3600)
def _cargar_catalogo_inversores_cached(mtime: float) -> dict:
    # OJO: el parámetro NO debe llevar guion bajo inicial — st.cache_data
    # excluye los args "_x" del hashing y el mtime no invalidaría nada.
    df = pd.read_excel(_EXCEL, sheet_name=_SHEET, header=2)
    df.columns = [str(c).strip() for c in df.columns]
    inversores = {}
    for _, r in df.iterrows():
        modelo = str(r.get("Modelo", "")).strip()
        if not modelo or modelo.lower() in ("", "nan"):
            continue
        completo = str(r.get("Datos completos (Si/No)", "")).strip().lower() == "si"
        costo = _f(r.get("Costo Inversor"))
        _p_ac_nom_kW = _f(r.get("Potencia AC nominal (kW)"))
        _p_dc_max_W  = _f(r.get("Potencia FV Max Recomendada (W)"))
        # P_ac_nom_W: columna directa > derivada de P_dc con factor 0.96 como fallback
        _p_ac_nom_W  = (_p_ac_nom_kW * 1000) if _p_ac_nom_kW else (
            _p_dc_max_W * 0.96 if _p_dc_max_W else None
        )
        inversores[modelo] = {
            "nombre":            modelo,
            "datos_completos":   completo,
            "costo_usd":         costo if (costo and costo > 0) else None,
            "Vdc_max":           _f(r.get("Tension DC Maxima (V)")),
            "V_arranque":        _f(r.get("Tension Arranque (V)")),
            "Vmppt_min":         _f(r.get("Rango MPPT Min (V)")),
            "Vmppt_max":         _f(r.get("Rango MPPT Max (V)")),
            "V_mppt_activo":     _f(r.get("Tension Minima MPPT Activo (V)")),
            "n_trackers":        _f(r.get("N Trackers")),
            "n_strings_tracker": _f(r.get("N Strings/Tracker")),
            "I_max_tracker":     _f(r.get("Corriente Maxima Tracker (A)")),
            "Isc_max_tracker":   _f(r.get("Corriente Cortocircuito Max Tracker (A)")),
            "P_dc_max_W":        _p_dc_max_W,
            # ── Potencia AC nominal ───────────────────────────────────────────
            "P_ac_nom_kW":       _p_ac_nom_kW,   # desde columna "Potencia AC nominal (kW)"
            "P_ac_nom_W":        _p_ac_nom_W,    # alias en Watts (para Dimensionamiento.py)
            # ── Aliases para dimensionamiento.py y Dimensionamiento.py ───
            "Vmppt_activo_min":  _f(r.get("Tension Minima MPPT Activo (V)")),  # = V_mppt_activo
            "N_mppt":            _f(r.get("N Trackers")),                       # = n_trackers
            # ── Compatibilidad con baterías (#25) — columnas opcionales ──────
            "es_hibrido":       str(r.get("Inversor Híbrido (Si/No)", "")).strip().lower() == "si",
            "bat_voltaje_min":  _f(r.get("Voltaje Batería Min (V)")),
            "bat_voltaje_max":  _f(r.get("Voltaje Batería Max (V)")),
            # Bug real corregido (30-ago-2026): estas 8 columnas (Marca,
            # Arquitectura, Inversor Híbrido, Voltaje Batería Min/Max,
            # Corriente Máxima Carga Batería, Confianza, Notas) las escribe
            # pages/15_Catálogo_Inversores_PDF.py desde hace tiempo, pero
            # NUNCA existieron como columnas reales en inversores_catalogo.xlsx
            # -- guardar_inversor_excel() solo escribe columnas que ya
            # existen en el encabezado, así que estos 8 campos se descartaban
            # en silencio en cada guardado. Confirmado con los 17 inversores
            # MUST reales ya guardados: los 17 son híbridos verdaderos, pero
            # los 17 mostraban es_hibrido=False. Columnas agregadas al Excel
            # y los 17 registros MUST corregidos -- ver
            # DIAGNOSTICO_EXTRACCION_INVERSORES_MUST.md.
            "marca":            str(r.get("Marca", "")).strip() or None,
            "arquitectura":     str(r.get("Arquitectura", "")).strip() or None,
            "bat_corriente_carga_max": _f(r.get("Corriente Máxima Carga Batería (A)")),
            "confianza":        str(r.get("Confianza", "")).strip() or None,
            "notas":            str(r.get("Notas", "")).strip() or None,
        }
    return inversores

# Compatibilidad: guardar/eliminar y la página 4 llaman .clear() sobre la pública
cargar_catalogo_inversores.clear = _cargar_catalogo_inversores_cached.clear
cargar_catalogo_inv_excel = cargar_catalogo_inversores


# ══════════════════════════════════════════════════════════════════════════════
# #122 — Diagnóstico del catálogo de inversores (mismo patrón que baterías #24)
# ══════════════════════════════════════════════════════════════════════════════

# Columnas del Excel según su impacto en el dimensionamiento de strings:
# críticas → sin ellas el dimensionamiento produce resultados incorrectos.
_COLS_CRITICAS = [
    "Modelo", "Tension DC Maxima (V)", "Rango MPPT Min (V)",
    "Rango MPPT Max (V)", "N Trackers", "Corriente Maxima Tracker (A)",
]
_COLS_IMPORTANTES = [
    "N Strings/Tracker", "Potencia FV Max Recomendada (W)",
    "Potencia AC nominal (kW)", "Tension Arranque (V)",
    "Tension Minima MPPT Activo (V)", "Corriente Cortocircuito Max Tracker (A)",
]
# Campos internos críticos por modelo (para reportar modelos incompletos)
_CAMPOS_CRITICOS_MODELO = ["Vdc_max", "Vmppt_min", "Vmppt_max",
                           "n_trackers", "I_max_tracker"]


@st.cache_data(ttl=3600)
def diagnostico_catalogo_inversores(mtime: float = 0.0) -> dict:
    """Diagnóstico del catálogo de inversores: hojas, columnas, duplicados.

    Nunca lanza — siempre devuelve un dict con 'estado' ("ok"/"parcial"/"error")
    para que la UI pinte el semáforo sin try/except propio.
    """
    info: dict = {"estado": "error", "hojas_disponibles": [],
                  "columnas_criticas_faltantes": [],
                  "columnas_importantes_faltantes": [],
                  "modelos_duplicados": [], "modelos_incompletos": [],
                  "modelos_cargados": 0}
    try:
        xl = pd.ExcelFile(_EXCEL, engine="openpyxl")
    except Exception as e:
        info["detalle"] = f"No se pudo abrir el Excel: {e}"
        return info

    info["hojas_disponibles"] = xl.sheet_names
    if _SHEET not in xl.sheet_names:
        info["detalle"] = (f"La hoja '{_SHEET}' no existe en el Excel. "
                           f"Hojas encontradas: {xl.sheet_names}")
        return info
    info["hoja_usada"] = _SHEET

    try:
        df = pd.read_excel(_EXCEL, sheet_name=_SHEET, header=2, engine="openpyxl")
        df.columns = [str(c).strip() for c in df.columns]
    except Exception as e:
        info["detalle"] = f"La hoja existe pero no se pudo leer (header fila 3): {e}"
        return info

    cols = set(df.columns)
    info["columnas_detectadas"] = [c for c in df.columns if "unnamed" not in c.lower()]
    info["columnas_criticas_faltantes"]    = [c for c in _COLS_CRITICAS if c not in cols]
    info["columnas_importantes_faltantes"] = [c for c in _COLS_IMPORTANTES if c not in cols]

    # ── Modelos duplicados (el dict se queda con el ÚLTIMO en silencio) ──────
    if "Modelo" in cols:
        _vistos: dict[str, list] = {}
        for _i, _v in df["Modelo"].items():
            _m = str(_v).strip()
            if _m and _m.lower() != "nan":
                # +4: fila real en Excel (2 filas de título + encabezado + base 1)
                _vistos.setdefault(_m, []).append(int(_i) + 4)
        info["modelos_duplicados"] = [
            {"modelo": m, "filas_excel": filas}
            for m, filas in _vistos.items() if len(filas) > 1
        ]

    # ── Modelos con campos críticos vacíos ────────────────────────────────────
    try:
        cat = cargar_catalogo_inversores()
        info["modelos_cargados"] = len(cat)
        for _n, _inv in cat.items():
            _falt = [c for c in _CAMPOS_CRITICOS_MODELO if not _inv.get(c)]
            if _falt:
                info["modelos_incompletos"].append({"modelo": _n, "campos_faltantes": _falt})
    except Exception as e:
        info["detalle"] = f"Las columnas existen pero el loader falló: {e}"
        return info

    if info["columnas_criticas_faltantes"] or info["modelos_cargados"] == 0:
        info["estado"] = "error"
        info.setdefault("detalle",
                        "Faltan columnas críticas o no se cargó ningún modelo.")
    elif (info["columnas_importantes_faltantes"] or info["modelos_duplicados"]
          or info["modelos_incompletos"]):
        info["estado"] = "parcial"
    else:
        info["estado"] = "ok"
    return info

def obtener_inversor_excel(nombre: str) -> dict:
    return cargar_catalogo_inversores().get(nombre, {})

def lista_inversores_excel() -> list:
    return sorted(cargar_catalogo_inversores().keys())


# ══════════════════════════════════════════════════════════════════════════════
# Escritura / edición / eliminación en el Excel
# ══════════════════════════════════════════════════════════════════════════════
# El archivo inversores_catalogo.xlsx tiene 2 filas de cabecera de título
# antes de la fila de encabezados de columna (header=2 en pd.read_excel →
# fila 3 de Excel en openpyxl).  Los datos empiezan en la fila 4.
_HEADER_ROW = 3   # 1-based row del encabezado de columnas en openpyxl
_DATA_START  = 4  # primera fila de datos

# Columna clave del catálogo de inversores
_KEY_COL = "Modelo"


def guardar_inversor_excel(datos: dict) -> str:
    """
    Agrega o actualiza un inversor en el Excel del catálogo.
    `datos` debe contener al menos la clave 'Modelo'.
    Retorna el nombre del inversor guardado e invalida el cache.
    """
    import openpyxl, datetime

    nombre = str(datos.get(_KEY_COL, "")).strip()
    if not nombre:
        raise ValueError("El campo 'Modelo' es obligatorio.")

    try:
        wb = openpyxl.load_workbook(_EXCEL)
    except FileNotFoundError:
        raise FileNotFoundError(f"No se encontró el archivo: {_EXCEL}")

    if _SHEET not in wb.sheetnames:
        raise ValueError(f"La hoja '{_SHEET}' no existe en {_EXCEL}.")

    ws = wb[_SHEET]

    # Leer encabezados desde la fila de cabecera real (fila 3 del Excel)
    headers = [
        str(c.value).strip() if c.value else ""
        for c in ws[_HEADER_ROW]
    ]

    if _KEY_COL not in headers:
        raise ValueError(f"La hoja no tiene columna '{_KEY_COL}'.")

    col_key = headers.index(_KEY_COL) + 1  # 1-based

    # Buscar fila existente o agregar al final
    fila_destino = None
    for row in ws.iter_rows(min_row=_DATA_START):
        val = str(row[col_key - 1].value or "").strip()
        if val == nombre:
            fila_destino = row[0].row
            break
    if fila_destino is None:
        fila_destino = ws.max_row + 1

    # Escribir solo columnas que existen en el encabezado
    for col_nombre, valor in datos.items():
        if col_nombre in headers:
            col_idx = headers.index(col_nombre) + 1
            ws.cell(row=fila_destino, column=col_idx, value=valor)

    # Anotar fecha si existe la columna
    for meta_col in ("FechaIngreso", "Fecha_Ingreso", "Fecha"):
        if meta_col in headers:
            ws.cell(
                row=fila_destino,
                column=headers.index(meta_col) + 1,
                value=datetime.date.today().isoformat(),
            )
            break

    wb.save(_EXCEL)

    try:
        cargar_catalogo_inversores.clear()
    except Exception:
        pass

    return nombre


def eliminar_inversor_excel(nombre: str) -> bool:
    """
    Elimina la fila del inversor con Modelo == nombre.
    Retorna True si se eliminó, False si no se encontró.
    """
    import openpyxl

    nombre = nombre.strip()
    if not nombre:
        raise ValueError("Nombre vacío.")

    try:
        wb = openpyxl.load_workbook(_EXCEL)
    except FileNotFoundError:
        raise FileNotFoundError(f"No se encontró el archivo: {_EXCEL}")

    if _SHEET not in wb.sheetnames:
        raise ValueError(f"La hoja '{_SHEET}' no existe.")

    ws = wb[_SHEET]
    headers = [str(c.value).strip() if c.value else "" for c in ws[_HEADER_ROW]]

    if _KEY_COL not in headers:
        raise ValueError(f"La hoja no tiene columna '{_KEY_COL}'.")

    col_key = headers.index(_KEY_COL) + 1
    fila_borrar = None
    for row in ws.iter_rows(min_row=_DATA_START):
        if str(row[col_key - 1].value or "").strip() == nombre:
            fila_borrar = row[0].row
            break

    if fila_borrar is None:
        return False

    ws.delete_rows(fila_borrar)
    wb.save(_EXCEL)

    try:
        cargar_catalogo_inversores.clear()
    except Exception:
        pass

    return True


def actualizar_inversor_excel(nombre_original: str, datos: dict) -> str:
    """
    Actualiza los campos de un inversor existente (por nombre_original).
    Si datos contiene 'Modelo' distinto, también renombra la entrada.
    Retorna el nombre final guardado.
    """
    datos_completos = {_KEY_COL: nombre_original}
    datos_completos.update(datos)
    return guardar_inversor_excel(datos_completos)
