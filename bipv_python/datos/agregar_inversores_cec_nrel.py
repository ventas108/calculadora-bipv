# -*- coding: utf-8 -*-
"""
Script de uso unico: agrega inversores reales del modelo Sandia/CEC de
NREL/SAM (datos/paneles_catalogo.xlsx tiene su equivalente de paneles;
este es el analogo para inversores) al catalogo real
(datos/inversores_catalogo.xlsx::Catalogo_Inversores) -- ver
DIAGNOSTICO_CATALOGO_INVERSORES_CEC_NREL.md para el detalle completo.

CRITERIO DISTINTO a los imports de paneles: el modelo Sandia de CEC NO
trae los campos MECANICOS que exige calculos.comparador_inversores.
filtrar_inversores_compatibles() y (desde el 4-sep-2026)
optimization.variables.variable_inversor() para el optimizador de Fase 4:
"N Trackers", "N Strings/Tracker", "Corriente Maxima Tracker (A)",
"Corriente Cortocircuito Max Tracker (A)" -- son datos de ficha mecanica,
el modelo electrico Sandia solo necesita una entrada DC agregada.

Por eso TODOS los inversores importados aqui quedan con "Datos completos"
= "No", y variable_inversor() los excluye automaticamente del
optimizador/comparador hasta que alguien los complete contra una ficha
real (mismo criterio ya usado para el resto del catalogo: nunca inventar
N_mppt). Quedan disponibles igual para comparar por clase de potencia
AC/DC y ventana de tension MPPT -- ese dato SI es 100% real.

Derivaciones desde el modelo Sandia (sin inventar nada):
- Potencia AC nominal (kW) = Paco / 1000 (dato Sandia directo).
- Potencia FV Max Recomendada (W) = Pdco -- ES la potencia DC de entrada
  que el modelo Sandia usa como referencia (a Vdc=Vdco produce
  exactamente Pac=Paco); NO es necesariamente la misma cifra que el
  fabricante publica como "sobredimensionamiento maximo recomendado"
  (ese es un limite de diseño/garantia, no un parametro del modelo
  electrico) -- se etiqueta en Notas para no confundir ambas cosas.
- Tension DC Maxima (V) = Vdcmax.
- Rango MPPT Min/Max (V) = Mppt_low / Mppt_high.
- Marca / Modelo: el campo Name de CEC sigue el patron real
  "Fabricante: Modelo {Vac}" (verificado en las 2.343 filas, 0 excepciones).

Fuente: NREL/SAM CEC Inverters.csv (github.com/NREL/SAM, rama patch,
deploy/libraries/CEC Inverters.csv) -- el mismo dataset que usa pvlib
para el modelo pvlib.inverter.sandia().

Uso:
    cd bipv_python
    python datos/agregar_inversores_cec_nrel.py
"""
import csv
import sys
from pathlib import Path
import openpyxl

EXCEL = Path(__file__).parent / "inversores_catalogo.xlsx"
SHEET = "Catalogo_Inversores"
HEADER_ROW = 3  # fila real del encabezado (pandas header=2 en el loader)

PATH_CEC = Path(r"C:\Users\Mauricio\AppData\Local\Temp\claude\C--Users-Mauricio\a42e99e2-0c0d-4df5-8dad-a306b9744d70\scratchpad\cec_inverters_sam.csv")


def _cargar_cec():
    with open(PATH_CEC, encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    # Las 2 primeras filas de datos son metadatos ("Units", "[0]") de SAM,
    # no inversores reales.
    return [r for r in rows if r["Name"] not in ("Units", "") and not r["Name"].startswith("[")]


def _construir_fila(r: dict) -> dict:
    nombre = r["Name"]
    marca, modelo = nombre.split(": ", 1)
    marca, modelo = marca.strip(), modelo.strip()

    Paco = float(r["Paco"])
    Pdco = float(r["Pdco"])
    Vac = float(r["Vac"])
    Vdcmax = float(r["Vdcmax"])
    Mppt_low = float(r["Mppt_low"])
    Mppt_high = float(r["Mppt_high"])

    notas = (
        "Fuente: NREL/SAM CEC Inverters.csv (modelo eléctrico Sandia oficial "
        "de CEC, el mismo que usa pvlib.inverter.sandia()). "
        f"Vac de referencia del modelo: {Vac:.0f}V (grid nominal de la "
        "caracterización CEC, no necesariamente el único soportado). "
        f"Potencia FV Max Recomendada = Pdco={Pdco:.0f}W -- es la potencia "
        "DC de referencia del modelo Sandia (a Vdc=Vdco produce exactamente "
        "Paco de salida), NO el límite oficial de sobredimensionamiento del "
        "fabricante (dato de diseño/garantía que el modelo eléctrico no "
        "reporta). "
        "⚠️ SIN datos mecánicos (N Trackers, Strings/Tracker, Corriente "
        "máxima/cortocircuito por tracker) -- el modelo Sandia no los "
        "necesita, son de ficha mecánica. Por eso queda 'Datos completos'="
        "'No' y optimization.variables.variable_inversor() lo excluye del "
        "optimizador/comparador de Fase 4 hasta completarlo contra una "
        "ficha real (ver DIAGNOSTICO_CATALOGO_INVERSORES_CEC_NREL.md) -- "
        "sirve hoy para comparar por clase de potencia AC/DC y ventana "
        "MPPT, ambos datos 100% reales."
    )

    return {
        "Datos completos (Si/No)": "No",
        "Modelo": modelo,
        "Costo Inversor ": None,
        "Archivo origen": "NREL/SAM CEC Inverters.csv",
        "Tension DC Maxima (V)": Vdcmax,
        "Tension Arranque (V)": None,
        "Rango MPPT Min (V)": Mppt_low,
        "Rango MPPT Max (V)": Mppt_high,
        "Tension Minima MPPT Activo (V)": None,
        "N Trackers": None,
        "N Strings/Tracker": None,
        "Corriente Maxima Tracker (A)": None,
        "Corriente Cortocircuito Max Tracker (A)": None,
        "Potencia FV Max Recomendada (W)": Pdco,
        "Potencia AC nominal (kW)": round(Paco / 1000.0, 4),
        "Marca": marca,
        "Arquitectura": None,
        "Inversor Híbrido (Si/No)": "No",
        "Voltaje Batería Min (V)": None,
        "Voltaje Batería Max (V)": None,
        "Corriente Máxima Carga Batería (A)": None,
        "Confianza": "alta (NREL/SAM CEC, modelo Sandia oficial) -- pendiente "
                     "completar datos mecánicos con ficha real",
        "Notas": notas,
    }


def main():
    if not PATH_CEC.exists():
        print(f"ERROR: no se encontro {PATH_CEC}")
        sys.exit(1)
    if not EXCEL.exists():
        print(f"ERROR: no se encontro {EXCEL}")
        sys.exit(1)

    filas_cec = _cargar_cec()
    print(f"Candidatos CEC Inverters.csv (filas de datos reales): {len(filas_cec)}")

    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]
    # SIN .strip() a proposito -- mismo patron que datos/agregar_paneles_*_nrel.py
    # (ver DIAGNOSTICO_CATALOGO_TRINA_NREL.md), para no repetir el bug real de
    # la columna "Tecnologia " con espacio.
    headers = [str(ws.cell(HEADER_ROW, c).value) if ws.cell(HEADER_ROW, c).value else "" for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    col_modelo = col_idx["Modelo"]
    col_marca = col_idx["Marca"]

    existentes = set()
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        v_modelo = ws.cell(r, col_modelo).value
        v_marca = ws.cell(r, col_marca).value
        if v_modelo:
            existentes.add((str(v_marca or "").strip(), str(v_modelo).strip()))

    agregados, ya_existian = 0, 0
    for r in filas_cec:
        fila = _construir_fila(r)
        clave = (fila["Marca"], fila["Modelo"])
        if clave in existentes:
            ya_existian += 1
            continue

        target_row = ws.max_row + 1
        for col_name, value in fila.items():
            if col_name not in col_idx:
                print(f"  AVISO: columna '{col_name}' no encontrada en el catalogo real (se omite)")
                continue
            ws.cell(row=target_row, column=col_idx[col_name], value=value)
        existentes.add(clave)
        agregados += 1

    wb.save(EXCEL)
    print(f"Agregados: {agregados}")
    print(f"Ya existian (omitidos, sin duplicar): {ya_existian}")
    print(f"Archivo guardado: {EXCEL}")


if __name__ == "__main__":
    main()
