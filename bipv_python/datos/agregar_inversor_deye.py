"""
Script de uso único: agrega DEYE SUN-7.6K-SG01LP1-US/EU al catálogo de inversores
si todavía no existe en la hoja Catalogo_Inversores.

Uso en el servidor:
    cd /var/www/bipv/calculadora-bipv
    source bipv_python/venv/bin/activate
    python bipv_python/datos/agregar_inversor_deye.py
"""
import sys
from pathlib import Path
import openpyxl

EXCEL = Path("/var/www/bipv/calculadora-bipv/bipv_python/datos/inversores_catalogo.xlsx")
SHEET = "Catalogo_Inversores"
MODELO = "SUN-7.6K-SG01LP1-US/EU"

# Datos técnicos extraídos de la ficha oficial DEYE
DEYE = {
    "Modelo":                                    MODELO,
    "Datos completos (Si/No)":                   "Si",
    "Costo Inversor":                            "",          # llenar cuando tenga cotización
    "Tension DC Maxima (V)":                     500,
    "Tension Arranque (V)":                      125,
    "Rango MPPT Min (V)":                        150,
    "Rango MPPT Max (V)":                        425,
    "Tension Minima MPPT Activo (V)":            200,
    "N Trackers":                                2,
    "N Strings/Tracker":                         2,
    "Corriente Maxima Tracker (A)":              26,
    "Corriente Cortocircuito Max Tracker (A)":   34,
    "Potencia FV Max Recomendada (W)":           9880,
}

def main():
    if not EXCEL.exists():
        print(f"ERROR: No se encontró el archivo: {EXCEL}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL)

    if SHEET not in wb.sheetnames:
        print(f"ERROR: No existe la hoja '{SHEET}' en {EXCEL.name}")
        print(f"Hojas disponibles: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[SHEET]

    # El header está en la fila 3 (header=2 en pandas → fila 3 en Excel)
    header_row = 3
    headers = [str(ws.cell(header_row, c).value).strip()
               for c in range(1, ws.max_column + 1)]

    # Verificar si el modelo ya existe (filas 4 en adelante)
    col_modelo = headers.index("Modelo") + 1  # columna 1-based
    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        val = str(row[col_modelo - 1].value).strip() if row[col_modelo - 1].value else ""
        if val.upper() == MODELO.upper():
            print(f"✓ El modelo '{MODELO}' ya existe en el catálogo. No se hizo ningún cambio.")
            return

    # Agregar nueva fila al final de los datos
    next_row = ws.max_row + 1
    for col_name, value in DEYE.items():
        if col_name in headers:
            col_idx = headers.index(col_name) + 1
            ws.cell(row=next_row, column=col_idx, value=value if value != "" else None)
        else:
            print(f"  AVISO: columna '{col_name}' no encontrada en la hoja (se omite)")

    wb.save(EXCEL)
    print(f"✓ Agregado: {MODELO} en fila {next_row}")
    print(f"  Archivo guardado: {EXCEL}")
    print()
    print("Datos ingresados:")
    for k, v in DEYE.items():
        print(f"  {k}: {v if v != '' else '(vacío - llenar cuando tenga cotización)'}")

if __name__ == "__main__":
    main()
