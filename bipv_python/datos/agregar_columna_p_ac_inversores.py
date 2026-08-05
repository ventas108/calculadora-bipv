"""
Script de migración: agrega la columna "Potencia AC nominal (kW)"
a la hoja Catalogo_Inversores del catálogo de inversores.

También rellena los valores conocidos para los modelos ya existentes.

Uso en el servidor:
    cd /var/www/bipv/calculadora-bipv
    source bipv_python/venv/bin/activate
    python bipv_python/datos/agregar_columna_p_ac_inversores.py
"""
import sys
from pathlib import Path
import openpyxl

EXCEL = Path("/var/www/bipv/calculadora-bipv/bipv_python/datos/inversores_catalogo.xlsx")
SHEET = "Catalogo_Inversores"
COL_NUEVA = "Potencia AC nominal (kW)"

# Potencias AC nominales (kW) por modelo — fuentes: fichas técnicas oficiales
P_AC_CONOCIDOS = {
    # Growatt
    "MID15KTL3-X":             15.0,
    "MID17KTL3-X":             17.0,
    "MID20KTL3-X":             20.0,
    "MID25KTL3-X":             25.0,
    "MID30KTL3-X":             30.0,
    "MIN3000TL-XH":             3.0,
    "MIN3600TL-XH":             3.6,
    "MIN4200TL-XH":             4.2,
    "MIN5000TL-XH":             5.0,
    "MIN6000TL-XH":             6.0,
    # APsystems
    "AHS-6.3-SP":               6.3,
    "AHS-6.3":                  6.3,
    "AHS-5-LV":                 5.0,
    # DEYE
    "SUN-7.6K-SG01LP1-US/EU":  7.6,
    "SUN-5K-SG01LP1-EU":        5.0,
    "SUN-8K-SG01LP1-EU":        8.0,
    "SUN-10K-SG01LP1-EU":      10.0,
    "SUN-12K-SG01LP1-EU":      12.0,
}


def main():
    if not EXCEL.exists():
        print(f"ERROR: No se encontró el archivo: {EXCEL}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL)
    if SHEET not in wb.sheetnames:
        print(f"ERROR: No existe la hoja '{SHEET}'")
        sys.exit(1)

    ws = wb[SHEET]
    header_row = 3  # header=2 en pandas → fila 3 en Excel

    headers = [
        str(ws.cell(header_row, c).value).strip()
        for c in range(1, ws.max_column + 1)
    ]

    # Verificar si la columna ya existe
    if COL_NUEVA in headers:
        print(f"✓ La columna '{COL_NUEVA}' ya existe. Se actualizarán valores faltantes.")
        col_p_ac = headers.index(COL_NUEVA) + 1
    else:
        # Agregar columna al final
        col_p_ac = ws.max_column + 1
        ws.cell(header_row, col_p_ac, value=COL_NUEVA)
        print(f"✓ Columna '{COL_NUEVA}' agregada en posición {col_p_ac}.")

    # Columna Modelo
    if "Modelo" not in headers:
        print("ERROR: No se encontró la columna 'Modelo'")
        sys.exit(1)
    col_modelo = headers.index("Modelo") + 1

    actualizados = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        modelo = ws.cell(row_idx, col_modelo).value
        if not modelo or str(modelo).strip().lower() in ("", "nan", "none"):
            continue
        modelo = str(modelo).strip()
        p_ac = P_AC_CONOCIDOS.get(modelo)
        celda = ws.cell(row_idx, col_p_ac)
        if p_ac is not None and not celda.value:
            celda.value = p_ac
            print(f"  → {modelo}: {p_ac} kW")
            actualizados += 1
        elif celda.value:
            print(f"  · {modelo}: ya tiene {celda.value} kW (sin cambio)")
        else:
            print(f"  ⚠ {modelo}: sin valor en P_AC_CONOCIDOS — llenar manualmente")

    wb.save(EXCEL)
    print(f"\n✓ Guardado. {actualizados} modelo(s) actualizados con Potencia AC nominal.")
    print(
        "\nNOTA: Para los modelos sin valor, completar manualmente la columna "
        f"'{COL_NUEVA}' en el Excel. El loader usará P_dc_max × 0.96 como fallback."
    )


if __name__ == "__main__":
    main()
