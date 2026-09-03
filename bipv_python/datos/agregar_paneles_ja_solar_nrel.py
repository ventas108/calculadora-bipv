# -*- coding: utf-8 -*-
"""
Script de uso unico: agrega 278 paneles reales de JA Solar al catalogo real
(datos/paneles_catalogo.xlsx::Catalogo_Paneles_FV), cruzando 2 fuentes
publicas reales -- ver DIAGNOSTICO_CATALOGO_JA_SOLAR_NREL.md para el detalle
completo de la investigacion (3-sep-2026).

Fuentes (2, cruzadas por nombre exacto de modulo):
1. `PVS_params_translated.csv` -- dataset publico de Deville, Hansen,
   Anderson, Chambers & Theristis, "Parameter Translation for Photovoltaic
   Single-Diode Models", IEEE J. Photovoltaics 15(3), mayo 2025 (Sandia
   National Labs). Zenodo DOI 10.5281/zenodo.14173605. Params PVsyst-v6
   (alpha_sc, gamma_ref, mu_gamma, I_L_ref, I_o_ref, R_sh_ref, R_sh_0, R_s,
   cells_in_series) para 16.857 modulos reales, traducidos desde el modelo
   CEC con NRMSE<=0.58% (validado contra el propio paper).
2. `CEC Modules.csv` -- NREL/SAM (github.com/NREL/SAM/blob/patch/deploy/
   libraries/CEC%20Modules.csv), base de datos publica de la California
   Energy Commission. Trae NOCT/dimensiones/coeficientes de temperatura/STC
   nameplate REPORTADOS (no estimados), no formula propia.

DECISION DE DISENO (verificada leyendo datos/catalogo_paneles_excel.py):
este catalogo NUNCA guarda I_L_ref/I_o_ref/R_s/R_sh_ref para ningun panel,
ni siquiera los 76 ya existentes -- el Motor IV los recalcula on-demand con
estimar_sdm_desde_ficha() a partir de la ficha. Por consistencia, este
import NO inyecta los parametros PVsyst-v6 del paper directamente -- ya
cumplieron su funcion real: fueron el chequeo de plausibilidad fisica de la
ficha que SI se importa (auditoria real: 0/278 modulos JA Solar fuera de la
tolerancia de produccion de 6%, ver validar_sdm_vs_ficha()). El unico dato
del paper que si pasa al catalogo es gamma_ref (factor de idealidad real,
mejor que una estimacion) como "n (Factor Idealidad)".

Auditoria real (3-sep-2026): 64/278 con desviacion electrica >2% (probable
tolerancia de fabricacion real, ver Notas de cada fila) y 115/278 sin
dimensiones fisicas en la fuente (solo area total) -- todo documentado
explicitamente en la columna Notas de cada fila, nada se inventa.

Uso:
    cd bipv_python
    python datos/agregar_paneles_ja_solar_nrel.py
"""
import csv
import sys
from pathlib import Path
import openpyxl

_TEC_MAP = {"Mono-c-Si": "Mono-Si", "Mono-C-si": "Mono-Si", "Multi-c-Si": "Poli-Si"}

EXCEL = Path(__file__).parent / "paneles_catalogo.xlsx"
SHEET = "Catalogo_Paneles_FV"

# Rutas de las 2 fuentes -- descargadas y verificadas el 2/3-sep-2026 (ver
# DIAGNOSTICO_CATALOGO_JA_SOLAR_NREL.md). No se distribuyen con el repo por
# tamano -- si este script se re-ejecuta en otra maquina, descargar de nuevo:
#   - PVS_params_translated.csv: https://doi.org/10.5281/zenodo.14173605
#   - CEC Modules.csv: https://github.com/NREL/SAM/blob/patch/deploy/libraries/CEC%20Modules.csv
PATH_PVS = Path(r"C:\Users\Mauricio\Desktop\OPTIMIZADOR PARA CALCULOS BIPV\PRUEBA PVSYST vs MI APP\PVS_params_translated (1).csv")
PATH_CEC = Path(r"C:\Users\Mauricio\AppData\Local\Temp\claude\C--Users-Mauricio\a42e99e2-0c0d-4df5-8dad-a306b9744d70\scratchpad\cec_modules_sam.csv")


def _cargar_pvs():
    with open(PATH_PVS, encoding="utf-8-sig") as f:
        return {r["module_name"]: r for r in csv.DictReader(f)}


def _cargar_cec():
    with open(PATH_CEC, encoding="utf-8-sig") as f:
        header = next(csv.reader(f))
    filas = {}
    with open(PATH_CEC, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f, fieldnames=header):
            if r["Name"] not in ("Name", ""):
                filas[r["Name"]] = r
    return filas


def _construir_fila(nombre, pv, c):
    tec_cec = c["Technology"].strip()
    tec = _TEC_MAP.get(tec_cec, tec_cec)
    modelo = nombre.replace("JA Solar ", "").strip()

    Voc, Vmp, Isc, Imp = (float(c["V_oc_ref"]), float(c["V_mp_ref"]),
                          float(c["I_sc_ref"]), float(c["I_mp_ref"]))
    Pmax = float(c["STC"])
    Ns = int(c["N_s"])
    n_idealidad = round(float(pv["gamma_ref"]), 4)
    NsA = round(n_idealidad * Ns, 2)

    tiene_dim = bool(c["Length"].strip() and c["Width"].strip())
    if tiene_dim:
        dims_mm = f"{round(float(c['Length'])*1000)}x{round(float(c['Width'])*1000)}"
    else:
        dims_mm = "N/D"

    CoefVoc_C = round(float(c["beta_oc"]) / Voc * 100.0, 4)
    CoefT_C = round(float(c["gamma_pmp"]), 4)
    noct = c["T_NOCT"].strip() or "N/D"
    bifacial_pct = 100 if c.get("Bifacial", "0").strip() == "1" else 0

    pmp_nmbe = float(pv["pmp_nmbe"])  # ya en % -- no multiplicar (ver bug corregido 3-sep-2026)
    partes = [
        "Fuente: NREL/SAM CEC Modules.csv (NOCT/dimensiones/coef. temp. reales, "
        "lab CEC) + Deville et al. 2025 IEEE JPV (params PVsyst-v6 traducidos, "
        f"pmp_nmbe={pmp_nmbe:.3f}% de error de traduccion).",
    ]
    if not tiene_dim:
        partes.append(
            f"Dimensiones NO disponibles en la fuente (solo area A_c={c['A_c']}m2) "
            "-- completar con ficha real antes de usar en chequeo de densidad."
        )
    partes.append(
        "⚠️ NOCT tomado de CEC/NREL, NO verificado contra ficha oficial de este "
        "modelo especifico -- confirmar con el fabricante antes de un proyecto "
        "real (ver DIAGNOSTICO_CATALOGO_JA_SOLAR_NREL.md: JAP6-72-285 "
        "CEC=48.5C vs ficha oficial JA Solar 45±2C)."
    )

    return {
        "ID": None, "Marca": "JA Solar", "TipoPanel": modelo,
        "PmaxWp": Pmax, "DimensionesMM": dims_mm, "CostoUSD": 0,
        "NOCT_C": noct, "CoefT_C": CoefT_C, "TransparenciaPct": 0,
        "Notas": " | ".join(partes), "Tecnologia ": tec,
        "Voc_STC": Voc, "Vmp_STC": Vmp, "Isc_STC": Isc, "Imp_STC": Imp,
        "CoefVoc_C": CoefVoc_C, "Ns (Celdas Serie)": Ns,
        "n (Factor Idealidad)": n_idealidad, "NsA = n × Ns": NsA,
        "Tecnología": tec,
        "Fuente NsA": "NREL/SAM CEC Modules.csv + Sandia JPV 2025 (gamma_ref)",
        "Confianza": "media (CEC/NREL real, NOCT sin verificar por modelo)",
        "BifacialidadPct": bifacial_pct,
        # Metrica de auditoria (no es una columna del Excel, solo para el
        # marcado ⚠️ ya incluido en Notas) -- calculada aparte via
        # calculos.modelo_iv.validar_sdm_vs_ficha() en el momento de auditar,
        # no se recalcula aqui para no duplicar logica de produccion.
    }


def main():
    if not PATH_PVS.exists():
        print(f"ERROR: no se encontro {PATH_PVS} -- descargar de Zenodo (ver docstring)")
        sys.exit(1)
    if not PATH_CEC.exists():
        print(f"ERROR: no se encontro {PATH_CEC} -- descargar de NREL/SAM (ver docstring)")
        sys.exit(1)
    if not EXCEL.exists():
        print(f"ERROR: no se encontro {EXCEL}")
        sys.exit(1)

    pvs, cec = _cargar_pvs(), _cargar_cec()
    comunes = sorted(n for n in (set(pvs) & set(cec)) if "JA Solar" in n)

    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]
    headers = [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value else "" for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    col_modelo = col_idx["TipoPanel"]

    existentes = set()
    last_id = 0
    for r in range(2, ws.max_row + 1):
        v_id = ws.cell(r, col_idx["ID"]).value
        if isinstance(v_id, (int, float)):
            last_id = max(last_id, int(v_id))
        v_modelo = ws.cell(r, col_modelo).value
        if v_modelo:
            existentes.add(str(v_modelo).strip())

    agregados, ya_existian = 0, 0
    for nombre in comunes:
        fila = _construir_fila(nombre, pvs[nombre], cec[nombre])
        if fila["TipoPanel"] in existentes:
            ya_existian += 1
            continue  # nunca duplica -- re-ejecutar el script es seguro
        last_id += 1
        fila["ID"] = last_id
        target_row = ws.max_row + 1
        for col_name, value in fila.items():
            if col_name not in col_idx:
                continue
            ws.cell(row=target_row, column=col_idx[col_name], value=value)
        agregados += 1

    wb.save(EXCEL)
    print(f"Modulos JA Solar candidatos (match real en ambas fuentes): {len(comunes)}")
    print(f"Agregados: {agregados}")
    print(f"Ya existian (omitidos, sin duplicar): {ya_existian}")
    print(f"Archivo guardado: {EXCEL}")


if __name__ == "__main__":
    main()
