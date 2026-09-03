# -*- coding: utf-8 -*-
"""
Script de uso unico: agrega paneles reales de SunPower (Mono-Si,
familias A-Series y Maxeon 3) al catalogo real
(datos/paneles_catalogo.xlsx::Catalogo_Paneles_FV) -- ver
DIAGNOSTICO_CATALOGO_SUNPOWER_NREL.md para el detalle completo.

DOS criterios de exclusion (distintos entre si):

1. Familia "-R" (SPR-MAX3-XXX-R, -COM-R, -BLK-R): son AC Modules
   ("Residential AC Module" -- microinversor integrado). Verificado real
   contra ficha publica de Maxeon 3 (secondsol/enfsolar: Voc=75.6V,
   Isc=6.58A, 104 celdas -> V/celda=0.727V): las variantes SIN "-R" dan
   V/celda=0.724-0.726V (coincide); las variantes CON "-R" dan siempre
   V/celda=0.362V (exactamente la mitad) e Isc aprox el doble -- no es un
   bug de traduccion del paper (a diferencia de Trina/Jinko), es un
   producto categoricamente distinto: un modulo AC con microinversor
   integrado no se modela como panel DC de string (N_serie a un inversor
   central no aplica). 21 excluidos.
2. Tolerancia SDM (validar_sdm_vs_ficha, 6%) sobre el resto -- 6
   excluidos, familia SPR-A-COM (72 celdas): Vmp*Imp no coincide con el
   Pmax de placa en el propio dato fuente de CEC (ej. SPR-A400-COM:
   Vmp*Imp=437.3W vs Pmax nameplate=400W, 9.3% de discrepancia) --
   inconsistencia real de la fuente (parece copiar el Pmax de la version
   de 66 celdas sin escalarlo a las 72 celdas reales), no artefacto de
   nuestro modelo.

Fuentes (mismas 2 que los imports anteriores):
1. PVS_params_translated.csv (Deville et al. 2025 IEEE JPV, Zenodo
   10.5281/zenodo.14173605) -- params PVsyst-v6.
2. CEC Modules.csv (NREL/SAM) -- NOCT/dimensiones/coef. temp. REPORTADOS.

DECISION DE DISENO (igual que los imports anteriores): no se inyectan
I_L_ref/I_o_ref/R_s/R_sh_ref -- el catalogo Excel nunca los guarda. El
unico dato del paper que si pasa al catalogo es gamma_ref, como "n
(Factor Idealidad)" -- solo usado por Motor IV/Mismatch/MPPT (SDM).

Uso:
    cd bipv_python
    python datos/agregar_paneles_sunpower_nrel.py
"""
import csv
import re
import sys
from pathlib import Path
import openpyxl

EXCEL = Path(__file__).parent / "paneles_catalogo.xlsx"
SHEET = "Catalogo_Paneles_FV"

# Rutas de las 2 fuentes -- descargadas y verificadas el 3-sep-2026.
# No se distribuyen con el repo por tamano.
PATH_PVS = Path(r"C:\Users\Mauricio\Desktop\OPTIMIZADOR PARA CALCULOS BIPV\PRUEBA PVSYST vs MI APP\PVS_params_translated (1).csv")
PATH_CEC = Path(r"C:\Users\Mauricio\AppData\Local\Temp\claude\C--Users-Mauricio\a42e99e2-0c0d-4df5-8dad-a306b9744d70\scratchpad\cec_modules_sam.csv")

TOLERANCIA_ELECTRICA_PCT = 6.0
_TEC_MAP = {"Mono-c-Si": "Mono-Si", "Mono-C-si": "Mono-Si", "Multi-c-Si": "Poli-Si"}


def _norm(s: str) -> str:
    return re.sub(r"[.,]", "", s).strip().lower()


def _cargar_pvs():
    with open(PATH_PVS, encoding="utf-8-sig") as f:
        return {r["module_name"]: r for r in csv.DictReader(f)}


def _cargar_cec():
    with open(PATH_CEC, encoding="utf-8-sig") as f:
        header = next(csv.reader(f))
    filas = {}
    with open(PATH_CEC, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f, fieldnames=header):
            if r["Name"] not in ("Name", ""):
                filas[r["Name"]] = r
    return filas


def _construir_pares_sunpower(pvs, cec):
    """Cruce normalizado, solo SunPower, primer match por clave
    normalizada."""
    pvs_norm = {}
    for n in pvs:
        pvs_norm.setdefault(_norm(n), n)
    cec_norm = {}
    for n in cec:
        cec_norm.setdefault(_norm(n), n)
    comunes = set(pvs_norm) & set(cec_norm)
    return [
        (cec_norm[k], pvs_norm[k])
        for k in comunes
        if cec[cec_norm[k]]["Manufacturer"].strip().lower() == "sunpower"
    ]


def _auditar_electrico(nombre, c, pv):
    sys.path.insert(0, str(Path(__file__).parent.parent))
    from calculos.modelo_iv import validar_sdm_vs_ficha  # import perezoso

    Voc, Vmp, Isc, Imp = (float(c["V_oc_ref"]), float(c["V_mp_ref"]),
                          float(c["I_sc_ref"]), float(c["I_mp_ref"]))
    Pmax = float(c["STC"])
    Ns = int(c["N_s"])
    n_idealidad = float(pv["gamma_ref"])
    Tk_alfa = float(c["alpha_sc"]) / Isc * 100.0
    tec = _TEC_MAP.get(c["Technology"].strip(), c["Technology"].strip())
    panel = {
        "nombre": nombre, "tecnologia": tec, "N_s": Ns,
        "gamma_ref": n_idealidad, "mu_gamma": float(pv["mu_gamma"]),
        "I_L_ref": float(pv["I_L_ref"]), "I_o_ref": float(pv["I_o_ref"]),
        "R_s": float(pv["R_s"]), "R_sh_ref": float(pv["R_sh_ref"]),
        "R_sh_0": float(pv["R_sh_0"]), "Tk_alfa": Tk_alfa,
        "Voc_stc": Voc, "Isc_stc": Isc, "Vmp_stc": Vmp, "Imp_stc": Imp, "Pmax_stc": Pmax,
    }
    return validar_sdm_vs_ficha(panel, tolerancia_pct=TOLERANCIA_ELECTRICA_PCT)


def _construir_fila(nombre_cec, pv, c):
    modelo = nombre_cec.replace("SunPower ", "").strip()
    Voc, Vmp, Isc, Imp = (float(c["V_oc_ref"]), float(c["V_mp_ref"]),
                          float(c["I_sc_ref"]), float(c["I_mp_ref"]))
    Pmax = float(c["STC"])
    Ns = int(c["N_s"])
    n_idealidad = round(float(pv["gamma_ref"]), 4)
    NsA = round(n_idealidad * Ns, 2)
    tec = _TEC_MAP.get(c["Technology"].strip(), c["Technology"].strip())

    tiene_dim = bool(c["Length"].strip() and c["Width"].strip())
    dims_mm = f"{round(float(c['Length'])*1000)}x{round(float(c['Width'])*1000)}" if tiene_dim else "N/D"

    CoefVoc_C = round(float(c["beta_oc"]) / Voc * 100.0, 4)
    CoefT_C = round(float(c["gamma_pmp"]), 4)
    noct = c["T_NOCT"].strip() or "N/D"

    pmp_nmbe = float(pv["pmp_nmbe"])
    partes = [
        "Fuente: NREL/SAM CEC Modules.csv + Deville et al. 2025 IEEE JPV "
        f"(params PVsyst-v6 traducidos, pmp_nmbe={pmp_nmbe:.3f}%).",
        "Verificado contra ficha pública real Maxeon 3 (75.6V/6.58A/104 "
        "celdas, secondsol/enfsolar) -- V/celda=0.72-0.73V coincide con "
        "las variantes sin sufijo -R importadas aquí (las -R son AC "
        "Module con microinversor, excluidas -- ver docstring del script).",
    ]
    if not tiene_dim:
        partes.append(
            f"Dimensiones NO disponibles en la fuente (solo area A_c={c['A_c']}m2) "
            "-- completar con ficha real antes de usar en chequeo de densidad."
        )
    partes.append(
        f"NOCT tomado de CEC/NREL ({noct}°C) -- confirmar con ficha específica "
        "antes de un proyecto real."
    )

    return {
        "ID": None, "Marca": "SunPower", "TipoPanel": modelo,
        "PmaxWp": Pmax, "DimensionesMM": dims_mm, "CostoUSD": 0,
        "NOCT_C": noct, "CoefT_C": CoefT_C, "TransparenciaPct": 0,
        "Notas": " | ".join(partes), "Tecnologia ": tec,
        "Voc_STC": Voc, "Vmp_STC": Vmp, "Isc_STC": Isc, "Imp_STC": Imp,
        "CoefVoc_C": CoefVoc_C, "Ns (Celdas Serie)": Ns,
        "n (Factor Idealidad)": n_idealidad, "NsA = n × Ns": NsA,
        "Tecnología": tec,
        "Fuente NsA": "NREL/SAM CEC Modules.csv + Sandia JPV 2025 (gamma_ref)",
        "Confianza": "alta (CEC + tolerancia SDM ≤6% + verificado vs ficha real Maxeon 3)",
        "BifacialidadPct": 0,
    }


def main():
    if not PATH_PVS.exists():
        print(f"ERROR: no se encontro {PATH_PVS} -- descargar de Zenodo (ver docstring)")
        sys.exit(1)
    if not PATH_CEC.exists():
        print(f"ERROR: no se encontro {PATH_CEC} -- descargar de NREL/SAM (ver docstring)")
        sys.exit(1)
    if not EXCEL.exists():
        print(f"ERROR: no se encontro {EXCEL}")
        sys.exit(1)

    pvs, cec = _cargar_pvs(), _cargar_cec()
    pares = _construir_pares_sunpower(pvs, cec)
    print(f"Candidatos SunPower (match normalizado): {len(pares)}")

    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]
    # SIN .strip() a proposito -- ver DIAGNOSTICO_CATALOGO_TRINA_NREL.md,
    # bug real del 3-sep-2026: el encabezado real tiene "Tecnologia " con
    # espacio al final; con .strip() aqui esa columna queda sin mapear y se
    # escribe en blanco sin ningun aviso.
    headers = [str(ws.cell(1, c).value) if ws.cell(1, c).value else "" for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    col_modelo = col_idx["TipoPanel"]

    existentes = set()
    last_id = 0
    for r in range(2, ws.max_row + 1):
        v_id = ws.cell(r, col_idx["ID"]).value
        if isinstance(v_id, (int, float)):
            last_id = max(last_id, int(v_id))
        v_modelo = ws.cell(r, col_modelo).value
        if v_modelo:
            existentes.add(str(v_modelo).strip())

    agregados, ya_existian, n_ac_module, n_excluidos_electrico = 0, 0, 0, 0
    for nombre_cec, nombre_pv in sorted(pares):
        pv, c = pvs[nombre_pv], cec[nombre_cec]

        if nombre_cec.rstrip().split("-")[-1] == "R":
            n_ac_module += 1
            continue

        val = _auditar_electrico(nombre_cec, c, pv)
        if not val["validacion_ok"]:
            n_excluidos_electrico += 1
            continue

        fila = _construir_fila(nombre_cec, pv, c)
        if fila["TipoPanel"] in existentes:
            ya_existian += 1
            continue

        last_id += 1
        fila["ID"] = last_id
        target_row = ws.max_row + 1
        for col_name, value in fila.items():
            if col_name not in col_idx:
                print(f"  AVISO: columna '{col_name}' no encontrada en el catalogo real (se omite)")
                continue
            ws.cell(row=target_row, column=col_idx[col_name], value=value)
        agregados += 1

    wb.save(EXCEL)
    print(f"Excluidos por ser AC Module (-R, microinversor integrado): {n_ac_module}")
    print(f"Excluidos por tolerancia SDM >6%: {n_excluidos_electrico}")
    print(f"Agregados: {agregados}")
    print(f"Ya existian (omitidos, sin duplicar): {ya_existian}")
    print(f"Archivo guardado: {EXCEL}")


if __name__ == "__main__":
    main()
