# -*- coding: utf-8 -*-
"""
Script de uso unico: agrega paneles reales de Trina Solar al catalogo real
(datos/paneles_catalogo.xlsx::Catalogo_Paneles_FV) -- ver
DIAGNOSTICO_CATALOGO_TRINA_NREL.md para el detalle completo de la
investigacion (3-sep-2026).

Mismo patron que datos/agregar_paneles_ja_solar_nrel.py (JA Solar), con una
diferencia real: el cruce se hace por nombre NORMALIZADO (sin puntuacion),
no exacto -- el cruce exacto solo encontraba 269 modulos Trina; con
normalizacion, 1.312. Investigado con 2 casos reales contra ficha oficial:
la diferencia de puntuacion en si misma NO es senal de colision entre
productos distintos (Trina simplemente escribe su nomenclatura con/sin
punto segun la fuente) -- lo que SI importa es el chequeo fisico real
(validar_sdm_vs_ficha, tolerancia 6%), que atrapa los casos problematicos
sin importar su causa exacta (error de traduccion del paper para
arquitecturas de alta potencia con strings en paralelo, o dato propio de
CEC que no coincide con el fabricante).

Fuentes (mismas 2 que JA Solar, ver ese script para las URLs completas):
1. PVS_params_translated.csv (Deville et al. 2025 IEEE JPV, Zenodo
   10.5281/zenodo.14173605) -- params PVsyst-v6.
2. CEC Modules.csv (NREL/SAM) -- NOCT/dimensiones/coef. temp. REPORTADOS.

Auditoria real: de 1.312 candidatos con match normalizado, 57 fallan la
tolerancia real de produccion (6%, validar_sdm_vs_ficha) -- EXCLUIDOS de
este import, no se fuerzan. Quedan 1.255 candidatos reales.

DECISION DE DISENO (igual que JA Solar): no se inyectan I_L_ref/I_o_ref/
R_s/R_sh_ref -- el catalogo Excel nunca los guarda, el Motor IV los
recalcula on-demand con estimar_sdm_desde_ficha(). El unico dato del paper
que si pasa al catalogo es gamma_ref, como "n (Factor Idealidad)".

Uso:
    cd bipv_python
    python datos/agregar_paneles_trina_nrel.py
"""
import csv
import re
import sys
from pathlib import Path
import openpyxl

_TEC_MAP = {"Mono-c-Si": "Mono-Si", "Mono-C-si": "Mono-Si", "Multi-c-Si": "Poli-Si"}

EXCEL = Path(__file__).parent / "paneles_catalogo.xlsx"
SHEET = "Catalogo_Paneles_FV"

TOLERANCIA_ELECTRICA_PCT = 6.0  # mismo umbral real de produccion (validar_sdm_vs_ficha)

# Rutas de las 2 fuentes -- descargadas y verificadas el 2/3-sep-2026 (ver
# DIAGNOSTICO_CATALOGO_JA_SOLAR_NREL.md / DIAGNOSTICO_CATALOGO_TRINA_NREL.md).
# No se distribuyen con el repo por tamano.
PATH_PVS = Path(r"C:\Users\Mauricio\Desktop\OPTIMIZADOR PARA CALCULOS BIPV\PRUEBA PVSYST vs MI APP\PVS_params_translated (1).csv")
PATH_CEC = Path(r"C:\Users\Mauricio\AppData\Local\Temp\claude\C--Users-Mauricio\a42e99e2-0c0d-4df5-8dad-a306b9744d70\scratchpad\cec_modules_sam.csv")


def _norm(s: str) -> str:
    return re.sub(r"[.,]", "", s).strip().lower()


def _cargar_pvs():
    with open(PATH_PVS, encoding="utf-8-sig") as f:
        return {r["module_name"]: r for r in csv.DictReader(f)}


def _cargar_cec():
    with open(PATH_CEC, encoding="utf-8-sig") as f:
        header = next(csv.reader(f))
    filas = {}
    with open(PATH_CEC, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f, fieldnames=header):
            if r["Name"] not in ("Name", ""):
                filas[r["Name"]] = r
    return filas


def _construir_pares_trina(pvs, cec):
    """Cruce normalizado, solo Trina Solar, con el primer match por clave
    normalizada (sin colisiones reales detectadas para esta marca, ver
    DIAGNOSTICO)."""
    pvs_norm = {}
    for n in pvs:
        pvs_norm.setdefault(_norm(n), n)
    cec_norm = {}
    for n in cec:
        cec_norm.setdefault(_norm(n), n)
    comunes = set(pvs_norm) & set(cec_norm)
    return [
        (cec_norm[k], pvs_norm[k])
        for k in comunes
        if "trina" in cec[cec_norm[k]]["Manufacturer"].lower()
    ]


def _construir_fila(nombre_cec, pv, c):
    tec_cec = c["Technology"].strip()
    tec = _TEC_MAP.get(tec_cec, tec_cec)
    modelo = nombre_cec.replace("Trina Solar ", "").strip()

    Voc, Vmp, Isc, Imp = (float(c["V_oc_ref"]), float(c["V_mp_ref"]),
                          float(c["I_sc_ref"]), float(c["I_mp_ref"]))
    Pmax = float(c["STC"])
    Ns = int(c["N_s"])
    n_idealidad = round(float(pv["gamma_ref"]), 4)
    NsA = round(n_idealidad * Ns, 2)

    tiene_dim = bool(c["Length"].strip() and c["Width"].strip())
    if tiene_dim:
        dims_mm = f"{round(float(c['Length'])*1000)}x{round(float(c['Width'])*1000)}"
    else:
        dims_mm = "N/D"

    CoefVoc_C = round(float(c["beta_oc"]) / Voc * 100.0, 4)
    CoefT_C = round(float(c["gamma_pmp"]), 4)
    noct = c["T_NOCT"].strip() or "N/D"
    bifacial_pct = 100 if c.get("Bifacial", "0").strip() == "1" else 0

    pmp_nmbe = float(pv["pmp_nmbe"])
    partes = [
        "Fuente: NREL/SAM CEC Modules.csv (NOCT/dimensiones/coef. temp. reales, "
        "lab CEC) + Deville et al. 2025 IEEE JPV (params PVsyst-v6 traducidos, "
        f"pmp_nmbe={pmp_nmbe:.3f}% de error de traduccion). Nombre cruzado por "
        f"coincidencia normalizada (fuente paper: '{pv['module_name']}').",
    ]
    if not tiene_dim:
        partes.append(
            f"Dimensiones NO disponibles en la fuente (solo area A_c={c['A_c']}m2) "
            "-- completar con ficha real antes de usar en chequeo de densidad."
        )
    partes.append(
        "⚠️ NOCT tomado de CEC/NREL, NO verificado contra ficha oficial de este "
        "modelo especifico -- confirmar con el fabricante antes de un proyecto "
        "real (ver DIAGNOSTICO_CATALOGO_TRINA_NREL.md)."
    )

    return {
        "ID": None, "Marca": "Trina Solar", "TipoPanel": modelo,
        "PmaxWp": Pmax, "DimensionesMM": dims_mm, "CostoUSD": 0,
        "NOCT_C": noct, "CoefT_C": CoefT_C, "TransparenciaPct": 0,
        "Notas": " | ".join(partes), "Tecnologia ": tec,
        "Voc_STC": Voc, "Vmp_STC": Vmp, "Isc_STC": Isc, "Imp_STC": Imp,
        "CoefVoc_C": CoefVoc_C, "Ns (Celdas Serie)": Ns,
        "n (Factor Idealidad)": n_idealidad, "NsA = n × Ns": NsA,
        "Tecnología": tec,
        "Fuente NsA": "NREL/SAM CEC Modules.csv + Sandia JPV 2025 (gamma_ref)",
        "Confianza": "media (CEC/NREL real, NOCT sin verificar por modelo)",
        "BifacialidadPct": bifacial_pct,
    }


def _auditar_electrico(fila, N_s):
    """Reproduce el chequeo real de validar_sdm_vs_ficha() (STC, PVsyst-v6)
    sin importar streamlit -- este script corre standalone. Devuelve True si
    pasa la tolerancia real de produccion en las 5 metricas."""
    sys.path.insert(0, str(Path(__file__).parent.parent))
    from calculos.modelo_iv import validar_sdm_vs_ficha  # import perezoso
    panel = {
        "nombre": fila["TipoPanel"], "tecnologia": fila["Tecnologia "].strip(),
        "N_s": N_s, "gamma_ref": fila["n (Factor Idealidad)"],
        "mu_gamma": fila["_mu_gamma"], "I_L_ref": fila["_I_L_ref"],
        "I_o_ref": fila["_I_o_ref"], "R_s": fila["_R_s"], "R_sh_ref": fila["_R_sh_ref"],
        "R_sh_0": fila["_R_sh_0"], "Tk_alfa": fila["_Tk_alfa"],
        "Voc_stc": fila["Voc_STC"], "Isc_stc": fila["Isc_STC"],
        "Vmp_stc": fila["Vmp_STC"], "Imp_stc": fila["Imp_STC"], "Pmax_stc": fila["PmaxWp"],
    }
    val = validar_sdm_vs_ficha(panel, tolerancia_pct=TOLERANCIA_ELECTRICA_PCT)
    return val["validacion_ok"]


def main():
    if not PATH_PVS.exists():
        print(f"ERROR: no se encontro {PATH_PVS} -- descargar de Zenodo (ver docstring)")
        sys.exit(1)
    if not PATH_CEC.exists():
        print(f"ERROR: no se encontro {PATH_CEC} -- descargar de NREL/SAM (ver docstring)")
        sys.exit(1)
    if not EXCEL.exists():
        print(f"ERROR: no se encontro {EXCEL}")
        sys.exit(1)

    pvs, cec = _cargar_pvs(), _cargar_cec()
    pares = _construir_pares_trina(pvs, cec)
    print(f"Candidatos Trina (match normalizado): {len(pares)}")

    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]
    # SIN .strip() a proposito -- ver DIAGNOSTICO_CATALOGO_TRINA_NREL.md,
    # bug real del 3-sep-2026: el encabezado real tiene "Tecnologia " con
    # espacio al final; con .strip() aqui esa columna queda sin mapear y se
    # escribe en blanco para las 1.255 filas de este import, sin aviso.
    headers = [str(ws.cell(1, c).value) if ws.cell(1, c).value else "" for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    col_modelo = col_idx["TipoPanel"]

    existentes = set()
    last_id = 0
    for r in range(2, ws.max_row + 1):
        v_id = ws.cell(r, col_idx["ID"]).value
        if isinstance(v_id, (int, float)):
            last_id = max(last_id, int(v_id))
        v_modelo = ws.cell(r, col_modelo).value
        if v_modelo:
            existentes.add(str(v_modelo).strip())

    agregados, ya_existian, excluidos_electrico = 0, 0, 0
    for nombre_cec, nombre_pv in pares:
        pv, c = pvs[nombre_pv], cec[nombre_cec]
        fila = _construir_fila(nombre_cec, pv, c)
        if fila["TipoPanel"] in existentes:
            ya_existian += 1
            continue

        fila["_mu_gamma"] = float(pv["mu_gamma"])
        fila["_I_L_ref"] = float(pv["I_L_ref"])
        fila["_I_o_ref"] = float(pv["I_o_ref"])
        fila["_R_s"] = float(pv["R_s"])
        fila["_R_sh_ref"] = float(pv["R_sh_ref"])
        fila["_R_sh_0"] = float(pv["R_sh_0"])
        fila["_Tk_alfa"] = float(pv["alpha_sc"]) / fila["Isc_STC"] * 100.0
        if not _auditar_electrico(fila, int(c["N_s"])):
            excluidos_electrico += 1
            continue

        last_id += 1
        fila["ID"] = last_id
        target_row = ws.max_row + 1
        for col_name, value in fila.items():
            if col_name.startswith("_"):
                continue
            if col_name not in col_idx:
                print(f"  AVISO: columna '{col_name}' no encontrada en el catalogo real (se omite)")
                continue
            ws.cell(row=target_row, column=col_idx[col_name], value=value)
        agregados += 1

    wb.save(EXCEL)
    print(f"Agregados: {agregados}")
    print(f"Excluidos por fallar tolerancia real (>{TOLERANCIA_ELECTRICA_PCT}%): {excluidos_electrico}")
    print(f"Ya existian (omitidos, sin duplicar): {ya_existian}")
    print(f"Archivo guardado: {EXCEL}")


if __name__ == "__main__":
    main()
