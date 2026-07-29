"""
Agrega la hoja 'Catalogo_Baterias' al Excel de inversores del servidor.
Incluye los 26 modelos de batería industrial (BR series + ATESS ESS).

Uso en el servidor:
    cd /var/www/bipv/calculadora-bipv
    source bipv_python/venv/bin/activate
    python bipv_python/datos/agregar_hoja_baterias.py
"""
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL = Path("/var/www/bipv/calculadora-bipv/bipv_python/datos/inversores_catalogo.xlsx")
SHEET = "Catalogo_Baterias"

# ── Colores ────────────────────────────────────────────────────────────────
FILL_VERDE   = PatternFill("solid", fgColor="1B5E20")
FILL_NARANJA = PatternFill("solid", fgColor="E65100")
FILL_GRIS    = PatternFill("solid", fgColor="37474F")
FILL_TITULO  = PatternFill("solid", fgColor="1565C0")
FILL_PAR     = PatternFill("solid", fgColor="E3F2FD")
FILL_WARN    = PatternFill("solid", fgColor="FFF9C4")

FONT_WHITE_B = Font(bold=True, color="FFFFFF", size=10)
FONT_BLACK_B = Font(bold=True, size=9)
FONT_NORMAL  = Font(size=9)

BORDER = Border(
    left=Side(style="thin", color="B0BEC5"),
    right=Side(style="thin", color="B0BEC5"),
    top=Side(style="thin", color="B0BEC5"),
    bottom=Side(style="thin", color="B0BEC5"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_WRAP   = Alignment(vertical="center", wrap_text=True)

# ── Definición de columnas ─────────────────────────────────────────────────
# (nombre, color, ancho)
COLUMNAS = [
    ("Modelo",                  FILL_VERDE,   18),
    ("Fabricante",              FILL_NARANJA, 20),
    ("Datos completos (Si/No)", FILL_VERDE,   12),
    ("Tecnologia",              FILL_VERDE,   16),
    ("Capacidad (kWh)",         FILL_VERDE,   12),
    ("DoD (%)",                 FILL_VERDE,   10),
    ("Eficiencia RTE (%)",      FILL_VERDE,   12),
    ("Ciclos de Vida",          FILL_VERDE,   12),
    ("Potencia Continua (kW)",  FILL_VERDE,   14),
    ("Potencia Pico (kW)",      FILL_NARANJA, 12),
    ("Voltaje Nominal (V)",     FILL_NARANJA, 13),
    ("Voltaje Min (V)",         FILL_NARANJA, 12),
    ("Voltaje Max (V)",         FILL_NARANJA, 12),
    ("Temperatura Min (C)",     FILL_GRIS,    12),
    ("Temperatura Max (C)",     FILL_GRIS,    12),
    ("Peso (kg)",               FILL_GRIS,    10),
    ("IP",                      FILL_GRIS,    8),
    ("Montaje",                 FILL_GRIS,    18),
    ("Garantia (anos)",         FILL_NARANJA, 11),
    ("Costo (USD)",             FILL_NARANJA, 12),
    ("Notas",                   FILL_GRIS,    55),
]

HV = "ALTA TENSION — requiere inversor HV comercial/industrial (NO compatible con 48V)."

# ── Datos de baterías (26 modelos) ─────────────────────────────────────────
def kw(kwh, c): return round(kwh * c, 2)

BATERIAS = [
    # ── Fabricante pendiente ───────────────────────────────────────────────
    ("BR172R","Pendiente confirmar","No","LiFePO4",172.032,None,None,6000,kw(172.032,0.5),kw(172.032,1),614.4,537.6,691.2,None,None,1511,"IP20","Rack interior",None,None,f"{HV} 12x14.336 kWh. BMS CAN. Pantalla 7\"."),
    ("BR186R","Pendiente confirmar","No","LiFePO4",186.368,None,None,6000,kw(186.368,0.5),kw(186.368,1),665.6,582.4,748.8,None,None,1624,"IP20","Rack interior",None,None,f"{HV} 13x14.336 kWh. BMS CAN."),
    ("BR200R","Pendiente confirmar","No","LiFePO4",200.704,None,None,6000,kw(200.704,0.5),kw(200.704,1),716.8,627.2,806.4,None,None,1737,"IP20","Rack interior",None,None,f"{HV} 14x14.336 kWh. BMS CAN."),
    ("BR215R","Pendiente confirmar","No","LiFePO4",215.040,None,None,6000,kw(215.040,0.5),kw(215.040,1),768.0,672.0,864.0,None,None,1850,"IP20","Rack interior",None,None,f"{HV} 15x14.336 kWh. BMS CAN."),
    # ── ATESS ESS — BC/BR45T-60T (24S1P, 7.68 kWh/mod) ──────────────────
    ("BC45T","ATESS ESS","No","LiFePO4",46.08,None,None,6000,kw(46.08,0.5),kw(46.08,1),460.8,403.2,525.6,None,None,716,"IP54","Gabinete exterior",None,None,f"{HV} 6x7.68 kWh. Variante interior: BR45T."),
    ("BR45T","ATESS ESS","No","LiFePO4",46.08,None,None,6000,kw(46.08,0.5),kw(46.08,1),460.8,403.2,525.6,None,None,539,"IP20","Rack interior",None,None,f"{HV} 6x7.68 kWh. Variante exterior IP54: BC45T."),
    ("BC50T","ATESS ESS","No","LiFePO4",53.76,None,None,6000,kw(53.76,0.5),kw(53.76,1),537.6,470.4,613.2,None,None,792,"IP54","Gabinete exterior",None,None,f"{HV} 7x7.68 kWh. Variante interior: BR50T."),
    ("BR50T","ATESS ESS","No","LiFePO4",53.76,None,None,6000,kw(53.76,0.5),kw(53.76,1),537.6,470.4,613.2,None,None,615,"IP20","Rack interior",None,None,f"{HV} 7x7.68 kWh. Variante exterior IP54: BC50T."),
    ("BC60T","ATESS ESS","No","LiFePO4",61.44,None,None,6000,kw(61.44,0.5),kw(61.44,1),614.4,537.6,700.8,None,None,868,"IP54","Gabinete exterior",None,None,f"{HV} 8x7.68 kWh. Variante interior: BR60T."),
    ("BR60T","ATESS ESS","No","LiFePO4",61.44,None,None,6000,kw(61.44,0.5),kw(61.44,1),614.4,537.6,700.8,None,None,691,"IP20","Rack interior",None,None,f"{HV} 8x7.68 kWh. Variante exterior IP54: BC60T."),
    # ── ATESS ESS — BC/BR75T-145T (12S2P, 7.68 kWh/mod) ─────────────────
    ("BC75T","ATESS ESS","No","LiFePO4",76.8,None,None,6000,kw(76.8,0.5),kw(76.8,1),384.0,336.0,438.0,None,None,1130,"IP54","Gabinete exterior",None,None,f"{HV} 10x7.68 kWh. Variante interior: BR75T."),
    ("BR75T","ATESS ESS","No","LiFePO4",76.8,None,None,6000,kw(76.8,0.5),kw(76.8,1),384.0,336.0,438.0,None,None,877,"IP20","Rack interior",None,None,f"{HV} 10x7.68 kWh. Variante exterior IP54: BC75T."),
    ("BC100T","ATESS ESS","No","LiFePO4",107.52,None,None,6000,kw(107.52,0.5),kw(107.52,1),537.6,470.4,613.2,None,None,1436,"IP54","Gabinete exterior",None,None,f"{HV} 14x7.68 kWh. Variante interior: BR100T."),
    ("BR100T","ATESS ESS","No","LiFePO4",107.52,None,None,6000,kw(107.52,0.5),kw(107.52,1),537.6,470.4,613.2,None,None,1183,"IP20","Rack interior",None,None,f"{HV} 14x7.68 kWh. Variante exterior IP54: BC100T."),
    ("BR138T","ATESS ESS","No","LiFePO4",138.24,None,None,6000,kw(138.24,0.5),kw(138.24,1),691.2,604.8,766.8,None,None,1547,"IP20","Rack interior",None,None,f"{HV} 18x7.68 kWh. Solo interior."),
    ("BR145T","ATESS ESS","No","LiFePO4",145.92,None,None,6000,kw(145.92,0.5),kw(145.92,1),729.6,638.4,832.2,None,None,1624,"IP20","Rack interior",None,None,f"{HV} 19x7.68 kWh. Solo interior."),
    # ── ATESS ESS — BR114R-157R (16S1P, 14.336 kWh/mod) ─────────────────
    ("BR114R","ATESS ESS","No","LiFePO4",114.688,None,None,6000,kw(114.688,0.5),kw(114.688,1),409.6,358.4,460.8,None,None,1064,"IP20","Rack interior",None,None,f"{HV} 8x14.336 kWh. 166 Wh/kg. BMS CAN+RS485."),
    ("BR129R","ATESS ESS","No","LiFePO4",129.024,None,None,6000,kw(129.024,0.5),kw(129.024,1),460.8,403.2,518.4,None,None,1157,"IP20","Rack interior",None,None,f"{HV} 9x14.336 kWh. 166 Wh/kg. BMS CAN+RS485."),
    ("BR143R","ATESS ESS","No","LiFePO4",143.36,None,None,6000,kw(143.36,0.5),kw(143.36,1),512.0,448.0,576.0,None,None,1270,"IP20","Rack interior",None,None,f"{HV} 10x14.336 kWh. 166 Wh/kg. BMS CAN+RS485."),
    ("BR157R","ATESS ESS","No","LiFePO4",157.696,None,None,6000,kw(157.696,0.5),kw(157.696,1),563.2,492.8,633.6,None,None,1383,"IP20","Rack interior",None,None,f"{HV} 11x14.336 kWh. 166 Wh/kg. BMS CAN+RS485."),
    # ── ATESS ESS — BC55RPB (16S1P, 5.12 kWh/mod, IP54) ─────────────────
    ("BC55RPB-6M","ATESS ESS","No","LiFePO4",30.72,None,None,6000,kw(30.72,0.5),kw(30.72,1),307.2,268.8,345.6,None,None,474,"IP54","Gabinete exterior",None,None,f"{HV} 6x5.12 kWh. 100Ah/mod."),
    ("BC55RPB-7M","ATESS ESS","No","LiFePO4",35.84,None,None,6000,kw(35.84,0.5),kw(35.84,1),358.4,313.6,403.2,None,None,518,"IP54","Gabinete exterior",None,None,f"{HV} 7x5.12 kWh."),
    ("BC55RPB-8M","ATESS ESS","No","LiFePO4",40.96,None,None,6000,kw(40.96,0.5),kw(40.96,1),409.6,358.4,460.8,None,None,562,"IP54","Gabinete exterior",None,None,f"{HV} 8x5.12 kWh."),
    ("BC55RPB-9M","ATESS ESS","No","LiFePO4",46.08,None,None,6000,kw(46.08,0.5),kw(46.08,1),460.8,403.2,518.4,None,None,606,"IP54","Gabinete exterior",None,None,f"{HV} 9x5.12 kWh."),
    ("BC55RPB-10M","ATESS ESS","No","LiFePO4",51.2,None,None,6000,kw(51.2,0.5),kw(51.2,1),512.0,448.0,576.0,None,None,650,"IP54","Gabinete exterior",None,None,f"{HV} 10x5.12 kWh."),
    ("BC55RPB-11M","ATESS ESS","No","LiFePO4",56.32,None,None,6000,kw(56.32,0.5),kw(56.32,1),563.2,492.8,642.4,None,None,694,"IP54","Gabinete exterior",None,None,f"{HV} 11x5.12 kWh."),
]


def aplicar_estilo(cell, fill=None, font=None, alignment=None, border=True):
    if fill:      cell.fill = fill
    if font:      cell.font = font
    if alignment: cell.alignment = alignment
    if border:    cell.border = BORDER


def main():
    if not EXCEL.exists():
        print(f"ERROR: No se encontró: {EXCEL}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL)

    if SHEET in wb.sheetnames:
        print(f"✓ La hoja '{SHEET}' ya existe — actualizando...")
        del wb[SHEET]

    ws = wb.create_sheet(SHEET)
    print(f"✓ Creando hoja '{SHEET}'...")

    # ── Fila 1: Título ─────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 24
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNAS))}1")
    c = ws["A1"]
    c.value = f"CATÁLOGO DE BATERÍAS — BIPV COLOMBIA  |  {len(BATERIAS)} modelos  |  Alta tensión 300–870 V"
    aplicar_estilo(c, FILL_TITULO, FONT_WHITE_B, ALIGN_CENTER)

    # ── Fila 2: Leyenda ────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 16
    for col_idx in range(1, len(COLUMNAS) + 1):
        aplicar_estilo(ws.cell(2, col_idx), border=True)
    ws["A2"].value = "🟢 OBLIGATORIO — requerido para calcular"
    ws["A2"].fill = FILL_VERDE
    ws["A2"].font = FONT_WHITE_B
    ws["A2"].alignment = ALIGN_CENTER
    ws.merge_cells("A2:D2")
    ws["E2"].value = "🟠 IMPORTANTE — completar con datos del proveedor"
    ws["E2"].fill = FILL_NARANJA
    ws["E2"].font = FONT_WHITE_B
    ws["E2"].alignment = ALIGN_CENTER
    ws.merge_cells("E2:M2")
    ws["N2"].value = "⬜ OPCIONAL"
    ws["N2"].fill = FILL_GRIS
    ws["N2"].font = FONT_WHITE_B
    ws["N2"].alignment = ALIGN_CENTER
    ws.merge_cells(f"N2:{get_column_letter(len(COLUMNAS))}2")

    # ── Fila 3: Encabezados ────────────────────────────────────────────────
    ws.row_dimensions[3].height = 36
    for ci, (nombre, fill, ancho) in enumerate(COLUMNAS, start=1):
        col_letter = get_column_letter(ci)
        ws.column_dimensions[col_letter].width = ancho
        c = ws.cell(3, ci)
        c.value = nombre
        aplicar_estilo(c, fill, FONT_WHITE_B, ALIGN_CENTER)

    # ── Filas de datos ─────────────────────────────────────────────────────
    for bi, datos in enumerate(BATERIAS):
        row = 4 + bi
        ws.row_dimensions[row].height = 50
        par = bi % 2 == 0
        fill_fila = FILL_PAR if par else None

        for ci, valor in enumerate(datos, start=1):
            c = ws.cell(row, ci)
            c.value = valor
            # Notas: siempre amarillo advertencia
            if ci == len(COLUMNAS):
                aplicar_estilo(c, FILL_WARN, FONT_NORMAL, ALIGN_WRAP)
            elif isinstance(valor, (int, float)):
                aplicar_estilo(c, fill_fila, FONT_BLACK_B,
                               Alignment(horizontal="center", vertical="center"))
            else:
                aplicar_estilo(c, fill_fila, FONT_NORMAL, ALIGN_WRAP)

    # ── Fila de advertencia final ──────────────────────────────────────────
    warn_row = 4 + len(BATERIAS) + 1
    ws.row_dimensions[warn_row].height = 20
    ws.merge_cells(f"A{warn_row}:{get_column_letter(len(COLUMNAS))}{warn_row}")
    c = ws.cell(warn_row, 1)
    c.value = ("⚠ Pendiente para TODOS los modelos: DoD (%), Eficiencia RTE (%), "
               "Garantía (años), Temperatura operación, Costo (USD). "
               "Solicitar al proveedor para marcar 'Datos completos = Si'.")
    aplicar_estilo(c, FILL_WARN, FONT_BLACK_B, ALIGN_WRAP)

    # ── Guardar ────────────────────────────────────────────────────────────
    wb.save(EXCEL)
    print(f"✓ Archivo guardado: {EXCEL}")
    print(f"✓ Hoja '{SHEET}' agregada con {len(BATERIAS)} modelos")
    print(f"\nHojas en el Excel ahora: {wb.sheetnames}")
    print("\nEjecute el diagnóstico para confirmar:")
    print("  python bipv_python/datos/diagnostico_catalogo_baterias.py")


if __name__ == "__main__":
    main()
