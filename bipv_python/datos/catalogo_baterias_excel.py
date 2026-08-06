"""
Loader del catálogo de baterías desde el mismo Excel de inversores.
Lee la hoja 'Catalogo_Baterias' (o 'Baterias' / 'Storage' como fallback).
Si la hoja no existe devuelve {} sin error.

Robusto frente a:
  - Headers con saltos de línea (\n) → se normalizan a espacio
  - Título/leyenda en filas 1-3 → se detecta automáticamente la fila de headers
  - Variantes de nombres de columna (~30 alias mapeados)
"""
import pathlib
import pandas as pd
import streamlit as st

_EXCEL  = "/var/www/bipv/calculadora-bipv/bipv_python/datos/inversores_catalogo.xlsx"
_SHEETS = ["Catalogo_Baterias", "Baterias", "Storage"]

# ── Modificación del Excel — usada para invalidar caché automáticamente ──
def excel_mtime() -> float:
    """Timestamp de modificación del Excel; 0.0 si no existe o no es accesible.
    Pasar como _mtime= a cargar_catalogo_baterias() y diagnostico_catalogo()
    garantiza que el caché se invalida al modificar el archivo sin reiniciar PM2.
    """
    try:
        return pathlib.Path(_EXCEL).stat().st_mtime
    except Exception:
        return 0.0


# ── Identificadores que confirman que una fila es el header real ──────────
_MODELO_ALIASES = {"modelo", "nombre", "model", "battery model", "bateria"}

# ── Mapa nombre-de-columna-Excel → clave interna ──────────────────────────
# Cubre variantes con/sin tildes, con/sin paréntesis, con espacios o guión bajo
_COL_MAP = {
    # Identificación
    "Modelo":                      "nombre",
    "modelo":                      "nombre",
    "Nombre":                      "nombre",
    "Battery Model":               "nombre",
    "Bateria":                     "nombre",
    # Fabricante
    "Fabricante":                  "fabricante",
    "Manufacturer":                "fabricante",
    "Marca":                       "fabricante",
    # Datos completos
    "Datos completos (Si/No)":     "_completos",
    "Datos Completos (Si/No)":     "_completos",
    "Datos Completos":             "_completos",
    "Completo":                    "_completos",
    "Complete":                    "_completos",
    # Energía / Capacidad
    "Capacidad (kWh)":             "capacidad_kWh",
    "Capacidad kWh":               "capacidad_kWh",
    "Capacidad_kWh":               "capacidad_kWh",
    "Energía Nominal (kWh)":       "capacidad_kWh",
    "Energia Nominal (kWh)":       "capacidad_kWh",
    "Energy (kWh)":                "capacidad_kWh",
    "Usable Capacity (kWh)":       "capacidad_kWh",
    # Potencia
    "Potencia Continua (kW)":      "potencia_kW",
    "Potencia kW":                 "potencia_kW",
    "Potencia_kW":                 "potencia_kW",
    "Potencia Max (kW)":           "potencia_kW",
    "Continuous Power (kW)":       "potencia_kW",
    "Max Power (kW)":              "potencia_kW",
    # Voltaje
    "Voltaje Nominal (V)":         "voltaje_V",
    "Voltaje V":                   "voltaje_V",
    "Voltaje_V":                   "voltaje_V",
    "Tensión Nominal (V)":         "voltaje_V",
    "Tension Nominal (V)":         "voltaje_V",
    "Nominal Voltage (V)":         "voltaje_V",
    # DoD
    "DoD (%)":                     "dod_pct",
    "DoD Máximo (%)":              "dod_pct",
    "DoD Maximo (%)":              "dod_pct",
    "DoD_pct":                     "dod_pct",
    "Profundidad Descarga (%)":    "dod_pct",
    "Depth of Discharge (%)":      "dod_pct",
    # Ciclos
    "Ciclos de Vida":              "ciclos_vida",
    "Ciclos Vida":                 "ciclos_vida",
    "Ciclos":                      "ciclos_vida",
    "Cycle Life":                  "ciclos_vida",
    "Cycles":                      "ciclos_vida",
    # Eficiencia
    "Eficiencia RTE (%)":          "eta_rte_pct",
    "Eficiencia (%)":              "eta_rte_pct",
    "Eficiencia_rte_pct":          "eta_rte_pct",
    "Rendimiento (%)":             "eta_rte_pct",
    "Round-trip Efficiency (%)":   "eta_rte_pct",
    "RTE (%)":                     "eta_rte_pct",
    # Tipo / Tecnología
    "Tecnología":                  "tipo",
    "Tecnologia":                  "tipo",
    "Tipo":                        "tipo",
    "Química":                     "tipo",
    "Quimica":                     "tipo",
    "Chemistry":                   "tipo",
    # Costo
    "Costo (USD)":                 "costo_usd",
    "Costo USD":                   "costo_usd",
    "Costo Batería":               "costo_usd",
    "Costo Bateria":               "costo_usd",
    "Precio (USD)":                "costo_usd",
    "Price (USD)":                 "costo_usd",
    # Garantía
    "Garantía (años)":             "garantia_anos",
    "Garantia (años)":             "garantia_anos",
    "Garantía (anos)":             "garantia_anos",
    "Garantia (anos)":             "garantia_anos",
    "Warranty (years)":            "garantia_anos",
    # Notas
    "Notas":                       "notas",
    "Observaciones":               "notas",
    "Notes":                       "notas",
}

# Claves numéricas internas
_NUM_KEYS = {"capacidad_kWh", "potencia_kW", "voltaje_V",
             "dod_pct", "ciclos_vida", "eta_rte_pct",
             "costo_usd", "garantia_anos"}

# ── #24 — Aliases canónicos por campo (para mensajes de acción en UI) ─────────
# Lista de nombres de columna sugeridos para agregar al Excel si el campo falta.
_CAMPO_ALIASES_SUGERIDOS: dict[str, list[str]] = {
    "capacidad_kWh": ["Capacidad (kWh)", "Energía Nominal (kWh)", "Energy (kWh)"],
    "potencia_kW":   ["Potencia Continua (kW)", "Potencia Max (kW)", "Continuous Power (kW)"],
    "voltaje_V":     ["Voltaje Nominal (V)", "Tensión Nominal (V)", "Nominal Voltage (V)"],
    "dod_pct":       ["DoD Máximo (%)", "Profundidad Descarga (%)", "Depth of Discharge (%)"],
    "ciclos_vida":   ["Ciclos de Vida", "Cycle Life", "Cycles"],
    "eta_rte_pct":   ["Eficiencia RTE (%)", "Round-trip Efficiency (%)", "RTE (%)"],
    "tipo":          ["Tecnología", "Química", "Chemistry"],
    "costo_usd":     ["Costo (USD)", "Precio (USD)", "Price (USD)"],
    "garantia_anos": ["Garantía (años)", "Warranty (years)"],
    "fabricante":    ["Fabricante", "Manufacturer"],
}

# Campos que bloquean el dimensionamiento si no se encuentran en el Excel
_CAMPOS_CRITICOS   = {"capacidad_kWh", "potencia_kW"}
# Campos que afectan precisión pero tienen defaults seguros
_CAMPOS_IMPORTANTES = {"dod_pct", "eta_rte_pct", "ciclos_vida"}


def _f(val, default=None):
    """Convierte a float; devuelve default para None, NaN o no-numérico."""
    try:
        import math
        v = float(val)
        return default if math.isnan(v) else v
    except Exception:
        return default


def _normalizar_col(nombre: str) -> str:
    """Normaliza un nombre de columna: strip + colapsar saltos de línea a espacio."""
    return " ".join(str(nombre).strip().split())


# ── #24 — Matching de columnas insensible a mayúsculas/tildes/espacios ────────
def _clave_col(nombre: str) -> str:
    """Clave canónica para comparar encabezados: minúsculas, sin tildes,
    espacios colapsados. Así 'CAPACIDAD (KWH)', 'Capacidad  (kWh)' y
    'Capacidad (kwh)' mapean a la misma columna interna."""
    import unicodedata
    s = _normalizar_col(nombre).lower()
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")


# _COL_MAP con claves normalizadas — la búsqueda real usa esta versión
_COL_MAP_NORM = {_clave_col(k): v for k, v in _COL_MAP.items()}


def _mapear_columnas_df(columnas) -> dict:
    """{nombre_columna_df: clave_interna} para las columnas reconocidas.
    Primer alias gana si dos columnas del Excel mapean a la misma interna."""
    mapa, usadas = {}, set()
    for col in columnas:
        interna = _COL_MAP_NORM.get(_clave_col(col))
        if interna and interna not in usadas:
            mapa[col] = interna
            usadas.add(interna)
    return mapa


def _detectar_header(df_raw: pd.DataFrame) -> pd.DataFrame | None:
    """
    Recibe el DataFrame leído con header=0 (fila 0 = títulos del archivo).
    Busca entre las primeras 5 filas cuál contiene 'Modelo' (o alias) en alguna celda.
    Devuelve el DataFrame releyendo con el header correcto, o None si no lo encuentra.
    """
    # Intentar primero con header=0 ya normalizado (#24: sin tildes/mayúsculas)
    cols0 = [_normalizar_col(c) for c in df_raw.columns]
    if any(_clave_col(c) in _MODELO_ALIASES for c in cols0):
        df_raw.columns = cols0
        return df_raw

    # Buscar en filas de datos (el header real está en fila 1, 2, 3…)
    for i, row in df_raw.head(5).iterrows():
        for val in row.values:
            if _clave_col(val) in _MODELO_ALIASES:
                return None  # señal para releer con header=i+1
    return None


@st.cache_data(ttl=3600)
def cargar_catalogo_baterias(_mtime: float = 0.0) -> dict:
    """Devuelve dict {nombre: {...}} con los parámetros de cada batería.

    Robusto frente a formatos de Excel con título en filas superiores.

    Args:
        _mtime: Pasar excel_mtime() para que el caché se invalide automáticamente
                cuando el archivo Excel cambia en disco (p.ej. tras agregar la hoja
                Catalogo_Baterias). Caché TTL de 1 hora como respaldo.
    """
    try:
        xl = pd.ExcelFile(_EXCEL, engine="openpyxl")
    except Exception:
        return {}

    sheet_found = next((s for s in _SHEETS if s in xl.sheet_names), None)
    if sheet_found is None:
        return {}

    # ── Detectar fila de encabezados probando header=0..4 ─────────────────
    df = None
    for h in range(5):
        try:
            df_cand = pd.read_excel(_EXCEL, sheet_name=sheet_found,
                                    header=h, engine="openpyxl")
            cols = [_normalizar_col(c) for c in df_cand.columns]
            if any(_clave_col(c) in _MODELO_ALIASES for c in cols):
                df_cand.columns = cols
                df = df_cand
                break
        except Exception:
            continue

    if df is None:
        return {}

    # ── Parsear filas ──────────────────────────────────────────────────────
    baterias = {}
    # #24 — mapeo por clave normalizada (insensible a mayúsculas/tildes)
    _mapa_df = _mapear_columnas_df(df.columns)
    _col_nombre_df    = next((c for c, i in _mapa_df.items() if i == "nombre"), None)
    _col_completos_df = next((c for c, i in _mapa_df.items() if i == "_completos"), None)
    for _, r in df.iterrows():
        # Buscar nombre del modelo
        nombre = None
        if _col_nombre_df is not None:
            val = str(r.get(_col_nombre_df, "")).strip()
            if val and val.lower() not in ("nan", ""):
                nombre = val
        if not nombre:
            continue
        # Ignorar filas que parezcan ser encabezados o notas
        if _clave_col(nombre) in _MODELO_ALIASES:
            continue
        if nombre.startswith("⚠") or nombre.startswith("*") or len(nombre) > 60:
            continue

        entry = {"nombre": nombre}

        # Mapear todas las columnas reconocidas
        for col_excel, col_int in _mapa_df.items():
            if col_int in ("nombre", "_completos"):
                continue
            raw = r.get(col_excel)
            if col_int in _NUM_KEYS:
                v = _f(raw)
                if v is not None:
                    entry[col_int] = v
            else:
                s = str(raw).strip() if raw is not None and str(raw).strip() != "nan" else None
                if s:
                    entry.setdefault(col_int, s)   # primer alias gana

        # Completitud
        completo_raw = ""
        if _col_completos_df is not None:
            completo_raw = str(r.get(_col_completos_df, "")).strip().lower()
        entry["datos_completos"] = (completo_raw == "si")

        # ── Defaults seguros para cálculo ─────────────────────────────────
        # Solo aplican si el dato NO viene en la ficha. Se registra qué campos
        # fueron rellenados para que el semáforo (#162) no los muestre como
        # datos verificados del fabricante.
        _defaults_aplicados = []
        if not entry.get("dod_pct"):
            entry["dod_pct"] = 80.0      # DoD conservador por defecto
            _defaults_aplicados.append("dod_pct")
        if not entry.get("eta_rte_pct"):
            entry["eta_rte_pct"] = 95.0  # RTE típico LFP
            _defaults_aplicados.append("eta_rte_pct")
        if not entry.get("tipo"):
            entry["tipo"] = "LFP"
        if not entry.get("ciclos_vida"):
            entry["ciclos_vida"] = 3000  # ciclos mínimo conservador
            _defaults_aplicados.append("ciclos_vida")
        entry["_defaults_aplicados"] = _defaults_aplicados

        baterias[nombre] = entry

    return baterias


def obtener_bateria(nombre: str) -> dict:
    return cargar_catalogo_baterias(_mtime=excel_mtime()).get(nombre, {})


def lista_baterias() -> list:
    return sorted(cargar_catalogo_baterias(_mtime=excel_mtime()).keys())


@st.cache_data(ttl=3600)
def diagnostico_catalogo(_mtime: float = 0.0) -> dict:
    """Diagnóstico del catálogo: detecta columnas no reconocidas, modelos incompletos, etc.

    Args:
        _mtime: Pasar excel_mtime() para invalidar caché automáticamente al
                cambiar el archivo. Útil para reflejar cambios sin reiniciar PM2.
    """
    try:
        xl = pd.ExcelFile(_EXCEL, engine="openpyxl")
    except Exception as e:
        return {"error": f"No se pudo abrir el Excel: {e}", "hojas": []}

    info = {"hojas_disponibles": xl.sheet_names}

    sheet_found = next((s for s in _SHEETS if s in xl.sheet_names), None)
    if not sheet_found:
        info["estado"] = f"Hoja no encontrada. Se buscó: {_SHEETS}"
        return info

    info["hoja_usada"] = sheet_found
    cat = cargar_catalogo_baterias(_mtime=_mtime)   # usa la misma entrada de caché
    info["modelos_cargados"] = len(cat)
    info["nombres"] = list(cat.keys())

    campos_criticos = ["capacidad_kWh", "dod_pct", "ciclos_vida", "eta_rte_pct"]
    incompletos = []
    for nombre, b in cat.items():
        falt = [c for c in campos_criticos if not b.get(c)]
        if falt or not b.get("datos_completos"):
            incompletos.append({"modelo": nombre, "campos_faltantes": falt,
                                "datos_completos": b.get("datos_completos")})
    info["modelos_incompletos"] = incompletos

    # Columnas del Excel que no están en _COL_MAP  +  campos sin ningún alias en Excel
    for h in range(5):
        try:
            df_cand = pd.read_excel(_EXCEL, sheet_name=sheet_found,
                                    header=h, engine="openpyxl")
            cols = [_normalizar_col(c) for c in df_cand.columns]
            if any(_clave_col(c) in _MODELO_ALIASES for c in cols):
                # #24 — comparación normalizada (mayúsculas/tildes no cuentan)
                no_mapeadas = [c for c in cols
                               if _clave_col(c) not in _COL_MAP_NORM
                               and _clave_col(c) not in _MODELO_ALIASES
                               and "unnamed" not in c.lower()]
                info["columnas_no_mapeadas"] = no_mapeadas

                # ── #24 — Varias columnas del Excel mapean al mismo campo ────
                # El loader toma la primera (orden de columnas del Excel) y las
                # demás se ignoran en silencio → avisar cuál se está usando.
                _por_interna: dict[str, list] = {}
                for c in cols:
                    _i = _COL_MAP_NORM.get(_clave_col(c))
                    if _i and _i not in ("nombre", "_completos"):
                        _por_interna.setdefault(_i, []).append(c)
                info["columnas_ambiguas"] = [
                    {"campo": k, "columnas": v, "usada": v[0]}
                    for k, v in _por_interna.items() if len(v) > 1
                ]

                # ── #24 — Campos internos cuyo alias NO aparece en el Excel ──────────
                # Distingue "columna existe pero celda vacía" de "columna ausente en total"
                cols_norm_set = {_clave_col(c) for c in cols}
                campos_sin_columna = []
                for campo_int, aliases_sug in _CAMPO_ALIASES_SUGERIDOS.items():
                    # Todos los alias de este campo interno en _COL_MAP (normalizados)
                    aliases_del_campo = [k for k, v in _COL_MAP_NORM.items() if v == campo_int]
                    if not any(a in cols_norm_set for a in aliases_del_campo):
                        campos_sin_columna.append({
                            "campo":              campo_int,
                            "critico":            campo_int in _CAMPOS_CRITICOS,
                            "importante":         campo_int in _CAMPOS_IMPORTANTES,
                            "columnas_sugeridas": aliases_sug,
                        })
                info["campos_sin_columna_excel"] = campos_sin_columna

                # ── #123 — Modelos duplicados en el Excel ────────────────────────
                # El loader usa el nombre como clave del dict: si un modelo aparece
                # dos veces (exacto o con espacios de más), solo sobrevive la última
                # fila y el resto se pierde en silencio.
                col_nombre = next((c for c in df_cand.columns
                                   if _clave_col(c) in _MODELO_ALIASES), None)
                ocurrencias = {}   # nombre_normalizado → {"nombre": ..., "filas": [...]}
                if col_nombre is not None:
                    for idx, raw in df_cand[col_nombre].items():
                        val = str(raw).strip()
                        if not val or val.lower() in ("nan", "") or _clave_col(val) in _MODELO_ALIASES:
                            continue
                        if val.startswith("⚠") or val.startswith("*") or len(val) > 60:
                            continue
                        clave = " ".join(val.split()).lower()   # colapsa espacios, ignora mayúsc.
                        fila_excel = h + 2 + idx                # fila real en el Excel (1-based)
                        d = ocurrencias.setdefault(clave, {"nombre": val, "filas": []})
                        d["filas"].append(int(fila_excel))
                info["modelos_duplicados"] = [
                    {"modelo": d["nombre"], "filas_excel": d["filas"], "n": len(d["filas"])}
                    for d in ocurrencias.values() if len(d["filas"]) > 1
                ]
                break
        except Exception:
            continue

    return info


# ══════════════════════════════════════════════════════════════════════════════
# #163 — Escritura del catálogo desde la app (sin SSH al Excel del servidor)
# ══════════════════════════════════════════════════════════════════════════════

# Encabezados canónicos (los que se usan al crear la hoja o columnas nuevas)
_CANON_COLS: dict[str, str] = {
    "nombre":        "Modelo",
    "fabricante":    "Fabricante",
    "_completos":    "Datos completos (Si/No)",
    "capacidad_kWh": "Capacidad (kWh)",
    "potencia_kW":   "Potencia Continua (kW)",
    "voltaje_V":     "Voltaje Nominal (V)",
    "dod_pct":       "DoD Máximo (%)",
    "ciclos_vida":   "Ciclos de Vida",
    "eta_rte_pct":   "Eficiencia RTE (%)",
    "tipo":          "Tecnología",
    "costo_usd":     "Costo (USD)",
    "garantia_anos": "Garantía (años)",
    "notas":         "Notas",
}


def _abrir_hoja(wb):
    """Devuelve (ws, header_row) de la hoja de baterías, creándola si no existe.

    header_row se detecta buscando en las primeras 5 filas una celda con un
    alias de 'Modelo' (mismo criterio que el loader). Si la hoja se crea,
    se escriben los encabezados canónicos en la fila 1.
    """
    nombre_hoja = next((s for s in _SHEETS if s in wb.sheetnames), None)
    if nombre_hoja is None:
        ws = wb.create_sheet(_SHEETS[0])
        for j, titulo in enumerate(_CANON_COLS.values(), start=1):
            ws.cell(row=1, column=j, value=titulo)
        return ws, 1

    ws = wb[nombre_hoja]
    for fila in range(1, 6):
        for celda in ws[fila]:
            if _clave_col(celda.value or "") in _MODELO_ALIASES:
                return ws, fila
    raise ValueError(
        f"La hoja '{nombre_hoja}' existe pero no se encontró la fila de "
        "encabezados (ninguna celda 'Modelo' en las primeras 5 filas)."
    )


def _mapa_columnas(ws, header_row: int) -> dict:
    """{clave_interna: índice_columna_1based} según los encabezados actuales."""
    mapa = {}
    for celda in ws[header_row]:
        nombre = _normalizar_col(celda.value or "")
        if not nombre:
            continue
        interna = _COL_MAP_NORM.get(_clave_col(nombre))
        if interna and interna not in mapa:
            mapa[interna] = celda.column
    return mapa


def _asegurar_columna(ws, header_row: int, mapa: dict, clave: str) -> int:
    """Devuelve la columna de `clave`, creándola al final si no existe."""
    if clave in mapa:
        return mapa[clave]
    nueva = ws.max_column + 1
    ws.cell(row=header_row, column=nueva, value=_CANON_COLS[clave])
    mapa[clave] = nueva
    return nueva


def _fila_de_modelo(ws, header_row: int, col_nombre: int, nombre: str):
    """Fila (1-based) del modelo, comparando sin mayúsculas ni espacios extra."""
    clave = " ".join(nombre.split()).lower()
    for fila in ws.iter_rows(min_row=header_row + 1):
        val = str(fila[col_nombre - 1].value or "").strip()
        if " ".join(val.split()).lower() == clave:
            return fila[0].row
    return None


def _invalidar_cache():
    try:
        cargar_catalogo_baterias.clear()
        diagnostico_catalogo.clear()
    except Exception:
        pass  # fuera de Streamlit (tests de consola) no hay caché que limpiar


class _LockExcel:
    """Lock advisorio sobre <_EXCEL>.lock para serializar escrituras entre
    sesiones Streamlit simultáneas (evita lost updates en el Excel compartido)."""

    def __enter__(self):
        import fcntl
        self._fh = open(_EXCEL + ".lock", "w")
        fcntl.flock(self._fh, fcntl.LOCK_EX)
        return self

    def __exit__(self, *exc):
        import fcntl
        fcntl.flock(self._fh, fcntl.LOCK_UN)
        self._fh.close()
        return False


def guardar_bateria_excel(datos: dict, nombre_original: str | None = None) -> str:
    """Agrega o actualiza una batería en la hoja Catalogo_Baterias.

    datos: claves internas ("nombre", "capacidad_kWh", "potencia_kW", ...).
           None = dejar la celda vacía (se escribe siempre: es la ficha completa
           tal como quedó en el formulario, no un merge).
    nombre_original: si se está editando y el usuario cambió el nombre del
           modelo, pasar el nombre anterior para actualizar esa fila.

    Retorna el nombre guardado. Lanza ValueError/FileNotFoundError con mensaje
    claro si algo impide escribir (nunca falla en silencio).
    """
    import openpyxl

    nombre = str(datos.get("nombre", "")).strip()
    if not nombre:
        raise ValueError("El nombre del modelo es obligatorio.")
    # El loader descarta en silencio nombres >60 chars — bloquear aquí con
    # mensaje claro en vez de "guardar" algo que luego desaparece.
    if len(nombre) > 60:
        raise ValueError(
            f"El nombre del modelo tiene {len(nombre)} caracteres; el máximo "
            "es 60 (el catálogo lo descartaría al recargar). Acórtalo."
        )

    with _LockExcel():
        try:
            wb = openpyxl.load_workbook(_EXCEL)
        except FileNotFoundError:
            raise FileNotFoundError(f"No se encontró el archivo Excel: {_EXCEL}")

        ws, header_row = _abrir_hoja(wb)
        mapa = _mapa_columnas(ws, header_row)
        col_nombre = _asegurar_columna(ws, header_row, mapa, "nombre")

        _norm = lambda s: " ".join(str(s).split()).lower()

        # Fila destino: la del nombre_original (edición), la del nombre
        # (ya existe) o una nueva al final.
        fila = None
        if nombre_original and nombre_original.strip():
            fila = _fila_de_modelo(ws, header_row, col_nombre, nombre_original)
            # Renombrado: si el nombre NUEVO ya existe en OTRA fila, bloquear
            # (si no, quedarían dos filas y el loader ocultaría una en silencio)
            if fila is not None and _norm(nombre) != _norm(nombre_original):
                fila_choque = _fila_de_modelo(ws, header_row, col_nombre, nombre)
                if fila_choque is not None and fila_choque != fila:
                    raise ValueError(
                        f"Ya existe otra batería llamada '{nombre}' en el "
                        "catálogo. Elige otro nombre o elimina primero la "
                        "existente."
                    )
        if fila is None:
            fila = _fila_de_modelo(ws, header_row, col_nombre, nombre)
        if fila is None:
            fila = ws.max_row + 1

        for clave in _CANON_COLS:
            if clave == "_completos":
                continue  # se escribe abajo según completitud real
            if clave == "nombre":
                ws.cell(row=fila, column=col_nombre, value=nombre)
                continue
            if clave not in datos and clave not in mapa:
                continue  # campo no provisto y sin columna: no crear columnas vacías
            col = _asegurar_columna(ws, header_row, mapa, clave)
            ws.cell(row=fila, column=col, value=datos.get(clave))

        # Completitud: Si cuando los campos que usa el dimensionamiento están llenos
        _claves_completa = ("capacidad_kWh", "potencia_kW", "voltaje_V",
                            "dod_pct", "ciclos_vida", "eta_rte_pct")
        completa = all(datos.get(k) not in (None, "", 0) for k in _claves_completa)
        col_comp = _asegurar_columna(ws, header_row, mapa, "_completos")
        ws.cell(row=fila, column=col_comp, value="Si" if completa else "No")

        wb.save(_EXCEL)
    _invalidar_cache()
    return nombre


def eliminar_bateria_excel(nombre: str) -> bool:
    """Elimina la fila de la batería `nombre`. True si la encontró y borró."""
    import openpyxl

    with _LockExcel():
        try:
            wb = openpyxl.load_workbook(_EXCEL)
        except FileNotFoundError:
            raise FileNotFoundError(f"No se encontró el archivo Excel: {_EXCEL}")

        ws, header_row = _abrir_hoja(wb)
        mapa = _mapa_columnas(ws, header_row)
        if "nombre" not in mapa:
            return False
        fila = _fila_de_modelo(ws, header_row, mapa["nombre"], nombre)
        if fila is None:
            return False
        ws.delete_rows(fila, 1)
        wb.save(_EXCEL)
    _invalidar_cache()
    return True
