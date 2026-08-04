"""
migrar_paneles_respaldo.py — Copia al catálogo actual los paneles que solo
existen en un respaldo del Excel (p. ej. paneles agregados desde la app en el
servidor antes de un git pull que reemplazó el archivo).

Uso (en el servidor):
    cd /var/www/bipv/calculadora-bipv/bipv_python
    venv/bin/python datos/migrar_paneles_respaldo.py /root/paneles_catalogo_backup_2026-08-04.xlsx

- Identifica cada panel por la columna `TipoPanel` (nombre exacto, sin
  mayúsculas/minúsculas ni espacios extra).
- Solo AGREGA los paneles del respaldo que no existen en el catálogo actual;
  nunca modifica ni borra filas existentes.
- Mapea por nombre de columna, así el respaldo puede tener menos columnas que
  el catálogo nuevo (p. ej. sin `BifacialidadPct`): lo que falte queda vacío.
- Antes de escribir crea una copia de seguridad `paneles_catalogo.pre_migracion.xlsx`.
"""
import os
import shutil
import sys

import openpyxl

_SHEET = "Catalogo_Paneles_FV"
_DESTINO = os.path.join(os.path.dirname(os.path.abspath(__file__)), "paneles_catalogo.xlsx")


def _headers(ws) -> dict:
    return {str(c.value).strip(): i + 1 for i, c in enumerate(ws[1]) if c.value is not None}


def _clave(nombre) -> str:
    return " ".join(str(nombre or "").split()).lower()


def main() -> int:
    if len(sys.argv) != 2:
        print("Uso: python datos/migrar_paneles_respaldo.py /ruta/al/respaldo.xlsx")
        return 2
    origen = sys.argv[1]
    for p, rol in ((origen, "respaldo"), (_DESTINO, "catálogo actual")):
        if not os.path.exists(p):
            print(f"❌ No existe el archivo de {rol}: {p}")
            return 2

    wb_src = openpyxl.load_workbook(origen, data_only=True)
    wb_dst = openpyxl.load_workbook(_DESTINO)
    ws_src = wb_src[_SHEET] if _SHEET in wb_src.sheetnames else wb_src.active
    ws_dst = wb_dst[_SHEET] if _SHEET in wb_dst.sheetnames else wb_dst.active

    h_src, h_dst = _headers(ws_src), _headers(ws_dst)
    if "TipoPanel" not in h_src or "TipoPanel" not in h_dst:
        print("❌ Alguno de los archivos no tiene la columna 'TipoPanel' — ¿es el Excel correcto?")
        return 2

    existentes = {
        _clave(ws_dst.cell(row=r, column=h_dst["TipoPanel"]).value)
        for r in range(2, ws_dst.max_row + 1)
    }
    existentes.discard("")

    nuevos, omitidos = [], 0
    for r in range(2, ws_src.max_row + 1):
        nombre = ws_src.cell(row=r, column=h_src["TipoPanel"]).value
        k = _clave(nombre)
        if not k:
            continue
        if k in existentes:
            omitidos += 1
            continue
        fila = {col: ws_src.cell(row=r, column=idx).value for col, idx in h_src.items()}
        nuevos.append((str(nombre).strip(), fila))
        existentes.add(k)

    if not nuevos:
        print(f"✅ Nada que migrar: los {omitidos} paneles del respaldo ya están en el catálogo actual.")
        return 0

    # Copia de seguridad del destino antes de tocarlo
    pre = _DESTINO.replace(".xlsx", ".pre_migracion.xlsx")
    shutil.copyfile(_DESTINO, pre)

    fila_dst = ws_dst.max_row
    cols_ignoradas = sorted(set(h_src) - set(h_dst))
    for nombre, fila in nuevos:
        fila_dst += 1
        for col, idx in h_dst.items():
            if col in fila and fila[col] is not None:
                ws_dst.cell(row=fila_dst, column=idx, value=fila[col])
        print(f"  ➕ {nombre}")

    wb_dst.save(_DESTINO)
    print(f"\n✅ Migrados {len(nuevos)} panel(es) al catálogo actual "
          f"({omitidos} ya existían y se omitieron).")
    if cols_ignoradas:
        print(f"ℹ️  Columnas del respaldo sin equivalente en el catálogo nuevo (ignoradas): {cols_ignoradas}")
    print(f"🗂️  Copia previa del catálogo guardada en: {pre}")
    print("🔁 Reinicia la app para que el catálogo se recargue: pm2 restart streamlit-bipv")
    return 0


if __name__ == "__main__":
    sys.exit(main())
