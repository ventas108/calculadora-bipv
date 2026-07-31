"""
Agrega los 6 modelos JA Solar JAM66D46-LB (Deep Blue 4.0 N-Type Bifacial)
al catálogo de paneles (paneles_catalogo.xlsx).
Fuente: Ficha técnica oficial JAM66D46 LB series (Global-EN-20250709C).
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "venv", "lib", "python3.12", "site-packages"))
import openpyxl

EXCEL = os.path.join(os.path.dirname(__file__), "paneles_catalogo.xlsx")

# ── Datos extraídos de la ficha técnica ──────────────────────────────────────
# Parámetros mecánicos (comunes a los 6 modelos)
# Dimensiones: 2384×1303×33 mm  |  Celdas: 132 físicas (6×22 half-cut)  →  N_s = 66 eléctrico
# NOCT: 45±2°C  |  Tecnología: N-Type TOPCon Bifacial  |  Bifacialidad: 80%±5%
# Coeficientes temperatura (comunes):
#   α_Isc = +0.045 %/°C   β_Voc = -0.250 %/°C   γ_Pmp = -0.290 %/°C

MODELOS = [
    # (Modelo,             Pmax, Voc,   Vmp,   Isc,   Imp,   Eta%)
    ("JAM66D46-715/LB",    715,  48.80, 41.00, 18.55, 17.44, 23.0),
    ("JAM66D46-720/LB",    720,  49.00, 41.19, 18.59, 17.48, 23.2),
    ("JAM66D46-725/LB",    725,  49.20, 41.39, 18.63, 17.52, 23.3),
    ("JAM66D46-730/LB",    730,  49.40, 41.58, 18.67, 17.56, 23.5),
    ("JAM66D46-735/LB",    735,  49.60, 41.77, 18.70, 17.60, 23.7),
    ("JAM66D46-740/LB",    740,  49.80, 41.96, 18.73, 17.64, 23.8),
]

COMUN = {
    "Marca":          "JA Solar",
    "Tecnologia":     "N-Type TOPCon Bifacial",
    "DimensionesMM":  "2384x1303x33 mm",
    "NOCT_C":         45.0,
    "CoefT_C":        -0.290,   # γ_Pmp  %/°C
    "CoefVoc_C":      -0.250,   # β_Voc  %/°C
    "alpha_isc":      0.045,    # α_Isc  %/°C  (informativo)
    "Ns (Celdas Serie)": 66,    # N_s eléctrico (half-cut: 132 físicas / 2 paralelo)
    "TransparenciaPct":  0.0,
    "Confianza":      "Alta — ficha oficial JA Solar Global-EN-20250709C",
    "Notas":          "N-Type n-Bycium+ 18BB Half-Cut Double Glass Bifacial. "
                      "Max sys voltage 1500 VDC. Bifacialidad 80%±5%. "
                      "0.4% degradación anual (30 años). NOCT 45±2°C.",
}

wb = openpyxl.load_workbook(EXCEL)
ws = wb.active

# ── Leer headers de la fila 1 ─────────────────────────────────────────────────
headers = [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value else "" for c in range(1, ws.max_column + 1)]

def col(name):
    """Retorna índice 1-based de la columna por nombre."""
    try:    return headers.index(name) + 1
    except: return None

# ── Detectar modelos ya presentes ────────────────────────────────────────────
existentes = set()
tipo_col = col("TipoPanel")
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[tipo_col - 1]:
        existentes.add(str(r[tipo_col - 1]).strip())

next_row = ws.max_row + 1
agregados = []

for modelo, pmax, voc, vmp, isc, imp, eta in MODELOS:
    nombre_panel = f"JA Solar {modelo}"
    if nombre_panel in existentes:
        print(f"⚠  Ya existe: {nombre_panel}")
        continue

    fila = {}
    fila["TipoPanel"]         = nombre_panel
    fila["PmaxWp"]            = pmax
    fila["Voc_STC"]           = voc
    fila["Vmp_STC"]           = vmp
    fila["Isc_STC"]           = isc
    fila["Imp_STC"]           = imp
    fila["EficienciaPct"]     = eta
    fila.update(COMUN)

    for col_name, valor in fila.items():
        c = col(col_name)
        if c:
            ws.cell(next_row, c).value = valor

    print(f"✓ Agregado fila {next_row}: {nombre_panel} — {pmax} W")
    agregados.append(nombre_panel)
    next_row += 1

wb.save(EXCEL)
print(f"\nTotal agregados: {len(agregados)}")
print("Archivo guardado:", EXCEL)
