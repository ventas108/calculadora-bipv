"""
Herramienta permanente para agregar paneles al catálogo Excel.

── USO EN REPLIT ────────────────────────────────────────────────────────────
    python3 bipv_python/datos/agregar_panel.py --datos datos_panel.json
    # o directamente editando PANEL_DATA al final de este archivo

── USO EN EL SERVIDOR ──────────────────────────────────────────────────────
    cd /var/www/bipv/calculadora-bipv
    source bipv_python/venv/bin/activate
    python bipv_python/datos/agregar_panel.py

── USO PROGRAMÁTICO ────────────────────────────────────────────────────────
    from bipv_python.datos.agregar_panel import agregar_panel_al_catalogo
    agregar_panel_al_catalogo({"TipoPanel": "...", "PmaxWp": 400, ...})

Columnas reconocidas (deben coincidir con fila 1 del Excel):
    TipoPanel, Marca, Tecnologia, PmaxWp, DimensionesMM, CostoUSD,
    NOCT_C, CoefT_C, CoefVoc_C, TransparenciaPct,
    Voc_STC, Vmp_STC, Isc_STC, Imp_STC,
    Ns (Celdas Serie), n (Factor Idealidad), NsA = n × Ns,
    Fuente NsA, Confianza, Notas
"""

import sys
import os
from pathlib import Path

# ── Detectar entorno y resolver ruta del Excel ───────────────────────────────
_THIS_DIR   = Path(__file__).parent
_EXCEL_LOCAL  = _THIS_DIR / "paneles_catalogo.xlsx"
_EXCEL_SERVER = Path("/var/www/bipv/calculadora-bipv/bipv_python/datos/paneles_catalogo.xlsx")
EXCEL = _EXCEL_SERVER if _EXCEL_SERVER.exists() else _EXCEL_LOCAL
SHEET = "Catalogo_Paneles_FV"


def _get_openpyxl():
    """Importa openpyxl buscando en el venv de la app si el entorno no lo tiene."""
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        pass

    # Buscar venv relativo al script (Replit: bipv_python/venv/lib/python3.12/...)
    venv_site = _THIS_DIR.parent / "venv" / "lib"
    if venv_site.exists():
        for py_ver in sorted(venv_site.iterdir(), reverse=True):
            sp = py_ver / "site-packages"
            if sp.exists() and str(sp) not in sys.path:
                sys.path.insert(0, str(sp))
        try:
            import openpyxl
            return openpyxl
        except ImportError:
            pass

    raise ImportError(
        "No se encontró openpyxl.\n"
        "En Replit: el venv debe estar en bipv_python/venv/\n"
        "En servidor: ejecutar con 'source bipv_python/venv/bin/activate'"
    )


def agregar_panel_al_catalogo(datos: dict, excel: Path = EXCEL) -> bool:
    """
    Agrega un panel al catálogo Excel si no existe ya.

    Args:
        datos: dict con claves = nombres de columna del Excel.
               'TipoPanel' es obligatorio y sirve como clave de deduplicación.
        excel: ruta al archivo paneles_catalogo.xlsx.

    Returns:
        True si se agregó, False si ya existía.

    Raises:
        ValueError si falta 'TipoPanel' o si el Excel no existe.
    """
    openpyxl = _get_openpyxl()

    modelo = str(datos.get("TipoPanel", "")).strip()
    if not modelo:
        raise ValueError("El dict debe incluir 'TipoPanel'.")

    if not excel.exists():
        raise FileNotFoundError(f"No se encontró el Excel: {excel}")

    wb = openpyxl.load_workbook(excel)
    if SHEET not in wb.sheetnames:
        raise ValueError(f"La hoja '{SHEET}' no existe en {excel.name}. "
                         f"Hojas disponibles: {wb.sheetnames}")
    ws = wb[SHEET]

    # Leer headers de fila 1 → {nombre_columna: número_columna_1based}
    headers = {
        str(ws.cell(1, c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value is not None
    }

    # Deduplicación por TipoPanel
    col_tipo = headers.get("TipoPanel", 1)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[col_tipo - 1]).strip() == modelo:
            print(f"ℹ️  Ya existe en catálogo: '{modelo}' — sin cambios.")
            return False

    # Escribir nueva fila
    next_row = ws.max_row + 1
    omitidas = []
    for col_name, value in datos.items():
        if col_name in headers:
            ws.cell(row=next_row, column=headers[col_name], value=value)
        else:
            omitidas.append(col_name)

    wb.save(excel)
    print(f"✅  Agregado: '{modelo}' → fila {next_row} de {excel.name}")
    if omitidas:
        print(f"   Columnas no encontradas (omitidas): {omitidas}")
    return True


def listar_paneles(excel: Path = EXCEL) -> list:
    """Devuelve lista de nombres de paneles en el catálogo."""
    openpyxl = _get_openpyxl()
    wb = openpyxl.load_workbook(excel, read_only=True, data_only=True)
    ws = wb[SHEET]
    headers = {str(ws.cell(1, c).value).strip(): c
               for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
    col = headers.get("TipoPanel", 1)
    nombres = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = row[col - 1]
        if v and str(v).strip():
            nombres.append(str(v).strip())
    wb.close()
    return sorted(nombres)


# ── Bloque de ejecución directa ───────────────────────────────────────────────
if __name__ == "__main__":
    import json, argparse

    parser = argparse.ArgumentParser(
        description="Agregar panel al catálogo Excel de la Calculadora BIPV"
    )
    parser.add_argument("--datos",  metavar="archivo.json",
                        help="JSON con los datos del panel (todas las columnas).")
    parser.add_argument("--listar", action="store_true",
                        help="Mostrar paneles actuales en el catálogo.")
    args = parser.parse_args()

    if args.listar:
        print(f"\nPaneles en catálogo ({EXCEL.name}):")
        for n in listar_paneles():
            print(f"  • {n}")
        sys.exit(0)

    if args.datos:
        with open(args.datos, encoding="utf-8") as f:
            panel_dict = json.load(f)
        agregar_panel_al_catalogo(panel_dict)
    else:
        # ── EDITA ESTE BLOQUE para uso directo sin JSON ───────────────────
        print("Uso: python agregar_panel.py --datos panel.json")
        print("     python agregar_panel.py --listar")
        print("\nO edita PANEL_DATA en este script y ejecuta directamente.")
        sys.exit(0)
