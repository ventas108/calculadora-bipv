# -*- coding: utf-8 -*-
"""
Script de uso unico: agrega 4 inversores reales Sungrow (serie comercial
SG5.0/7.0/8.0/10RT, 1000 Vdc) al catalogo real
(datos/inversores_catalogo.xlsx::Catalogo_Inversores), con ficha
MECANICA completa (N Trackers, corriente por tracker) -- a diferencia
del import masivo CEC/Sandia (datos/agregar_inversores_cec_nrel.py, 2.343
modelos SIN estos datos), estos 4 quedan "Datos completos"="Si" y
disponibles de inmediato en el optimizador de Fase 4 y el diagrama
unifilar.

Elegido por presencia real confirmada en Colombia (~1.5 GW ya instalados,
Bemco nombrado nuevo distribuidor oficial para expansion residencial/
comercial/industrial, 25 GW entregados en LatAm -- verificado via
busqueda web el 4-sep-2026, evidencia mas fuerte que otros candidatos
del mismo segmento).

Fuente: ficha oficial real Sungrow, alojada en el propio dominio de
soporte del fabricante:
https://info-support.sungrowpower.com/application/pdf/2022/08/18/
DS_20220818_SG5.0_7.0_8.0_10RT_Datasheet_V11_EN(AU).pdf
(SG5.0/7.0/8.0/10RT -- Multi-MPPT String Inverter for 1000 Vdc System,
version 11, 2022, Sungrow Power Supply Co., Ltd.)

NOTA IMPORTANTE sobre asimetria real de MPPT (no es un dato inventado,
es una limitacion real del esquema del catalogo): para SG7.0RT/8.0RT/
10RT, el MPPT1 real soporta 2 strings/25A y el MPPT2 real soporta 1
string/12.5A (asimetrico) -- la ficha oficial lo confirma
("No. of PV strings per MPPT" = "2/1"). El catalogo solo tiene un campo
por inversor para "N Strings/Tracker"/"Corriente Maxima Tracker", no
uno por MPPT individual. Se usa el valor MAS CONSERVADOR (1 string,
12.5A) para los 2 trackers -- evita que el chequeo automatico de
compatibilidad sobreestime la capacidad real; la asimetria completa
(y la capacidad extra real del MPPT1) queda documentada en Notas para
que un ingeniero la aproveche manualmente si corresponde.

Uso:
    cd bipv_python
    python datos/agregar_inversores_sungrow_ficha_real.py
"""
import sys
from pathlib import Path
import openpyxl

EXCEL = Path(__file__).parent / "inversores_catalogo.xlsx"
SHEET = "Catalogo_Inversores"
HEADER_ROW = 3

FUENTE = (
    "Ficha oficial real Sungrow SG5.0/7.0/8.0/10RT (Multi-MPPT String "
    "Inverter for 1000 Vdc System, V11, 2022) -- "
    "info-support.sungrowpower.com (dominio propio del fabricante). "
    "Marca con presencia real confirmada en Colombia (~1.5 GW ya "
    "instalados, distribuidor oficial Bemco desde 2026 para expansión "
    "residencial/comercial/industrial)."
)

_NOTA_ASIMETRIA = (
    "⚠️ MPPT asimétrico real (confirmado en ficha): MPPT1 soporta 2 "
    "strings/25A, MPPT2 soporta 1 string/12.5A -- 'N Strings/Tracker' y "
    "'Corriente Máxima Tracker' aquí usan el valor MÁS CONSERVADOR (1 "
    "string, 12.5A) para los 2 trackers, para no sobreestimar la "
    "capacidad real en el chequeo automático. La capacidad extra real "
    "del MPPT1 (string adicional, +12.5A) no está representada -- "
    "verificar contra la ficha para aprovecharla en un proyecto real."
)

# Datos extraidos exactos de la ficha oficial (pagina 2, tabla "Type
# designation"). Campos comunes a los 4: Vdc_max=1100V, V_arranque=180V,
# Vmppt_min=160V, Vmppt_max=1000V, N_mppt=2, Vmppt_activo_min=160V
# (mismo valor real, sin divergencia documentada).
MODELOS = [
    {
        "modelo": "SG5.0RT", "P_dc_max_W": 7500.0, "P_ac_nom_kW": 5.0,
        "N_strings_tracker": 1, "I_max_tracker": 12.5, "Isc_max_tracker": 18.0,
        "simetria": "simétrico real (1 string/12.5A en ambos MPPT, sin asimetría).",
    },
    {
        "modelo": "SG7.0RT", "P_dc_max_W": 10500.0, "P_ac_nom_kW": 6.999,
        "N_strings_tracker": 1, "I_max_tracker": 12.5, "Isc_max_tracker": 18.0,
        "simetria": _NOTA_ASIMETRIA,
    },
    {
        "modelo": "SG8.0RT", "P_dc_max_W": 12000.0, "P_ac_nom_kW": 8.0,
        "N_strings_tracker": 1, "I_max_tracker": 12.5, "Isc_max_tracker": 18.0,
        "simetria": _NOTA_ASIMETRIA,
    },
    {
        "modelo": "SG10RT", "P_dc_max_W": 15000.0, "P_ac_nom_kW": 10.0,
        "N_strings_tracker": 1, "I_max_tracker": 12.5, "Isc_max_tracker": 18.0,
        "simetria": _NOTA_ASIMETRIA,
    },
]


def _construir_fila(m: dict) -> dict:
    return {
        "Datos completos (Si/No)": "Si",
        "Modelo": m["modelo"],
        "Costo Inversor ": None,
        "Archivo origen": "Sungrow SG5.0_7.0_8.0_10RT Datasheet V11 (sungrowpower.com)",
        "Tension DC Maxima (V)": 1100.0,
        "Tension Arranque (V)": 180.0,
        "Rango MPPT Min (V)": 160.0,
        "Rango MPPT Max (V)": 1000.0,
        "Tension Minima MPPT Activo (V)": 160.0,
        "N Trackers": 2,
        "N Strings/Tracker": m["N_strings_tracker"],
        "Corriente Maxima Tracker (A)": m["I_max_tracker"],
        "Corriente Cortocircuito Max Tracker (A)": m["Isc_max_tracker"],
        "Potencia FV Max Recomendada (W)": m["P_dc_max_W"],
        "Potencia AC nominal (kW)": m["P_ac_nom_kW"],
        "Marca": "Sungrow",
        "Arquitectura": "String",
        "Inversor Híbrido (Si/No)": "No",
        "Voltaje Batería Min (V)": None,
        "Voltaje Batería Max (V)": None,
        "Corriente Máxima Carga Batería (A)": None,
        "Confianza": "alta (ficha oficial real Sungrow, verificada contra dominio propio del fabricante)",
        "Notas": FUENTE + " " + m["simetria"],
    }


def main():
    if not EXCEL.exists():
        print(f"ERROR: no se encontro {EXCEL}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]
    headers = [str(ws.cell(HEADER_ROW, c).value) if ws.cell(HEADER_ROW, c).value else "" for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    col_modelo = col_idx["Modelo"]
    col_marca = col_idx["Marca"]

    existentes = set()
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        v_modelo = ws.cell(r, col_modelo).value
        v_marca = ws.cell(r, col_marca).value
        if v_modelo:
            existentes.add((str(v_marca or "").strip(), str(v_modelo).strip()))

    agregados, ya_existian = 0, 0
    for m in MODELOS:
        fila = _construir_fila(m)
        clave = (fila["Marca"], fila["Modelo"])
        if clave in existentes:
            ya_existian += 1
            continue

        target_row = ws.max_row + 1
        for col_name, value in fila.items():
            if col_name not in col_idx:
                print(f"  AVISO: columna '{col_name}' no encontrada en el catalogo real (se omite)")
                continue
            ws.cell(row=target_row, column=col_idx[col_name], value=value)
        existentes.add(clave)
        agregados += 1

    wb.save(EXCEL)
    print(f"Agregados: {agregados}")
    print(f"Ya existian (omitidos, sin duplicar): {ya_existian}")
    print(f"Archivo guardado: {EXCEL}")


if __name__ == "__main__":
    main()
