# -*- coding: utf-8 -*-
"""
Script de uso único: agrega el inversor INVT MG750TL al catálogo real
(Catalogo_Inversores) si todavía no existe.

Motivo: es el inversor real usado en la corrida de PVsyst 8.1.5 que validó
el motor CdTe contra un caso real (ver DIAGNOSTICO_JRC_HULD_PRIMARIO_CDTE.md
y DIAGNOSTICO_RECOMBINACION_CDTE.md) -- el usuario no lo encontraba en el
catálogo de la app para reproducir el mismo diseño.

Fuentes de los datos (2 fuentes reales, cruzadas):
1. Ficha oficial del fabricante: "INVT MG-0.75-6kW Single-Phase Grid-tied
   Solar Inverter", rev. 2020.07 V1.0, INVT Solar Technology (Shenzhen) Co.
   -- fila MG750TL de la tabla "Specification".
2. Pantalla real de PVsyst 8.1.5 ("Definición del inversor de red", módulo
   MG750TL (750w), fabricante "INVT Solar technology", fuente de datos
   "Manufacturer 2017") -- capturada durante la corrida real validada.

Las 2 fuentes coinciden en lo esencial (potencia CA nominal 750W, tensión
DC máxima 400V, eficiencia máxima 96,80% EXACTA) pero difieren en la
ventana MPPT (ficha 2020: 50-400V; pantalla PVsyst: 60-350V) -- se usa la
ventana de PVsyst por ser la que realmente se validó en la corrida real,
con la discrepancia documentada en la columna "Notas", no oculta.

Uso:
    cd bipv_python
    python datos/agregar_inversor_invt_mg750tl.py
"""
import sys
from pathlib import Path
import openpyxl

EXCEL = Path(__file__).parent / "inversores_catalogo.xlsx"
SHEET = "Catalogo_Inversores"

INVERSOR = {
    "Modelo":                                    "MG750TL",
    "Datos completos (Si/No)":                   "Si",
    "Costo Inversor ":                           "",
    "Archivo origen":                             "INVT-MG-0.75-6kW_datasheet_2020.07_V1.0.pdf "
                                                    "+ verificado contra pantalla real PVsyst 8.1.5",
    # PV Input -- ventana MPPT de la pantalla real de PVsyst (la que se
    # validó), no la de la ficha 2020 (50-400V) -- ver docstring.
    "Tension DC Maxima (V)":                     400,   # coincide EXACTO en ambas fuentes
    "Tension Arranque (V)":                      60,
    "Rango MPPT Min (V)":                        60,
    "Rango MPPT Max (V)":                        350,
    "Tension Minima MPPT Activo (V)":            60,
    "N Trackers":                                1,
    "N Strings/Tracker":                         1,
    # Corriente máxima de entrada: NINGUNA de las 2 fuentes la publica
    # directamente (PVsyst mostró "N/A" en pantalla). Sin este dato,
    # evaluar_compatibilidad_string() (calculos/dimensionamiento.py) devuelve
    # evaluable=False para CUALQUIER configuración con este inversor --
    # encontrado el 2-sep-2026 corriendo la config real ya validada (3 en
    # serie x 4 strings) y viendo que fallaba por "ficha incompleta", no por
    # incompatibilidad real. Se deriva (no se inventa) del límite físico real
    # ya publicado: Max. DC input power (900W, ficha oficial) / Vmppt_min
    # (60V, ver arriba) = 15A -- el peor caso real de corriente a la potencia
    # máxima del inversor operando en el extremo inferior de su ventana MPPT,
    # no un número sacado de la nada. Documentado como derivado, no medido.
    "Corriente Maxima Tracker (A)":              15.0,
    "Corriente Cortocircuito Max Tracker (A)":   15.0,
    "Potencia FV Max Recomendada (W)":           900,   # real, ficha oficial ("Max. DC input power")
    "Potencia AC nominal (kW)":                  0.75,  # real, ambas fuentes coinciden
    "Marca":                                      "INVT Solar Technology",
    "Notas": (
        "Eficiencia máxima 96,80% (idéntica en ficha oficial y pantalla PVsyst). "
        "Eficiencia Euro: 95,95% (ficha oficial) / 96,00% (PVsyst) -- diferencia menor, no "
        "resuelta. Corriente AC nominal ficha PVsyst: 3,26A (=750W/230V); corriente AC máxima "
        "ficha oficial: 3,6A (=~828W/230V, coherente con 'Potencia CA máxima 0,80kW' vista en "
        "PVsyst). Ventana MPPT de la ficha oficial 2020 es más amplia (50-400V) que la usada en "
        "la corrida real de PVsyst (60-350V) -- se usó esta última por ser la validada. "
        "Corriente máxima de entrada (15A): DERIVADA de 900W/60V (potencia DC máxima real / "
        "Vmppt mínimo real), no publicada directamente por ninguna fuente -- sin este valor, "
        "evaluar_compatibilidad_string() no podía evaluar ninguna configuración con este inversor."
    ),
}


def main():
    if not EXCEL.exists():
        print(f"ERROR: No se encontró el archivo: {EXCEL}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL)
    if SHEET not in wb.sheetnames:
        print(f"ERROR: No existe la hoja '{SHEET}' en {EXCEL.name}")
        sys.exit(1)

    ws = wb[SHEET]
    header_row = 3
    headers = [str(ws.cell(header_row, c).value).strip() if ws.cell(header_row, c).value else ""
               for c in range(1, ws.max_column + 1)]

    col_modelo = headers.index("Modelo") + 1

    modelo = INVERSOR["Modelo"]
    fila_existente = None
    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        val = str(row[col_modelo - 1].value).strip() if row[col_modelo - 1].value else ""
        if val.upper() == modelo.upper():
            fila_existente = row[0].row
            break

    target_row = fila_existente if fila_existente else ws.max_row + 1
    for col_name, value in INVERSOR.items():
        if col_name not in headers:
            print(f"  AVISO: columna '{col_name}' no encontrada en el catálogo real (se omite)")
            continue
        col_idx = headers.index(col_name) + 1
        ws.cell(row=target_row, column=col_idx, value=value if value != "" else None)

    wb.save(EXCEL)
    accion = "Actualizado" if fila_existente else "Agregado"
    print(f"{accion}: {modelo} en la fila {target_row}. Archivo guardado: {EXCEL}")


if __name__ == "__main__":
    main()
