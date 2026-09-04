# -*- coding: utf-8 -*-
"""
Script de uso unico: agrega paneles reales de ZNSHINE PV-TECH (Mono-Si)
al catalogo real (datos/paneles_catalogo.xlsx::Catalogo_Paneles_FV) --
ver DIAGNOSTICO_CATALOGO_ZNSHINE_NREL.md para el detalle completo.

Mismo patron que JA Solar/Trina/Jinko/Canadian Solar/LONGi/Risen:
exclusion por tolerancia SDM (validar_sdm_vs_ficha, 6%). 0/99 excluidos
-- lote mas limpio junto con LONGi.

Elegido por presencia real confirmada en el mercado colombiano (oficina
propia en Bogota, 6 distribuidores locales nombrados: Energitel Solar,
Emergente Energia Sostenible, Ferragro, Solar On Colombia, Eco Green
Solar, Colpilastiendasolar -- verificado via busqueda web el 3-sep-2026)
y fabricante activo (a diferencia de LG Electronics, descartado por el
usuario por riesgo real de garantia/reposicion tras su salida del
negocio de fabricacion de paneles en 2022).

Fuentes (mismas 2 que los imports anteriores):
1. PVS_params_translated.csv (Deville et al. 2025 IEEE JPV, Zenodo
   10.5281/zenodo.14173605) -- params PVsyst-v6.
2. CEC Modules.csv (NREL/SAM) -- NOCT/dimensiones/coef. temp. REPORTADOS.

Uso:
    cd bipv_python
    python datos/agregar_paneles_znshine_nrel.py
"""
import csv
import re
import sys
from pathlib import Path
import openpyxl

EXCEL = Path(__file__).parent / "paneles_catalogo.xlsx"
SHEET = "Catalogo_Paneles_FV"

PATH_PVS = Path(r"C:\Users\Mauricio\Desktop\OPTIMIZADOR PARA CALCULOS BIPV\PRUEBA PVSYST vs MI APP\PVS_params_translated (1).csv")
PATH_CEC = Path(r"C:\Users\Mauricio\AppData\Local\Temp\claude\C--Users-Mauricio\a42e99e2-0c0d-4df5-8dad-a306b9744d70\scratchpad\cec_modules_sam.csv")

TOLERANCIA_ELECTRICA_PCT = 6.0
_TEC_MAP = {"Mono-c-Si": "Mono-Si", "Mono-C-si": "Mono-Si", "Multi-c-Si": "Poli-Si"}


def _norm(s: str) -> str:
    return re.sub(r"[.,]", "", s).strip().lower()


def _cargar_pvs():
    with open(PATH_PVS, encoding="utf-8-sig") as f:
        return {r["module_name"]: r for r in csv.DictReader(f)}


def _cargar_cec():
    with open(PATH_CEC, encoding="utf-8-sig") as f:
        header = next(csv.reader(f))
    filas = {}
    with open(PATH_CEC, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f, fieldnames=header):
            if r["Name"] not in ("Name", ""):
                filas[r["Name"]] = r
    return filas


def _construir_pares_znshine(pvs, cec):
    pvs_norm = {}
    for n in pvs:
        pvs_norm.setdefault(_norm(n), n)
    cec_norm = {}
    for n in cec:
        cec_norm.setdefault(_norm(n), n)
    comunes = set(pvs_norm) & set(cec_norm)
    return [
        (cec_norm[k], pvs_norm[k])
        for k in comunes
        if cec[cec_norm[k]]["Manufacturer"].strip().lower() == "znshine pv-tech co ltd"
    ]


def _auditar_electrico(nombre, c, pv):
    sys.path.insert(0, str(Path(__file__).parent.parent))
    from calculos.modelo_iv import validar_sdm_vs_ficha  # import perezoso

    Voc, Vmp, Isc, Imp = (float(c["V_oc_ref"]), float(c["V_mp_ref"]),
                          float(c["I_sc_ref"]), float(c["I_mp_ref"]))
    Pmax = float(c["STC"])
    Ns = int(c["N_s"])
    n_idealidad = float(pv["gamma_ref"])
    Tk_alfa = float(c["alpha_sc"]) / Isc * 100.0
    tec = _TEC_MAP.get(c["Technology"].strip(), c["Technology"].strip())
    panel = {
        "nombre": nombre, "tecnologia": tec, "N_s": Ns,
        "gamma_ref": n_idealidad, "mu_gamma": float(pv["mu_gamma"]),
        "I_L_ref": float(pv["I_L_ref"]), "I_o_ref": float(pv["I_o_ref"]),
        "R_s": float(pv["R_s"]), "R_sh_ref": float(pv["R_sh_ref"]),
        "R_sh_0": float(pv["R_sh_0"]), "Tk_alfa": Tk_alfa,
        "Voc_stc": Voc, "Isc_stc": Isc, "Vmp_stc": Vmp, "Imp_stc": Imp, "Pmax_stc": Pmax,
    }
    return validar_sdm_vs_ficha(panel, tolerancia_pct=TOLERANCIA_ELECTRICA_PCT)


def _construir_fila(nombre_cec, pv, c):
    modelo = nombre_cec.replace("ZNSHINE PV-TECH Co Ltd ", "").strip()
    Voc, Vmp, Isc, Imp = (float(c["V_oc_ref"]), float(c["V_mp_ref"]),
                          float(c["I_sc_ref"]), float(c["I_mp_ref"]))
    Pmax = float(c["STC"])
    Ns = int(c["N_s"])
    n_idealidad = round(float(pv["gamma_ref"]), 4)
    NsA = round(n_idealidad * Ns, 2)
    tec = _TEC_MAP.get(c["Technology"].strip(), c["Technology"].strip())

    tiene_dim = bool(c["Length"].strip() and c["Width"].strip())
    dims_mm = f"{round(float(c['Length'])*1000)}x{round(float(c['Width'])*1000)}" if tiene_dim else "N/D"

    CoefVoc_C = round(float(c["beta_oc"]) / Voc * 100.0, 4)
    CoefT_C = round(float(c["gamma_pmp"]), 4)
    noct = c["T_NOCT"].strip() or "N/D"

    pmp_nmbe = float(pv["pmp_nmbe"])
    partes = [
        "Fuente: NREL/SAM CEC Modules.csv + Deville et al. 2025 IEEE JPV "
        f"(params PVsyst-v6 traducidos, pmp_nmbe={pmp_nmbe:.3f}%).",
        "Marca con presencia real confirmada en Colombia (oficina propia en "
        "Bogotá; 6 distribuidores nombrados: Energitel Solar, Emergente "
        "Energía Sostenible, Ferragro, Solar On Colombia, Eco Green Solar, "
        "Colpilastiendasolar).",
    ]
    if not tiene_dim:
        partes.append(
            f"Dimensiones NO disponibles en la fuente (solo area A_c={c['A_c']}m2) "
            "-- completar con ficha real antes de usar en chequeo de densidad."
        )
    partes.append(
        f"NOCT tomado de CEC/NREL ({noct}°C) -- confirmar con ficha específica "
        "antes de un proyecto real."
    )

    return {
        "ID": None, "Marca": "ZNSHINE", "TipoPanel": modelo,
        "PmaxWp": Pmax, "DimensionesMM": dims_mm, "CostoUSD": 0,
        "NOCT_C": noct, "CoefT_C": CoefT_C, "TransparenciaPct": 0,
        "Notas": " | ".join(partes), "Tecnologia ": tec,
        "Voc_STC": Voc, "Vmp_STC": Vmp, "Isc_STC": Isc, "Imp_STC": Imp,
        "CoefVoc_C": CoefVoc_C, "Ns (Celdas Serie)": Ns,
        "n (Factor Idealidad)": n_idealidad, "NsA = n × Ns": NsA,
        "Tecnología": tec,
        "Fuente NsA": "NREL/SAM CEC Modules.csv + Sandia JPV 2025 (gamma_ref)",
        "Confianza": "alta (CEC + tolerancia SDM ≤6%)",
        "BifacialidadPct": 0,
    }


def main():
    if not PATH_PVS.exists():
        print(f"ERROR: no se encontro {PATH_PVS} -- descargar de Zenodo (ver docstring)")
        sys.exit(1)
    if not PATH_CEC.exists():
        print(f"ERROR: no se encontro {PATH_CEC} -- descargar de NREL/SAM (ver docstring)")
        sys.exit(1)
    if not EXCEL.exists():
        print(f"ERROR: no se encontro {EXCEL}")
        sys.exit(1)

    pvs, cec = _cargar_pvs(), _cargar_cec()
    pares = _construir_pares_znshine(pvs, cec)
    print(f"Candidatos ZNSHINE (match normalizado): {len(pares)}")

    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]
    # SIN .strip() a proposito -- ver DIAGNOSTICO_CATALOGO_TRINA_NREL.md,
    # bug real del 3-sep-2026: el encabezado real tiene "Tecnologia " con
    # espacio al final; con .strip() aqui esa columna queda sin mapear y se
    # escribe en blanco sin ningun aviso.
    headers = [str(ws.cell(1, c).value) if ws.cell(1, c).value else "" for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    col_modelo = col_idx["TipoPanel"]

    existentes = set()
    last_id = 0
    for r in range(2, ws.max_row + 1):
        v_id = ws.cell(r, col_idx["ID"]).value
        if isinstance(v_id, (int, float)):
            last_id = max(last_id, int(v_id))
        v_modelo = ws.cell(r, col_modelo).value
        if v_modelo:
            existentes.add(str(v_modelo).strip())

    agregados, ya_existian, n_excluidos_electrico = 0, 0, 0
    for nombre_cec, nombre_pv in sorted(pares):
        pv, c = pvs[nombre_pv], cec[nombre_cec]

        val = _auditar_electrico(nombre_cec, c, pv)
        if not val["validacion_ok"]:
            n_excluidos_electrico += 1
            continue

        fila = _construir_fila(nombre_cec, pv, c)
        if fila["TipoPanel"] in existentes:
            ya_existian += 1
            continue

        last_id += 1
        fila["ID"] = last_id
        target_row = ws.max_row + 1
        for col_name, value in fila.items():
            if col_name not in col_idx:
                print(f"  AVISO: columna '{col_name}' no encontrada en el catalogo real (se omite)")
                continue
            ws.cell(row=target_row, column=col_idx[col_name], value=value)
        agregados += 1

    wb.save(EXCEL)
    print(f"Excluidos por tolerancia SDM >6%: {n_excluidos_electrico}")
    print(f"Agregados: {agregados}")
    print(f"Ya existian (omitidos, sin duplicar): {ya_existian}")
    print(f"Archivo guardado: {EXCEL}")


if __name__ == "__main__":
    main()
