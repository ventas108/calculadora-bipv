"""
Test del módulo calculos/export_cotizacion.py

Verifica:
  1. generar_cotizacion_excel(datos) → bytes; abre con openpyxl y contiene el total.
  2. generar_cotizacion_pdf(datos)   → bytes; empieza con %PDF y pesa > 1 KB.
  3. Caso sin ítems activos ⇒ ValueError claro.

Uso:
    /tmp/venv/bin/python bipv_python/scripts/test_export_cotizacion.py
"""
import io
import os
import sys

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from calculos.export_cotizacion import (  # noqa: E402
    generar_cotizacion_excel,
    generar_cotizacion_pdf,
    formato_cop,
    nombre_archivo_cotizacion,
)


def _datos_sinteticos() -> dict:
    trm = 4000.0
    items = [
        # Perfilería
        {"categoria": "Perfilería y Estructura", "descripcion": "Estructura de aluminio BIPV",
         "cantidad": 120.0, "unidad": "m", "unitario_cop": 85000.0, "total_cop": 10_200_000.0},
        {"categoria": "Perfilería y Estructura", "descripcion": "Fijaciones y anclajes",
         "cantidad": 1.0, "unidad": "glb", "unitario_cop": 3_000_000.0, "total_cop": 3_000_000.0},
        # Equipos
        {"categoria": "Equipos principales", "descripcion": "Módulos BIPV (catálogo)",
         "cantidad": 80.0, "unidad": "un", "unitario_cop": 260000.0, "total_cop": 20_800_000.0},
        {"categoria": "Equipos principales", "descripcion": "Inversor string 25 kW",
         "cantidad": 1.0, "unidad": "un", "unitario_cop": 7_400_000.0, "total_cop": 7_400_000.0},
        # Mano de obra
        {"categoria": "Mano de obra", "descripcion": "Instalación, montaje y certificación RETIE",
         "cantidad": 1.0, "unidad": "glb", "unitario_cop": 9_600_000.0, "total_cop": 9_600_000.0},
    ]
    subtotal = sum(it["total_cop"] for it in items)          # 51_000_000
    blandos  = 6_000_000.0
    indirect = 5_000_000.0
    conting  = 4_000_000.0
    total_cop = subtotal + blandos + indirect + conting      # 66_000_000
    return {
        "empresa": "SolTech Energy S.A.S.",
        "proyecto": "Fachada BIPV Torre Empresarial",
        "cliente": "Inmobiliaria Andina",
        "fecha": "15/08/2026",
        "validez_dias": 15,
        "trm": trm,
        "items": items,
        "subtotal_cop": subtotal,
        "costos_blandos_cop": blandos,
        "indirectos_cop": indirect,
        "contingencia_cop": conting,
        "total_cop": total_cop,
        "total_usd": total_cop / trm,
        "notas": "",   # usa notas por defecto
    }


def test_formato_cop():
    assert formato_cop(12345678) == "$ 12.345.678", formato_cop(12345678)
    assert formato_cop(0) == "$ 0"
    assert formato_cop(None) == "$ 0"
    print("✅ formato_cop → separador de miles con punto")


def test_excel():
    from openpyxl import load_workbook
    datos = _datos_sinteticos()
    data = generar_cotizacion_excel(datos)
    assert isinstance(data, (bytes, bytearray)), "El Excel debe devolver bytes"
    assert len(data) > 1024, f"Excel demasiado pequeño: {len(data)} bytes"

    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    # Recolectar todos los valores numéricos y textos de la hoja.
    numeros = set()
    textos = []
    for fila in ws.iter_rows(values_only=True):
        for v in fila:
            if isinstance(v, (int, float)):
                numeros.add(round(float(v)))
            elif isinstance(v, str):
                textos.append(v)
    total_esperado = round(datos["total_cop"])
    assert total_esperado in numeros, (
        f"El total esperado {total_esperado} no está en el Excel. Números: {sorted(numeros)[-8:]}"
    )
    assert any("TOTAL" in t.upper() for t in textos), "Falta la etiqueta TOTAL"
    assert any("Cliente" in t for t in textos), "Falta el dato del cliente"
    assert any("Notas" in t for t in textos), "Faltan las notas/condiciones"
    print(f"✅ Excel OK — {len(data)} bytes, total {formato_cop(total_esperado)} presente")


def test_pdf():
    datos = _datos_sinteticos()
    data = generar_cotizacion_pdf(datos)
    assert isinstance(data, (bytes, bytearray)), "El PDF debe devolver bytes"
    assert data[:4] == b"%PDF", f"El PDF no empieza con %PDF: {data[:8]!r}"
    assert len(data) > 1024, f"PDF demasiado pequeño: {len(data)} bytes"
    print(f"✅ PDF OK — {len(data)} bytes, cabecera {data[:5]!r}")


def test_sin_items():
    datos = _datos_sinteticos()
    datos["items"] = []
    try:
        generar_cotizacion_excel(datos)
        raise AssertionError("Excel: se esperaba ValueError con lista de ítems vacía")
    except ValueError as e:
        assert "activo" in str(e).lower(), f"Mensaje poco claro: {e}"
    try:
        generar_cotizacion_pdf(datos)
        raise AssertionError("PDF: se esperaba ValueError con lista de ítems vacía")
    except ValueError as e:
        assert "activo" in str(e).lower(), f"Mensaje poco claro: {e}"

    # Ítems presentes pero todos en cero (equivalente a inactivos) → también falla.
    datos2 = _datos_sinteticos()
    for it in datos2["items"]:
        it["total_cop"] = 0.0
        it["cantidad"] = 0.0
    try:
        generar_cotizacion_excel(datos2)
        raise AssertionError("Se esperaba ValueError con ítems en cero")
    except ValueError:
        pass
    print("✅ Caso sin ítems activos → ValueError claro")


def test_nombre_archivo():
    n = nombre_archivo_cotizacion("Fachada BIPV Torre Empresarial", "20260815", "pdf")
    assert n == "Cotizacion_Fachada_BIPV_Torre_Empresarial_20260815.pdf", n
    print(f"✅ nombre_archivo → {n}")


if __name__ == "__main__":
    test_formato_cop()
    test_excel()
    test_pdf()
    test_sin_items()
    test_nombre_archivo()
    print("\n🎉 Todos los tests de export_cotizacion pasaron.")
