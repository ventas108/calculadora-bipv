"""
Pruebas del guardado/edición/eliminación de baterías en el Excel (#163).

Usa un Excel temporal (no toca el catálogo real del servidor).

Uso:  python scripts/test_catalogo_baterias_crud.py
"""
import os
import sys
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
import datos.catalogo_baterias_excel as cbe

PASS = FAIL = 0


def check(nombre, cond, detalle=""):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  ✅ {nombre}")
    else:
        FAIL += 1
        print(f"  ❌ {nombre} {detalle}")


def leer_catalogo():
    """Lee el catálogo directamente (sin caché de Streamlit)."""
    return cbe.cargar_catalogo_baterias.__wrapped__(_mtime=cbe.excel_mtime())


tmp = tempfile.mkdtemp()
cbe._EXCEL = os.path.join(tmp, "test_catalogo.xlsx")

print("── CRUD del catálogo de baterías en Excel ──")

# 1. Excel sin hoja de baterías → guardar crea la hoja con encabezados canónicos
wb = openpyxl.Workbook()
wb.active.title = "Inversores"
wb.save(cbe._EXCEL)
cbe.guardar_bateria_excel({"nombre": "BYD HVM 11.0", "fabricante": "BYD",
                           "capacidad_kWh": 11.04, "potencia_kW": 5.0,
                           "voltaje_V": 409.6, "dod_pct": 90, "eta_rte_pct": 96,
                           "ciclos_vida": 4000, "tipo": "LFP",
                           "costo_usd": 4200, "garantia_anos": 10})
cat = leer_catalogo()
check("Crea hoja y guarda batería nueva", "BYD HVM 11.0" in cat, f"({list(cat)})")
b = cat.get("BYD HVM 11.0", {})
check("Valores numéricos correctos",
      b.get("capacidad_kWh") == 11.04 and b.get("voltaje_V") == 409.6)
check("Ficha completa marcada Si", b.get("datos_completos") is True)

# 2. Guardar con el mismo nombre → actualiza la fila, no duplica
cbe.guardar_bateria_excel({"nombre": "BYD HVM 11.0", "capacidad_kWh": 11.04,
                           "potencia_kW": 5.0, "voltaje_V": 409.6,
                           "dod_pct": 90, "eta_rte_pct": 96, "ciclos_vida": 4000,
                           "costo_usd": 4500})
cat = leer_catalogo()
diag = cbe.diagnostico_catalogo.__wrapped__(_mtime=cbe.excel_mtime())
check("Actualiza sin duplicar", cat["BYD HVM 11.0"]["costo_usd"] == 4500
      and not diag.get("modelos_duplicados"), f"({diag.get('modelos_duplicados')})")

# 3. Renombrar (nombre_original) → una sola fila con el nombre nuevo
cbe.guardar_bateria_excel({"nombre": "BYD Battery-Box HVM 11.0",
                           "capacidad_kWh": 11.04, "potencia_kW": 5.0,
                           "voltaje_V": 409.6},
                          nombre_original="BYD HVM 11.0")
cat = leer_catalogo()
check("Renombrar reemplaza la fila",
      "BYD Battery-Box HVM 11.0" in cat and "BYD HVM 11.0" not in cat,
      f"({list(cat)})")

# 4. Campos None quedan vacíos + ficha incompleta → 'No'
cbe.guardar_bateria_excel({"nombre": "Pylontech US3000C", "capacidad_kWh": 3.55,
                           "voltaje_V": 48, "potencia_kW": None, "dod_pct": None})
cat = leer_catalogo()
p = cat["Pylontech US3000C"]
check("Ficha incompleta marcada No", p.get("datos_completos") is False)
check("Campo vacío recibe default del loader (marcado)",
      "dod_pct" in p.get("_defaults_aplicados", []), f"({p.get('_defaults_aplicados')})")

# 5. Hoja con título en fila 1 y headers en fila 2 → escribe en la fila correcta
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "Catalogo_Baterias"
ws2.cell(row=1, column=1, value="CATÁLOGO DE BATERÍAS — actualizado 2026")
for j, h in enumerate(["Modelo", "Capacidad (kWh)", "Potencia Continua (kW)",
                       "Voltaje Nominal (V)"], start=1):
    ws2.cell(row=2, column=j, value=h)
ws2.append(["Huawei LUNA2000-5", 5.0, 2.5, 360])
cbe._EXCEL = os.path.join(tmp, "test_titulo.xlsx")
wb2.save(cbe._EXCEL)
cbe.guardar_bateria_excel({"nombre": "Growatt ARK 2.5H", "capacidad_kWh": 2.56,
                           "potencia_kW": 1.28, "voltaje_V": 51.2})
cat = leer_catalogo()
check("Header en fila 2: conserva el existente y agrega el nuevo",
      "Huawei LUNA2000-5" in cat and "Growatt ARK 2.5H" in cat, f"({list(cat)})")
wb_chk = openpyxl.load_workbook(cbe._EXCEL)
check("No pisó la fila de título",
      "CATÁLOGO" in str(wb_chk["Catalogo_Baterias"].cell(row=1, column=1).value))

# 6. Columna nueva se crea si el header no la tenía (DoD no existía)
cbe.guardar_bateria_excel({"nombre": "Growatt ARK 2.5H", "capacidad_kWh": 2.56,
                           "potencia_kW": 1.28, "voltaje_V": 51.2, "dod_pct": 90})
cat = leer_catalogo()
check("Columna DoD creada al vuelo", cat["Growatt ARK 2.5H"].get("dod_pct") == 90)

# 7. Eliminar existente / inexistente
check("Eliminar existente → True", cbe.eliminar_bateria_excel("Huawei LUNA2000-5"))
cat = leer_catalogo()
check("Fila eliminada del catálogo", "Huawei LUNA2000-5" not in cat)
check("Eliminar inexistente → False", not cbe.eliminar_bateria_excel("No Existe X"))

# 8. Guardar sin nombre → ValueError explícito
try:
    cbe.guardar_bateria_excel({"capacidad_kWh": 5.0})
    check("Sin nombre → ValueError", False)
except ValueError:
    check("Sin nombre → ValueError", True)

# 9a. Renombrar hacia un nombre que YA existe en otra fila → ValueError
cbe.guardar_bateria_excel({"nombre": "Modelo A", "capacidad_kWh": 5})
cbe.guardar_bateria_excel({"nombre": "Modelo B", "capacidad_kWh": 10})
try:
    cbe.guardar_bateria_excel({"nombre": "Modelo B", "capacidad_kWh": 5},
                              nombre_original="Modelo A")
    check("Renombrar a nombre existente → ValueError", False)
except ValueError:
    check("Renombrar a nombre existente → ValueError", True)
cat = leer_catalogo()
check("Sin duplicados tras el intento",
      cat["Modelo A"]["capacidad_kWh"] == 5 and cat["Modelo B"]["capacidad_kWh"] == 10)

# 9b. Nombre >60 caracteres → ValueError (el loader lo descartaría en silencio)
try:
    cbe.guardar_bateria_excel({"nombre": "X" * 61, "capacidad_kWh": 5})
    check("Nombre >60 chars → ValueError", False)
except ValueError:
    check("Nombre >60 chars → ValueError", True)

# 9. Excel inexistente → FileNotFoundError explícito (nunca silencio)
cbe._EXCEL = os.path.join(tmp, "no_existe.xlsx")
try:
    cbe.guardar_bateria_excel({"nombre": "X", "capacidad_kWh": 1})
    check("Excel ausente → FileNotFoundError", False)
except FileNotFoundError:
    check("Excel ausente → FileNotFoundError", True)

print(f"\n{'='*60}\nRESULTADO: {PASS} OK · {FAIL} FALLOS")
sys.exit(1 if FAIL else 0)
